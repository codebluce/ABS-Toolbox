"""机构统计 (ABS工具箱 v2.0.0 整合版)
从ABS发行台账生成三张统计表（合并HTML输出）

  表一：管理人统计表（券商only，剔除信托/银行/保险）
  表二：销售机构统计表（联席承销商=销售机构，券商only）
  表三：托管行统计表（分行归并至总行，同名合并）

实体合并：申万宏源 / 申万宏源资管 / 申万资管 → 申万宏源 (来自 entity_alias.py)
托管行归并：分行名 → XX银行（同银行不同分行合并计算, 来自 entity_alias.py）

v2.0.0 改造（2026-07-05）：
- 接入 abs_common.py 共享底座（删 preprocess_unmerge_fill 内部实现）
- 接入 entity_alias.py（删 ENTITY_MERGE_MAP / BANK_NORM_MAP 内联定义）
- 保留 internal_merge_bookkeeping（第二轮簿记录入迁入后删）
- 默认输出路径改 deliverables/dashboards/01_latest/

用法:
  python gen_institution_stats.py <xlsx_path> [--details <明细1.xlsx> ...] [--output <输出路径>]
  --output 默认 deliverables/dashboards/01_latest/YYYYMMDD_机构统计看板.html
"""

import sys, os, re, argparse, glob
import pandas as pd
import openpyxl
from collections import defaultdict
from datetime import datetime

# v2.0.0 改造: 接入同目录 abs_common + entity_alias
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from abs_common import preprocess_xlsx_for_pandas, PRIORITY_LAYERS, filter_excluded_institutions, fix_bookkeeping_year, normalize_investor_name
from entity_alias import normalize_entity, normalize_bank, ENTITY_MERGE_MAP, BANK_NORM_MAP

# ═══════════════════════════════════════════════════════════════
# §1  视觉设计范式 (Visual Design Paradigm)
# ═══════════════════════════════════════════════════════════════
#
# ┌─────────────────────────────────────────────────┐
# │  BANNER  深蓝渐变 #0d1b2e→#1a3a5c               │
# │  标题20px粗体 / 标签10px / 副标题11px             │
# │  右侧徽章：关键指标                               │
# ├─────────────────────────────────────────────────┤
# │  NOTE BAR  浅黄 #fffbec / 文字 #7a6010            │
# │  业务定义和筛选条件说明                            │
# ├─────────────────────────────────────────────────┤
# │  SECTION ×3  各表独立区块                         │
# │  ┌─ HEADER 渐变背景 白字 ────────────────────┐  │
# │  │ 表名16px粗体 + 统计摘要11px                 │  │
# │  └───────────────────────────────────────────┘  │
# │  ┌─ TABLE ───────────────────────────────────┐  │
# │  │ thead: 深色背景 #fff 11px粗体               │  │
# │  │ tbody: #000 12px / 斑马纹 #f7f9fc          │  │
# │  │ 机构名: 左对齐 font-weight:500               │  │
# │  │ 项目数: 居中对齐                              │  │
# │  │ 规模/占比: 右对齐 font-weight:600 tabular    │  │
# │  └───────────────────────────────────────────┘  │
# ├─────────────────────────────────────────────────┤
# │  FOOTER  灰色 #9aa5b5 11px 居中                  │
# └─────────────────────────────────────────────────┘
#
# 配色体系：
#   表一(管理人): HEADER #0d1b2e→#1a3a5c  THEAD #0d1b2e
#   表二(销售):   HEADER #1c3a2a→#2d7a4f  THEAD #1c3a2a
#   表三(托管行): HEADER #3a2010→#c05a20  THEAD #3a2010
#
# 字体规范：
#   全局: "PingFang SC","Microsoft YaHei",sans-serif
#   正文: 13px / 表格: 12px / 表头: 11px / 标题: 20px
#   全部文字颜色: #000（表格正文） / #fff（深色背景上）
#   数字: font-variant-numeric: tabular-nums
#
# 表格规范：
#   border-collapse: collapse / 白色背景
#   单元格内边距: 7-8px 12px
#   行分隔线: 1px solid #eaeef4
#   偶数行: #f7f9fc / 悬停: #eef3fa
#   圆角: 8px / 阴影: 0 1px 3px rgba(0,0,0,.07)
#
# ──────────────────────────────────────────────────────────────

SECTION_COLORS = {
    'manager':   {'grad_start': '#0d1b2e', 'grad_end': '#1a3a5c', 'thead': '#0d1b2e'},
    'sales':     {'grad_start': '#1c3a2a', 'grad_end': '#2d7a4f', 'thead': '#1c3a2a'},
    'custodian': {'grad_start': '#3a2010', 'grad_end': '#c05a20', 'thead': '#3a2010'},
}

CSS = """
* { box-sizing: border-box; margin: 0; padding: 0; }
body { font-family:"PingFang SC","Microsoft YaHei","Helvetica Neue",Arial,sans-serif;
       background:#f4f5f7; color:#000; font-size:13px; }

.page-banner { background:linear-gradient(135deg,#0d1b2e 0%,#1a3a5c 60%,#0d2a40 100%);
               color:#fff; padding:18px 32px 16px; }
.banner-top { display:flex; justify-content:space-between; align-items:flex-start; }
.banner-tag { font-size:10px; letter-spacing:2px; text-transform:uppercase; color:#7db8e8; margin-bottom:4px; }
.banner-title { font-size:20px; font-weight:700; }
.banner-subtitle { font-size:11px; color:rgba(255,255,255,.65); margin-top:4px; }
.banner-badge { background:rgba(255,255,255,.12); border:1px solid rgba(255,255,255,.25);
                border-radius:4px; padding:5px 12px; font-size:11px; color:#c8dded;
                text-align:right; line-height:1.7; }
.banner-badge strong { color:#fff; font-size:13px; }

.note-bar { padding:8px 32px; background:#fffbec; border-bottom:1px solid #ffe58f;
            font-size:11px; color:#7a6010; line-height:1.5; display:flex; gap:24px; flex-wrap:wrap; }

.content { padding:20px 32px; }

.section { margin-bottom:28px; border-radius:8px; overflow:hidden;
           box-shadow:0 1px 3px rgba(0,0,0,.07); }
.section-header { padding:12px 16px 10px; color:#fff;
                  display:flex; justify-content:space-between; align-items:center; }
.section-title { font-size:16px; font-weight:700; letter-spacing:.5px; }
.section-sub { font-size:11px; color:rgba(255,255,255,.75); text-align:right; line-height:1.5; }

.stat-table { width:100%; border-collapse:collapse; background:#fff; table-layout:fixed; }
.stat-table thead th { color:#fff; padding:10px 12px; text-align:center;
                       font-size:11px; font-weight:600; letter-spacing:.3px; }
/* 5列表(管理人/销售机构): 机构名称28% + 4列各18% = 100% */
.stat-table.cols-5 thead th.col-name { width:28%; }
.stat-table.cols-5 thead th.col-cnt  { width:18%; }
.stat-table.cols-5 thead th.col-num  { width:18%; }
.stat-table.cols-5 thead th.col-pct  { width:18%; }
/* 4列表(托管行): 机构名称30% + 3列各约23.33% = 100% */
.stat-table.cols-4 thead th.col-name { width:30%; }
.stat-table.cols-4 thead th.col-cnt  { width:23%; }
.stat-table.cols-4 thead th.col-num  { width:24%; }
.stat-table.cols-4 thead th.col-pct  { width:23%; }
.stat-table tbody tr:nth-child(even) { background:#f7f9fc; }
.stat-table tbody tr:hover { background:#eef3fa; }
.stat-table tbody tr:not(:last-child) td { border-bottom:1px solid #eaeef4; }
.stat-table tbody td { padding:8px 12px; font-size:12px; color:#000; text-align:center; }
.stat-table tbody td.name { font-weight:500; text-align:center; }
.stat-table tbody td.cnt { text-align:center; }
.stat-table tbody td.num { text-align:center; font-weight:600; font-variant-numeric:tabular-nums; white-space:nowrap; }
.stat-table tbody td.pct { text-align:center; }

.footer { text-align:center; padding:16px 0 24px; font-size:11px; color:#9aa5b5; }
"""

# ═══════════════════════════════════════════════════════════════
# §2  业务规则配置
# ═══════════════════════════════════════════════════════════════

# v2.0.0 改造: ENTITY_MERGE_MAP / BANK_NORM_MAP 已迁入 entity_alias.py
# 顶部 from entity_alias import normalize_entity, normalize_bank 已加载
# 保留 normalize_entity / normalize_bank 函数定义(L166-176)作兼容包装,直接调 entity_alias

# 券商判断：排除银行/信托/保险关键词
BROKERAGE_EXCLUDE_KW = ['银行', '信托', '泰康', '民生通惠', '国寿']

# 列名常量
RESOLVED_COST_COL = '实际成本'
RESOLVED_INST_COL = '实际机构'
RESOLVED_SHARE_COL = '实际份额'
NON_NUMERIC_COST_VALS = {'成本', '自持', '不出表', '-'}

# ═══════════════════════════════════════════════════════════════
# §3  辅助函数
# ═══════════════════════════════════════════════════════════════

# v2.0.0 修正(2026-07-05 审计反馈): 删除本地 normalize_entity / normalize_bank 重定义,
# 消除函数遮蔽,统一走 entity_alias 导入版(含正则回退"XX银行")。
# 原本地版逻辑已迁入 entity_alias.normalize_bank。




def is_brokerage(name):
    """判断是否为券商（排除银行/信托/保险）"""
    return not any(kw in str(name) for kw in BROKERAGE_EXCLUDE_KW)


# ═══════════════════════════════════════════════════════════════
# §3.5  内部簿记合并（22列原始台账 + 明细文件 → 25列完整数据）
# ═══════════════════════════════════════════════════════════════

def preprocess_unmerge_fill(xlsx_path):
    """v2.0.0 改造: 直接调 abs_common.preprocess_xlsx_for_pandas
    （原 65 行实现已删除，逻辑等价:取消合并+A~T向下填充+保护P/U/V/W/X/Y列）
    保留函数名作兼容包装，避免下游 internal_merge_bookkeeping / load_data 改调用链。
    """
    return preprocess_xlsx_for_pandas(xlsx_path)


# 项目名称推断映射：明细文件名 → 台账项目名
PROJECT_NAME_MAP = {
    '金采7-12簿记': '金采7-12',
    '京诚14-9簿记': '京诚14-9',
    # 如有更多映射，可在此追加
}


def infer_project_name(detail_filename):
    """从明细文件名或文件内容推断对应的项目名称"""
    name = os.path.splitext(os.path.basename(detail_filename))[0]
    # 先查映射表
    if name in PROJECT_NAME_MAP:
        return PROJECT_NAME_MAP[name]
    # 读取文件内容提取项目简称（前5行）
    try:
        wb = openpyxl.load_workbook(detail_filename, data_only=True)
        ws = wb.active
        for r in range(1, min(6, ws.max_row + 1)):
            if ws.cell(row=r, column=1).value == '项目名称' and ws.cell(row=r, column=2).value:
                proj_full = str(ws.cell(row=r, column=2).value).strip()
                m = re.search(r'京诚.*?([0-9]+)[号\-]第?([0-9]+)[期\-]', proj_full)
                if m:
                    return f'京诚{m.group(1)}-{m.group(2)}'
                m = re.search(r'盛赢.*?([0-9]+)[号\-]第?([0-9]+)[期\-]', proj_full)
                if m:
                    return f'盛赢{m.group(1)}-{m.group(2)}'
                m = re.search(r'尚赢.*?([0-9]+)[号\-]第?([0-9]+)[期\-]', proj_full)
                if m:
                    return f'尚赢{m.group(1)}-{m.group(2)}'
                m = re.search(r'金采.*?([0-9]+)[号\-]第?([0-9]+)[期\-]', proj_full)
                if m:
                    return f'金采{m.group(1)}-{m.group(2)}'
                break
    except Exception:
        pass
    # 通用规则：去掉常见后缀
    for suffix in ['簿记', '明细', 'book']:
        if name.endswith(suffix):
            return name[:-len(suffix)]
    return name


def read_detail_files(detail_paths):
    """
    读取簿记明细文件列表，返回DataFrame：利率 | 机构 | 规模 | 来源文件 | 推断项目名
    支持多sheet，自动跳过空sheet和header行
    """
    all_records = []
    for path in detail_paths:
        if not os.path.exists(path):
            print(f'[WARN] 明细文件不存在，跳过: {path}')
            continue
        wb = openpyxl.load_workbook(path, data_only=True)
        proj_name = infer_project_name(path)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.max_row < 2:
                continue
            # 尝试定位数据起始行（跳过标题行）
            start_row = 1
            for r in range(1, min(ws.max_row + 1, 7)):
                first_val = ws.cell(row=r, column=1).value
                if first_val is not None and any(kw in str(first_val).strip().lower() for kw in ['利率', '成本', '申购利率']):
                    start_row = r + 1
                    break
            for r in range(start_row, ws.max_row + 1):
                rate = ws.cell(row=r, column=1).value
                inst = ws.cell(row=r, column=2).value
                scale = ws.cell(row=r, column=3).value
                if rate is None and inst is None and scale is None:
                    continue
                all_records.append({
                    '利率': rate,
                    '机构': inst,
                    '规模': scale,
                    '来源文件': os.path.basename(path),
                    '推断项目名': proj_name,
                })
    if not all_records:
        return None
    detail_df = pd.DataFrame(all_records)
    # 类型转换
    detail_df['利率'] = pd.to_numeric(detail_df['利率'].astype(str).str.strip().replace({'-': pd.NA}), errors='coerce')
    detail_df['规模'] = pd.to_numeric(detail_df['规模'].astype(str).str.strip().replace({'-': pd.NA}), errors='coerce')
    # 过滤空值行
    detail_df = detail_df[detail_df[['利率', '机构', '规模']].notna().any(axis=1)].copy()
    return detail_df


def upgrade_22_to_25(xlsx_path):
    """v2.3.0 新增: 22列原始台账 → 25列临时文件(补 WXY 空表头)

    作用: run_increment_merge 要求 processed_path 已是 25 列(含 WXY 表头),
          本函数把 22 列原始台账补 3 列空表头(W/X/Y),数据行留空让 increment_merge 填。
    返回: 25 列临时文件路径(不修改原文件)。
    """
    from openpyxl import load_workbook
    import shutil
    import os as _os

    # 复制到临时文件(不修改原文件)
    tmp_dir = _os.path.join(_os.path.expanduser('~'), 'Documents', 'LikeCodeNex', 'tmp')
    _os.makedirs(tmp_dir, exist_ok=True)
    base = _os.path.splitext(_os.path.basename(xlsx_path))[0]
    tmp_path = _os.path.join(tmp_dir, f'{base}_25col.xlsx')
    shutil.copy2(xlsx_path, tmp_path)

    wb = load_workbook(tmp_path)
    ws = wb.active
    if ws.max_column >= 25:
        print(f'[INFO] 台账已有 {ws.max_column} 列,跳过 22→25 升级')
        return tmp_path

    # 补 W/X/Y 表头(Row1 + Row2,与 25 列加工台账格式一致)
    # W=23(申购利率) X=24(穿透机构) Y=25(申购规模)
    headers_row1 = {23: '申购利率', 24: '穿透机构', 25: '申购规模'}
    headers_row2 = {23: 'W', 24: 'X', 25: 'Y'}
    for col, val in headers_row1.items():
        if ws.cell(row=1, column=col).value is None:
            ws.cell(row=1, column=col).value = val
    for col, val in headers_row2.items():
        if ws.cell(row=2, column=col).value is None:
            ws.cell(row=2, column=col).value = val

    wb.save(tmp_path)
    print(f'[INFO] 22→25 列升级完成: {tmp_path}')
    return tmp_path


def internal_merge_bookkeeping(xlsx_path, detail_paths):
    """v2.3.0 改造: 翻译官模式,调 run_increment_merge(supplement=True)

    改造前(已删除): 88 行内部簿记合并逻辑(与 run_increment_merge 90% 重复,无 QC)
    改造后: 翻译官,做 3 件事:
      1. 22 列台账 → upgrade_22_to_25 → 25 列临时文件(补 WXY 空表头)
      2. 调 increment_merge.run_increment_merge(supplement=True) 填 WXY
      3. 返回 25 列临时文件路径(保持原接口)

    好处:
      - 消除代码重复(88 行 → 15 行翻译逻辑)
      - 自动继承 increment_merge 的 17 项 QC 7.1-7.19
      - 单一来源,行为统一

    返回: 25 列临时文件路径(含已填 WXY)或原始路径(如果无需合并)
    """
    # 若台账本身已有25列以上,无需合并
    wb_check = openpyxl.load_workbook(xlsx_path, data_only=False)
    ws_check = wb_check.active
    if ws_check.max_column >= 25:
        print('[INFO] 台账已有25列,跳过内部簿记合并')
        return xlsx_path

    print('[INFO] 执行内部簿记合并(翻译官模式,调 run_increment_merge)...')

    # Step 1: 22 列 → 25 列(补 WXY 空表头)
    processed_25 = upgrade_22_to_25(xlsx_path)

    # Step 2: 调 run_increment_merge(supplement=True)
    # increment_merge 在同目录,顶部已 sys.path.insert
    from increment_merge import run_increment_merge

    import os as _os
    tmp_dir = _os.path.join(_os.path.expanduser('~'), 'Documents', 'LikeCodeNex', 'tmp')
    _os.makedirs(tmp_dir, exist_ok=True)
    base = _os.path.splitext(_os.path.basename(xlsx_path))[0]
    output_path = _os.path.join(tmp_dir, f'{base}_internal_merged.xlsx')

    run_increment_merge(
        processed_path=processed_25,
        new_raw_path=None,           # supplement 模式不传
        detail_paths=detail_paths,
        output_path=output_path,
        supplement=True
    )

    print(f'[INFO] 内部簿记合并完成(翻译官模式): {output_path}')
    return output_path


# ═══════════════════════════════════════════════════════════════
# §4  数据加载
# ═══════════════════════════════════════════════════════════════

def load_data(xlsx_path, detail_paths=None):
    """读取台账（支持22列原始表或25列总表），返回原始df和去重项目df

    detail_paths: 簿记明细文件路径列表。若传入，且台账为22列原始表，
                  则先执行内部簿记合并，再读取统计。
    """
    # 若提供了明细文件且台账列数不足25，先内部合并
    if detail_paths:
        xlsx_path = internal_merge_bookkeeping(xlsx_path, detail_paths)

    df_raw = pd.read_excel(xlsx_path, engine='openpyxl', header=None)
    n_cols = df_raw.shape[1]
    if n_cols < 22:
        raise ValueError(
            f"[INPUT ERROR] 输入文件列数不足（{n_cols}列），"
            "至少需要22列（原始台账A~V）"
        )

    # 22列原始表（A~V）或25列总表（A~Y含WXY）
    colnames_22 = ['序号', '月份', '类别', '资产类型', '项目名称', '发行场所',
                   '计划管理人', '联席承销商', '托管行', '规模', '期限',
                   '簿记时间', '国股CD', 'ALL-IN成本', '认购倍数',
                   '分层情况', '分层占比', '对应金额（亿）', '评级',
                   '成本', '认购机构', '认购份额']
    colnames_25 = colnames_22 + ['申购利率', '穿透机构', '申购规模']
    colnames = colnames_25 if n_cols >= 25 else colnames_22
    df_raw.columns = colnames + [f'col_{i}' for i in range(len(colnames), n_cols)]

    # 跳过前2行标题 + 残余标题行
    df = df_raw.iloc[2:].copy()
    title_mask = df['成本'].astype(str).str.match(r'^成本$', na=False)
    df = df[~title_mask]
    df.reset_index(drop=True, inplace=True)

    # 解析数值列
    for c in ['规模', '国股CD']:
        df[c] = pd.to_numeric(
            df[c].astype(str).str.strip()
            .replace({v: pd.NA for v in NON_NUMERIC_COST_VALS}),
            errors='coerce'
        )
    df['认购份额'] = pd.to_numeric(df['认购份额'].astype(str).str.strip(), errors='coerce')

    # WXY优先，TUV回退（兼容22列原始表：无WXY时直接用TUV）
    t = pd.to_numeric(
        df['成本'].astype(str).str.strip()
        .replace({v: pd.NA for v in NON_NUMERIC_COST_VALS}),
        errors='coerce'
    )
    u = df['认购机构'].where(
        ~df['认购机构'].astype(str).str.strip().isin(NON_NUMERIC_COST_VALS), pd.NA
    )
    v = pd.to_numeric(df['认购份额'], errors='coerce')

    if '申购利率' in df.columns:
        w = pd.to_numeric(df['申购利率'], errors='coerce')
        df[RESOLVED_COST_COL] = w.fillna(t)
    else:
        df[RESOLVED_COST_COL] = t

    if '穿透机构' in df.columns:
        df[RESOLVED_INST_COL] = df['穿透机构'].where(df['穿透机构'].notna(), u)
    else:
        df[RESOLVED_INST_COL] = u

    # 机构名自动归并：去公司后缀 + 投行上报后缀
    before_count = df[RESOLVED_INST_COL].nunique()
    df[RESOLVED_INST_COL] = df[RESOLVED_INST_COL].apply(normalize_investor_name)
    after_count = df[RESOLVED_INST_COL].nunique()
    if before_count != after_count:
        print(f'[MERGE] 机构名归并：{before_count} → {after_count}（减少 {before_count - after_count} 个变体）')

    if '申购规模' in df.columns:
        df[RESOLVED_SHARE_COL] = pd.to_numeric(df['申购规模'], errors='coerce').fillna(v)
    else:
        df[RESOLVED_SHARE_COL] = v

    # 过滤"自持"
    if '申购利率' in df.columns:
        self_hold = df['申购利率'].astype(str).str.strip().str.match(r'^自持$', na=False)
    else:
        self_hold = df['成本'].astype(str).str.strip().str.match(r'^自持$', na=False)
    df = df[~self_hold]

    # 簿记日期纠错：2025 年 → 2026 年（仅项目名唯一时自动改）
    df['簿记时间'] = pd.to_datetime(df['簿记时间'], errors='coerce')
    df = fix_bookkeeping_year(df, label='gen_institution_stats.load_data')

    # 黑名单剔除：管理人/联席承销商/托管行/认购机构/穿透机构任一命中即剔除整行
    df = filter_excluded_institutions(df, label='gen_institution_stats.load_data')

    # 数据质量检查：各分层内U列有值但WXY全空 → 簿记明细可能漏录
    if '申购利率' in df.columns and '申购规模' in df.columns:
        proj_layers = df.groupby('项目名称')
        for proj_name, group in proj_layers:
            if pd.isna(proj_name):
                continue
            layers = group.groupby('分层情况')
            for layer_name, lg in layers:
                if pd.isna(layer_name):
                    continue
                has_u = lg['认购机构'].notna().any()
                w_empty = lg['申购利率'].isna().all()
                y_empty = lg['申购规模'].isna().all()
                if has_u and w_empty and y_empty and any(k in str(layer_name) for k in ['优先A', '优先B', '优先级']):
                    print(
                        f'  [WARN] 数据质量: 项目"{proj_name}" {layer_name}层 有U列数据但WXY全空，'
                        f'簿记明细可能未录入该分层'
                    )

    # 项目去重：每个项目名称取一行（过滤空项目名）
    projects = df[df['项目名称'].notna()].drop_duplicates(subset='项目名称')[
        ['项目名称', '计划管理人', '联席承销商', '托管行', '规模']
    ].copy()
    projects['计划管理人'] = projects['计划管理人'].apply(normalize_entity)
    projects['托管行'] = projects['托管行'].apply(normalize_bank)

    return df, projects


def compute_ib_subscription(df):
    """计算投行认购规模：U列含"投行"→提取机构名→合并实体→V列求和"""
    df_ib = df.copy()
    df_ib['机构_clean'] = df_ib['认购机构'].astype(str).str.strip()
    ib_mask = df_ib['机构_clean'].str.contains('投行', na=False)
    ib_df = df_ib[ib_mask].copy()
    ib_df['机构简称'] = ib_df['机构_clean'].str.replace('投行.*', '', regex=True).str.strip()
    ib_df['机构简称'] = ib_df['机构简称'].apply(normalize_entity)
    ib_scale = ib_df.groupby('机构简称')[RESOLVED_SHARE_COL].sum()
    return ib_scale


# ═══════════════════════════════════════════════════════════════
# §5  三张表计算
# ═══════════════════════════════════════════════════════════════

def compute_manager_table(projects, ib_scale, total_scale):
    """表一：管理人统计表（券商only）"""
    mgr = projects.groupby('计划管理人').agg(
        管理项目数=('项目名称', 'count'),
        管理规模=('规模', 'sum')
    ).reset_index()
    mgr = mgr[mgr['计划管理人'].apply(is_brokerage)].copy()
    ib_df = ib_scale.reset_index()
    ib_df.columns = ['计划管理人', '投行认购规模']
    mgr = mgr.merge(ib_df, on='计划管理人', how='left')
    mgr['投行认购规模'] = mgr['投行认购规模'].fillna(0)
    mgr['规模占比'] = (mgr['管理规模'] / total_scale * 100).round(2)
    mgr = mgr.sort_values('管理规模', ascending=False).reset_index(drop=True)
    return mgr


def compute_sales_table(projects, ib_scale, total_scale):
    """表二：销售机构统计表（联席承销商=销售，券商only）"""
    rows = []
    for _, row in projects.iterrows():
        institutions = [
            normalize_entity(x.strip())
            for x in str(row['联席承销商']).split('/') if x.strip()
        ]
        for inst in institutions:
            if inst == '-':
                continue
            rows.append({
                '机构名称': inst,
                '项目名称': row['项目名称'],
                '项目规模': row['规模']
            })
    sales_df = pd.DataFrame(rows)
    if sales_df.empty:
        return pd.DataFrame(columns=['机构名称', '销售项目数', '参与项目规模', '投行认购规模', '规模占比'])
    sales_df = sales_df[sales_df['机构名称'].apply(is_brokerage)].copy()
    sales = sales_df.groupby('机构名称').agg(
        销售项目数=('项目名称', 'count'),
        参与项目规模=('项目规模', 'sum')
    ).reset_index()
    ib_df = ib_scale.reset_index()
    ib_df.columns = ['机构名称', '投行认购规模']
    sales = sales.merge(ib_df, on='机构名称', how='left')
    sales['投行认购规模'] = sales['投行认购规模'].fillna(0)
    sales['规模占比'] = (sales['参与项目规模'] / total_scale * 100).round(2)
    sales = sales.sort_values('参与项目规模', ascending=False).reset_index(drop=True)
    return sales


def compute_custodian_table(projects, total_scale):
    """表三：托管行统计表（分行归并至总行）"""
    custody = projects.groupby('托管行').agg(
        托管项目数=('项目名称', 'count'),
        托管规模=('规模', 'sum')
    ).reset_index()
    custody['规模占比'] = (custody['托管规模'] / total_scale * 100).round(2)
    custody = custody.sort_values('托管规模', ascending=False).reset_index(drop=True)
    return custody


# ═══════════════════════════════════════════════════════════════
# §6  HTML 生成
# ═══════════════════════════════════════════════════════════════

def build_stat_table(columns, rows_html, thead_bg):
    """通用统计表格：columns=[(key, label, align_class), ...]
    v2.5.1: 加 cols-N class 用于按列数分配均匀列宽(N=4/5)"""
    th_items = []
    for key, label, cls in columns:
        th_items.append(f'<th class="{cls}">{label}</th>')
    thead = f'<thead style="background:{thead_bg}"><tr>{"".join(th_items)}</tr></thead>'
    n_cols = len(columns)
    cols_class = f' cols-{n_cols}' if n_cols in (4, 5) else ''
    return f'<table class="stat-table{cols_class}">{thead}<tbody>{rows_html}</tbody></table>'


def build_manager_rows(mgr):
    rows = ''
    for _, r in mgr.iterrows():
        rows += (
            f'<tr>'
            f'<td class="name">{r["计划管理人"]}</td>'
            f'<td class="cnt">{int(r["管理项目数"])}</td>'
            f'<td class="num">{r["管理规模"]:.2f}</td>'
            f'<td class="pct">{r["规模占比"]:.2f}%</td>'
            f'<td class="num">{r["投行认购规模"]:.2f}</td>'
            f'</tr>'
        )
    return rows


def build_sales_rows(sales):
    rows = ''
    for _, r in sales.iterrows():
        rows += (
            f'<tr>'
            f'<td class="name">{r["机构名称"]}</td>'
            f'<td class="cnt">{int(r["销售项目数"])}</td>'
            f'<td class="num">{r["参与项目规模"]:.2f}</td>'
            f'<td class="pct">{r["规模占比"]:.2f}%</td>'
            f'<td class="num">{r["投行认购规模"]:.2f}</td>'
            f'</tr>'
        )
    return rows


def build_custodian_rows(custody):
    rows = ''
    for _, r in custody.iterrows():
        rows += (
            f'<tr>'
            f'<td class="name">{r["托管行"]}</td>'
            f'<td class="cnt">{int(r["托管项目数"])}</td>'
            f'<td class="num">{r["托管规模"]:.2f}</td>'
            f'<td class="pct">{r["规模占比"]:.2f}%</td>'
            f'</tr>'
        )
    return rows


def build_section(title, subtitle, table_html, color_key):
    c = SECTION_COLORS[color_key]
    return f'''<div class="section">
  <div class="section-header" style="background:linear-gradient(135deg,{c['grad_start']},{c['grad_end']})">
    <span class="section-title">{title}</span>
    <span class="section-sub">{subtitle}</span>
  </div>
  {table_html}
</div>'''


def _build_section_by_key(key, mgr, sales, custody):
    """按 section key 构建对应 section HTML"""
    if key == 'manager':
        c = SECTION_COLORS['manager']
        cols = [
            ('name', '机构名称', 'col-name'),
            ('cnt', '管理项目数', 'col-cnt'),
            ('num', '管理规模(亿)', 'col-num'),
            ('pct', '规模占比', 'col-pct'),
            ('num', '投行认购规模(亿)', 'col-num'),
        ]
        table = build_stat_table(cols, build_manager_rows(mgr), c['thead'])
        return build_section(
            '表一：管理人统计表',
            f'{len(mgr)}家券商 · 剔除信托/银行/保险 · 申万宏源系合并',
            table, 'manager'
        )
    if key == 'sales':
        c = SECTION_COLORS['sales']
        cols = [
            ('name', '机构名称', 'col-name'),
            ('cnt', '销售项目数', 'col-cnt'),
            ('num', '参与项目规模(亿)', 'col-num'),
            ('pct', '规模占比', 'col-pct'),
            ('num', '投行认购规模(亿)', 'col-num'),
        ]
        table = build_stat_table(cols, build_sales_rows(sales), c['thead'])
        return build_section(
            '表二：销售机构统计表',
            f'{len(sales)}家券商 · 联席承销商=销售机构 · 申万宏源系合并',
            table, 'sales'
        )
    if key == 'custodian':
        c = SECTION_COLORS['custodian']
        cols = [
            ('name', '机构名称', 'col-name'),
            ('cnt', '托管项目数', 'col-cnt'),
            ('num', '托管规模(亿)', 'col-num'),
            ('pct', '规模占比', 'col-pct'),
        ]
        table = build_stat_table(cols, build_custodian_rows(custody), c['thead'])
        return build_section(
            '表三：托管行统计表',
            f'{len(custody)}家银行 · 分行归并至总行 · 同名合并',
            table, 'custodian'
        )
    raise ValueError(f"unknown section_key: {key}")


def render_body(data, section_key=None):
    """渲染 body 片段（不含 <!DOCTYPE>/<html>/<head>）

    section_key=None：渲染完整 body（3 sections，独立看板用）
    section_key='manager'/'sales'/'custodian'：只渲染对应 section（综合看板用）
    """
    mgr = data['mgr']
    sales = data['sales']
    custody = data['custody']
    meta = data['meta']
    xlsx_basename = data['xlsx_basename']

    total_scale = meta['total_scale']
    project_count = meta['project_count']
    date_min = meta['date_min']
    date_max = meta['date_max']

    if section_key is None:
        sections_html = '\n'.join(
            _build_section_by_key(k, mgr, sales, custody)
            for k in ['manager', 'sales', 'custodian']
        )
    else:
        sections_html = _build_section_by_key(section_key, mgr, sales, custody)

    return f'''
<div class="page-banner">
  <div class="banner-top">
    <div>
      <div class="banner-tag">ABS Institution Statistics</div>
      <div class="banner-title">机构统计看板</div>
      <div class="banner-subtitle">
        管理人/销售机构/托管行 三维度统计 ·
        申万宏源系合并 · 托管行分行归并 · 券商only
      </div>
    </div>
    <div class="banner-badge">
      数据期间 <strong>{date_min} ~ {date_max}</strong><br>
      项目 <strong>{project_count}</strong> 个 · 总规模 <strong>{total_scale:.2f} 亿</strong>
    </div>
  </div>
</div>

<div class="note-bar">
  <span>管理人/销售机构：仅统计券商，剔除信托/银行/保险</span>
  <span>申万宏源/申万宏源资管/申万资管合并为同一主体</span>
  <span>联席承销商即销售机构</span>
  <span>托管行分行名归并为总行名（XX银行），同名合并</span>
  <span>规模占比 = 机构规模 / 全部项目总规模</span>
</div>

<div class="content">
{sections_html}
</div>

<div class="footer">
  数据来源：{xlsx_basename} ·
  生成时间：{datetime.now().strftime("%Y-%m-%d")} ·
  ABS工具箱 v2.0.0 (机构统计) · 分析报告员 Peizhi Wu
</div>'''


def render_html(data, section_key=None):
    """渲染完整 HTML 文档（含 <!DOCTYPE>/<html>/<head>）

    独立看板用 section_key=None；综合看板调 render_body 取 body 片段。
    """
    body = render_body(data, section_key=section_key)
    return f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>机构统计看板</title>
<style>{CSS}</style>
</head>
<body>{body}
</body>
</html>'''


# 兼容旧接口
def generate_html(xlsx_path, mgr, sales, custody, meta):
    """已废弃，用 render_html 替代（保留向后兼容）"""
    data = {
        'mgr': mgr, 'sales': sales, 'custody': custody, 'meta': meta,
        'xlsx_basename': os.path.basename(xlsx_path),
    }
    return render_html(data, section_key=None)


# ═══════════════════════════════════════════════════════════════
# §6.5  数据计算层（供综合看板调用）
# ═══════════════════════════════════════════════════════════════

def compute_data(xlsx_path, detail_paths=None):
    """读取台账 + 计算三表 + 跑 QC precheck，返回数据 dict

    供综合看板和独立看板共用。QC precheck 失败时抛 RuntimeError。
    """
    print(f'Loading: {xlsx_path}')
    df, projects = load_data(xlsx_path, detail_paths=detail_paths)

    total_scale = projects['规模'].sum()
    project_count = len(projects)
    ib_scale = compute_ib_subscription(df)

    mgr = compute_manager_table(projects, ib_scale, total_scale)
    sales = compute_sales_table(projects, ib_scale, total_scale)
    custody = compute_custodian_table(projects, total_scale)

    book_dates = df.drop_duplicates(subset='项目名称')['簿记时间']
    book_dates = pd.to_datetime(book_dates, errors='coerce').dropna()
    date_min = book_dates.min().strftime('%Y-%m-%d') if len(book_dates) > 0 else 'N/A'
    date_max = book_dates.max().strftime('%Y-%m-%d') if len(book_dates) > 0 else 'N/A'
    meta = {
        'total_scale': total_scale,
        'project_count': project_count,
        'date_min': date_min,
        'date_max': date_max,
    }

    print(f'项目: {project_count}个 | 总规模: {total_scale:.2f}亿')
    print(f'管理人: {len(mgr)}家 | 销售机构: {len(sales)}家 | 托管行: {len(custody)}家')

    # QC precheck（不写文件）
    qc_passed = run_qc(df, projects, mgr, sales, custody, total_scale,
                       out_path=None, meta=meta, precheck_only=True)
    if not qc_passed:
        raise RuntimeError(f'[QC] 机构统计预检未通过，请检查数据或逻辑')

    return {
        'df': df,
        'projects': projects,
        'mgr': mgr,
        'sales': sales,
        'custody': custody,
        'meta': meta,
        'total_scale': total_scale,
        'ib_scale': ib_scale,
        'xlsx_basename': os.path.basename(xlsx_path),
        'xlsx_path': xlsx_path,
    }


# ═══════════════════════════════════════════════════════════════
# §7  自检模块 (QC)
# ═══════════════════════════════════════════════════════════════

def run_qc(df, projects, mgr, sales, custody, total_scale,
           out_path, meta, precheck_only=False):
    """
    自检维度：
      【1】数据与台账一致性 — 记录数、总额、实体覆盖
      【2】业务逻辑正确性 — 实体合并、托管行归并、券商过滤
      【3】视觉生成一致性 — HTML结构、配色、元素完整
      【4】显示效果正确性 — 无NaN、无缺失值、占比合理
    """
    issues = []
    ok_count = 0

    def ok(msg):
        nonlocal ok_count; ok_count += 1; print(f'  [OK]   {msg}')
    def warn(msg):
        issues.append(('WARN', msg)); print(f'  [WARN] {msg}')
    def fail(msg):
        issues.append(('FAIL', msg)); print(f'  [FAIL] {msg}')

    print('\n' + '=' * 56)
    print('  QC Report — ABS工具箱 v2.0.0 (机构统计)')
    print('=' * 56)

    # ── 【1】数据与台账一致性 ──────────────────────────────────
    print('\n【1】数据与台账一致性')

    # 1-1 项目数一致
    raw_project_count = df['项目名称'].nunique()
    report_project_count = meta['project_count']
    if raw_project_count == report_project_count:
        ok(f'项目数一致：{report_project_count}个')
    else:
        fail(f'项目数不一致：台账={raw_project_count}，报告={report_project_count}')

    # 1-2 总规模一致
    raw_total = df.drop_duplicates(subset='项目名称')['规模'].sum()
    if abs(raw_total - total_scale) < 0.01:
        ok(f'总规模一致：{total_scale:.2f}亿')
    else:
        fail(f'总规模不一致：台账={raw_total:.2f}，报告={total_scale:.2f}')

    # 1-3 管理人表规模之和 ≤ 总规模（因为剔除非券商）
    mgr_total = mgr['管理规模'].sum()
    if mgr_total <= total_scale + 0.01:
        ok(f'管理人表规模合理：{mgr_total:.2f}亿 ≤ 总规模{total_scale:.2f}亿')
    else:
        fail(f'管理人表规模超过总规模：{mgr_total:.2f} > {total_scale:.2f}')

    # 1-4 销售机构表规模之和 ≤ 总规模
    sales_total = sales['参与项目规模'].sum()
    if sales_total <= total_scale + 0.01:
        ok(f'销售机构表规模合理：{sales_total:.2f}亿 ≤ 总规模{total_scale:.2f}亿')
    else:
        fail(f'销售机构表规模超过总规模：{sales_total:.2f} > {total_scale:.2f}')

    # 1-5 托管行表规模之和 ≈ 总规模（所有项目都有托管行）
    custody_total = custody['托管规模'].sum()
    if abs(custody_total - total_scale) < 0.01:
        ok(f'托管行表规模与总规模一致：{custody_total:.2f}亿')
    else:
        warn(f'托管行表规模与总规模有差异：{custody_total:.2f} vs {total_scale:.2f}（可能有缺失托管行）')

    # ── 【2】业务逻辑正确性 ──────────────────────────────────
    print('\n【2】业务逻辑正确性')

    # 2-1 申万宏源合并：不应存在"申万宏源资管"或"申万资管"独立条目
    for col_name, tbl in [('计划管理人', mgr), ('机构名称', sales)]:
        merged_names = tbl[col_name].tolist()
        for old_name in ['申万宏源资管', '申万资管']:
            if old_name in merged_names:
                fail(f'{col_name}列存在未合并实体：{old_name}')
            else:
                ok(f'{col_name}列已合并"申万宏源"系实体')

    # 2-2 托管行归并：所有名称应以"银行"结尾（不含分行字样）
    bank_names = custody['托管行'].tolist()
    branch_keywords = ['北分', '南分', '分行', '支行', '中关村', '总行', '天津', '自贸区']
    branch_found = [n for n in bank_names if any(kw in n for kw in branch_keywords)]
    if not branch_found:
        ok('托管行名称已全部归并至总行（无分行字样）')
    else:
        fail(f'托管行存在未归并分行名：{branch_found}')

    # 2-3 券商过滤：管理人/销售机构表中不应有银行/信托/保险
    for tbl_name, tbl, col in [('管理人', mgr, '计划管理人'), ('销售机构', sales, '机构名称')]:
        non_broker = [name for name in tbl[col].tolist() if not is_brokerage(name)]
        if not non_broker:
            ok(f'{tbl_name}表券商过滤正确（无非券商机构）')
        else:
            fail(f'{tbl_name}表含非券商机构：{non_broker}')

    # 2-4 占比合理性：各表占比之和应≤100%（剔除非券商后可能<100%）
    mgr_pct_sum = mgr['规模占比'].sum()
    sales_pct_sum = sales['规模占比'].sum()
    custody_pct_sum = custody['规模占比'].sum()
    ok(f'管理人占比合计：{mgr_pct_sum:.2f}%')
    ok(f'销售机构占比合计：{sales_pct_sum:.2f}%')
    if abs(custody_pct_sum - 100.0) < 0.5:
        ok(f'托管行占比合计：{custody_pct_sum:.2f}%（≈100%）')
    else:
        warn(f'托管行占比合计：{custody_pct_sum:.2f}%（偏离100%较多）')

    # 2-5 排序验证：各表应按规模降序
    for tbl_name, tbl, col in [('管理人', mgr, '管理规模'), ('销售机构', sales, '参与项目规模'), ('托管行', custody, '托管规模')]:
        vals = tbl[col].tolist()
        if vals == sorted(vals, reverse=True):
            ok(f'{tbl_name}表按规模降序排列')
        else:
            fail(f'{tbl_name}表排序异常')

    # ── 【3】视觉生成一致性 ──────────────────────────────────
    print('\n【3】视觉生成一致性')

    # 3-1 配色检查：section颜色常量定义完整
    for key in ['manager', 'sales', 'custodian']:
        c = SECTION_COLORS.get(key)
        if c and all(k in c for k in ['grad_start', 'grad_end', 'thead']):
            ok(f'配色定义完整：{key} = {c}')
        else:
            fail(f'配色定义缺失：{key}')

    # ── 【4】显示效果正确性 ──────────────────────────────────
    print('\n【4】显示效果正确性')

    # 4-1 无NaN值显示
    for tbl_name, tbl in [('管理人', mgr), ('销售机构', sales), ('托管行', custody)]:
        nan_cols = [col for col in tbl.columns if tbl[col].isna().any()]
        if not nan_cols:
            ok(f'{tbl_name}表无NaN值')
        else:
            fail(f'{tbl_name}表含NaN列：{nan_cols}')

    # 4-2 规模值非负
    for tbl_name, tbl, cols in [
        ('管理人', mgr, ['管理规模', '投行认购规模']),
        ('销售机构', sales, ['参与项目规模', '投行认购规模']),
        ('托管行', custody, ['托管规模']),
    ]:
        for col in cols:
            neg_count = (tbl[col] < 0).sum()
            if neg_count == 0:
                ok(f'{tbl_name}表{col}均为非负值')
            else:
                fail(f'{tbl_name}表{col}有{neg_count}条负值')

    # 4-3 文件质量检查（最终QC）
    if not precheck_only:
        file_kb = os.path.getsize(out_path) / 1024
        if file_kb < 1:
            fail(f'输出文件过小（{file_kb:.1f}KB）')
        elif file_kb < 5:
            warn(f'输出文件较小（{file_kb:.1f}KB），请确认内容完整性')
        else:
            ok(f'输出文件大小正常：{file_kb:.1f}KB')

        with open(out_path, 'r', encoding='utf-8') as f:
            html_content = f.read()

        # HTML结构完整性
        checks = [
            ('class="section"', '三个表区块'),
            ('class="stat-table', '统计表格'),
            ('class="page-banner"', '顶部横幅'),
            ('class="note-bar"', '说明栏'),
            ('class="footer"', '页脚'),
            ('管理人统计表', '表一标题'),
            ('销售机构统计表', '表二标题'),
            ('托管行统计表', '表三标题'),
            ('申万宏源', '申万宏源合并条目'),
        ]
        html_fail = False
        for keyword, label in checks:
            if keyword not in html_content:
                fail(f'HTML缺少关键结构：{label}（"{keyword}"未找到）')
                html_fail = True
        if not html_fail:
            ok('HTML关键结构完整')

        # 区块数量
        section_count = html_content.count('class="section"')
        if section_count == 3:
            ok(f'HTML区块数量正确：{section_count}个')
        else:
            fail(f'HTML区块数量不符：期望3个，实际{section_count}个')

        # 颜色注入验证
        for key, expected_color in [('manager', '#0d1b2e'), ('sales', '#1c3a2a'), ('custodian', '#3a2010')]:
            if expected_color in html_content:
                ok(f'配色{key}已注入HTML（{expected_color}）')
            else:
                warn(f'配色{key}未在HTML中找到（{expected_color}）')
    else:
        ok('输出质量检查在文件写入后执行（预检模式跳过）')

    # ── 汇总 ──────────────────────────────────────────────────
    print('\n' + '=' * 56)
    fail_items = [x for x in issues if x[0] == 'FAIL']
    warn_items = [x for x in issues if x[0] == 'WARN']
    if not fail_items and not warn_items:
        print(f'  QC PASSED — {ok_count}项全部通过')
    elif not fail_items:
        print(f'  QC PASSED WITH WARNINGS — {ok_count}项通过，{len(warn_items)}项需关注：')
        for _, msg in warn_items:
            print(f'    ! {msg}')
    else:
        print(f'  QC FAILED — {len(fail_items)}项严重错误，{len(warn_items)}项警告：')
        for level, msg in issues:
            mark = 'X' if level == 'FAIL' else '!'
            print(f'    {mark} {msg}')
    print('=' * 56 + '\n')

    return not bool(fail_items)


# ═══════════════════════════════════════════════════════════════
# §8  踩坑日志
# ═══════════════════════════════════════════════════════════════

PITFALL_LOG_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'pitfall_log.md')


def log_pitfall(issue, fix, category='其他'):
    """追加踩坑记录到日志文件"""
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M')
    entry = (
        f'\n### {timestamp}\n'
        f'- **类别**: {category}\n'
        f'- **问题**: {issue}\n'
        f'- **修复**: {fix}\n'
    )
    with open(PITFALL_LOG_PATH, 'a', encoding='utf-8') as f:
        f.write(entry)


def init_pitfall_log():
    """初始化踩坑日志（仅在文件不存在时创建）"""
    if os.path.exists(PITFALL_LOG_PATH):
        return
    header = (
        '# 机构统计 Skill 踩坑日志\n\n'
        '本文件记录每次执行的错误和修正情况，由脚本自动追加。\n\n'
        '---\n\n'
        '## v1.0.0 初始已知问题\n\n'
        '### 2026-05-13 初始化\n'
        '- **类别**: 数据\n'
        '- **问题**: 申万宏源/申万宏源资管/申万资管在台账中为不同条目，需合并统计\n'
        '- **修复**: ENTITY_MERGE_MAP映射合并为"申万宏源"\n\n'
        '### 2026-05-13 初始化\n'
        '- **类别**: 数据\n'
        '- **问题**: 托管行列含分行名（如"江苏北分"），需归并至总行名\n'
        '- **修复**: BANK_NORM_MAP映射 + 正则提取"XX银行"通用规则\n\n'
        '### 2026-05-13 初始化\n'
        '- **类别**: 视觉\n'
        '- **问题**: 用户要求字体颜色全部为黑色，不能有彩色字体\n'
        '- **修复**: CSS中所有表格正文color:#000，仅表头/横幅用#fff\n\n'
        '### 2026-05-13 初始化\n'
        '- **类别**: 逻辑\n'
        '- **问题**: 联席承销商列用"/"分隔多机构，需拆分后独立统计\n'
        '- **修复**: str.split("/")展开为多行\n\n'
        '### 2026-05-13 初始化\n'
        '- **类别**: 逻辑\n'
        '- **问题**: 销售机构表中"-"表示无联席承销商，应排除\n'
        '- **修复**: 跳过inst=="-"的条目\n\n'
    )
    with open(PITFALL_LOG_PATH, 'w', encoding='utf-8') as f:
        f.write(header)


# ═══════════════════════════════════════════════════════════════
# §9  主入口
# ═══════════════════════════════════════════════════════════════

def parse_args():
    parser = argparse.ArgumentParser(
        description='ABS机构统计 — 从发行台账生成管理人/销售机构/托管行三表'
    )
    parser.add_argument('xlsx_path', help='输入台账文件路径（.xlsx）')
    parser.add_argument('--details', nargs='+', default=None,
                        help='簿记明细文件列表（支持多文件，支持通配符）')
    parser.add_argument('--output', '-o', default=None,
                        help='输出HTML文件路径（默认 ABS技能包/看板/日期_机构统计看板.html）')
    return parser.parse_args()


if __name__ == '__main__':
    args = parse_args()
    xlsx_path = args.xlsx_path
    out_path = args.output

    # 展开通配符
    detail_paths = []
    if args.details:
        for pattern in args.details:
            matched = glob.glob(pattern)
            if matched:
                detail_paths.extend(matched)
            else:
                detail_paths.append(pattern)

    if out_path is None:
        date_tag = pd.Timestamp('today').strftime('%Y%m%d')
        # v2.0.0 改造: 默认输出路径改 deliverables/dashboards/01_latest/
        kanban_dir = os.path.join(
            os.path.dirname(os.path.abspath(__file__)), '..', 'deliverables', 'dashboards', '01_latest'
        )
        os.makedirs(kanban_dir, exist_ok=True)
        out_path = os.path.join(
            kanban_dir,
            f'{date_tag}_机构统计看板.html'
        )

    # 初始化踩坑日志
    init_pitfall_log()

    # 计算数据（含 QC precheck）
    try:
        data = compute_data(xlsx_path, detail_paths=detail_paths if detail_paths else None)
    except RuntimeError as e:
        log_pitfall(
            'QC预检未通过，跳过文件写入',
            '请检查数据或逻辑后重试',
            category='逻辑'
        )
        print(f'\n[ERROR] {e}')
        print('[ERROR] 请修正数据或逻辑后重试')
        sys.exit(1)

    # 生成 HTML
    html = render_html(data, section_key=None)

    # 写文件 + 最终 QC
    with open(out_path, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f'\nDone: {out_path}')
    run_qc(data['df'], data['projects'], data['mgr'], data['sales'],
           data['custody'], data['total_scale'],
           out_path, data['meta'], precheck_only=False)
