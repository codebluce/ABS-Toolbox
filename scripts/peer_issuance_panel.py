"""《互联网金融ABS内部信息共享》同业发行动态正式面板生成器。"""

from __future__ import annotations

import argparse
import hashlib
import html
import json
import re
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT_DIR = ROOT / "deliverables" / "dashboards" / "02_history" / "peer_issuance"
DEFAULT_BASELINE_PATH = ROOT / "deliverables" / "dashboards" / "04_reference" / "2025年互联网金融ABS同业发行明细.xlsx"
ZERO = Decimal("0")
HUNDRED = Decimal("100")

# 资产类型固定展示顺序
ASSET_TYPE_ORDER = ("消金分期类", "消金提现类", "小微toB", "小微toC")
TOP_N = 12
TRUST_TOP_N = 5
BAR_BLUE = "#2a78d6"
ASSET_FAMILY_ORDER = ("蚂蚁系", "网商系", "腾讯系", "微众系", "字节系", "美团系")
UNKNOWN_ASSET_FAMILY = "未知资产"
ASSET_FAMILY_COLORS = {
    "蚂蚁系": "#2a78d6",
    "网商系": "#4e9f8b",
    "腾讯系": "#7c6eb0",
    "微众系": "#c98747",
    "字节系": "#cf6b6b",
    "美团系": "#6f8e45",
    UNKNOWN_ASSET_FAMILY: "#98a1ad",
}
ASSET_FAMILY_KEYWORDS = {
    "蚂蚁系": ("花呗", "借呗"),
    "网商系": ("网商",),
    "腾讯系": ("腾讯", "分付"),
    "微众系": ("微众", "微粒贷"),
    "字节系": ("抖音", "放心借", "字节"),
    "美团系": ("美团", "月付", "生活费"),
}


class PanelInputError(ValueError):
    """无法可靠生成面板时抛出。"""


# ---------------------------------------------------------------------------
# 输入解析
# ---------------------------------------------------------------------------

def header_key(raw: Any) -> str | None:
    """把表头单元格映射为语义列名（兼容 2025 无场所列、2026 有场所列）。"""
    text = str(raw or "").replace("\n", "").replace(" ", "").strip()
    if not text:
        return None
    if text.startswith("簿记日期"):
        return "week"
    if text.startswith("产品名称"):
        return "product"
    if text.startswith("产品类型"):
        return "product_type"
    if text.startswith("场所"):
        return "venue"
    if text.startswith("原始权益人"):
        return "originator"
    if text.startswith("基础资产"):
        return "base_asset"
    if text.startswith("资产类型"):
        return "asset_type"
    if text.startswith("规模"):
        return "amount"
    if text.startswith("簿记时间"):
        return "bookkeep_time"
    if text.startswith("设立时间"):
        return "establish_time"
    if text.startswith("AAA证券票面利率"):
        return "aaa_rate"
    if text.startswith("AAA档证券期限"):
        return "aaa_term"
    return None


REQUIRED_COLUMNS = ("week", "product", "originator", "base_asset", "asset_type",
                    "amount", "bookkeep_time", "aaa_rate", "aaa_term")


def first_number(value: Any) -> Decimal | None:
    """取多档文本（如“1.89%、1.93%”“0.91、1.19”）的首个数值。"""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    match = re.match(r"\s*(\d+(?:\.\d+)?)", str(value))
    return Decimal(match.group(1)) if match else None


def parse_rate(value: Any) -> Decimal | None:
    """票面利率统一为小数形式：0.0175 或 “1.89%、1.93%” → 0.0189。"""
    number = first_number(value)
    if number is None:
        return None
    if isinstance(value, str) and "%" in value:
        return number / HUNDRED
    if number > Decimal("0.5"):  # 防御：裸写的 1.89 视为百分数
        return number / HUNDRED
    return number


def parse_term(value: Any) -> Decimal | None:
    return first_number(value)


def as_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                pass
    return None


def issue(items: list[dict], level: str, code: str, message: str, **detail: Any) -> None:
    items.append({"level": level, "code": code, "message": message, **detail})


def input_fingerprint(path: Path, rows: int) -> dict[str, Any]:
    """记录正式面板输入指纹，便于周度结果回溯。"""
    digest = hashlib.sha256(path.read_bytes()).hexdigest()
    workbook = load_workbook(path, read_only=True)
    return {"file": path.name, "sheet": workbook.worksheets[0].title, "rows": rows, "sha256": digest}


def parse_dynamics(path: Path) -> list[dict[str, Any]]:
    """解析发行动态 Sheet，返回记录列表（周标签前向填充）。"""
    if not path.exists():
        raise PanelInputError(f"输入文件不存在：{path}")
    workbook = load_workbook(path, data_only=True, read_only=False)
    sheet = workbook.worksheets[0]
    columns = {}
    for col in range(1, sheet.max_column + 1):
        key = header_key(sheet.cell(1, col).value)
        if key and key not in columns:
            columns[key] = col
    missing = [name for name in REQUIRED_COLUMNS if name not in columns]
    if missing:
        raise PanelInputError(f"{path.name} 缺少必需列：{missing}")

    records: list[dict[str, Any]] = []
    current_week = None
    for row in range(2, sheet.max_row + 1):
        cell = lambda key: sheet.cell(row, columns[key]).value  # noqa: E731
        week_raw = str(cell("week") or "").replace("\n", "").strip()
        if week_raw:
            current_week = week_raw
        product = str(cell("product") or "").strip()
        if not product:
            continue
        amount_raw = cell("amount")
        try:
            amount = Decimal(str(amount_raw).replace(",", "").strip()) if amount_raw is not None else ZERO
        except InvalidOperation as exc:
            raise PanelInputError(f"{path.name} 第{row}行规模非法：{amount_raw!r}") from exc
        records.append({
            "week": current_week,
            "product": product,
            "product_type": str(cell("product_type") or "").strip(),
            "venue": str(cell("venue") or "").strip() if "venue" in columns else "",
            "originator": str(cell("originator") or "").strip(),
            "base_asset": str(cell("base_asset") or "").strip(),
            "asset_type": str(cell("asset_type") or "").strip(),
            "amount": amount,
            "date": as_date(cell("bookkeep_time")),
            "aaa": parse_rate(cell("aaa_rate")),
            "term": parse_term(cell("aaa_term")),
        })
    if not records:
        raise PanelInputError(f"{path.name} 未读取到有效发行动态记录")
    return records


# ---------------------------------------------------------------------------
# 周更增量与漂移校验
# ---------------------------------------------------------------------------

DRIFT_FIELDS = ("product_type", "venue", "originator", "base_asset", "asset_type", "amount", "date", "aaa", "term")


def record_key(record: dict[str, Any]) -> str:
    """当前发行动态中产品名称唯一；复合信息保留在快照中供诊断。"""
    return record["product"]


def record_fingerprint(record: dict[str, Any]) -> str:
    """排除周标签展示字段后的业务事实指纹。"""
    payload = {field: str(record.get(field) or "") for field in DRIFT_FIELDS}
    return hashlib.sha256(json.dumps(payload, ensure_ascii=False, sort_keys=True).encode("utf-8")).hexdigest()


def compare_weekly_snapshot(previous: list[dict[str, Any]], current: list[dict[str, Any]]) -> tuple[dict[str, Any], list[dict]]:
    """比较两份追加式快照，识别周新增、历史回补、删除与实质修订。"""
    qc: list[dict] = []
    previous_by_key = {record_key(record): record for record in previous}
    current_by_key = {record_key(record): record for record in current}
    if len(previous_by_key) != len(previous):
        issue(qc, "FAIL", "BASELINE_DUPLICATE_KEY", "基线快照存在重复产品名称")
    if len(current_by_key) != len(current):
        issue(qc, "FAIL", "CURRENT_DUPLICATE_KEY", "当前快照存在重复产品名称")

    previous_keys, current_keys = set(previous_by_key), set(current_by_key)
    additions = [current_by_key[key] for key in sorted(current_keys - previous_keys)]
    deletions = [previous_by_key[key] for key in sorted(previous_keys - current_keys)]
    modifications = []
    for key in sorted(previous_keys & current_keys):
        before, after = previous_by_key[key], current_by_key[key]
        if record_fingerprint(before) != record_fingerprint(after):
            changed_fields = [field for field in DRIFT_FIELDS if before.get(field) != after.get(field)]
            modifications.append({"product": key, "changed_fields": changed_fields})

    current_dates = [record["date"] for record in current if record["date"]]
    latest_date = max(current_dates) if current_dates else None
    # Excel 常按最新周到历史周倒序，但快照/测试可有任意顺序；以周标签中的起始日期确定本周。
    def week_start(week: Any) -> date:
        match = re.search(r"(20\d{2})\.(\d{1,2})\.(\d{1,2})", str(week or ""))
        return date(int(match.group(1)), int(match.group(2)), int(match.group(3))) if match else date.min
    current_week = max((record["week"] for record in current if record["week"]), key=week_start, default=None)
    weekly_additions = [record for record in additions if record.get("week") == current_week]
    backfills = [record for record in additions if record.get("week") != current_week]
    added_amount = sum((record["amount"] for record in additions), ZERO)
    weekly_amount = sum((record["amount"] for record in weekly_additions), ZERO)
    backfill_amount = sum((record["amount"] for record in backfills), ZERO)

    if deletions:
        issue(qc, "FAIL", "HISTORICAL_DELETION", f"检测到 {len(deletions)} 条历史产品删除", products=[record["product"] for record in deletions])
    if modifications:
        issue(qc, "FAIL", "HISTORICAL_REVISION", f"检测到 {len(modifications)} 条历史产品业务字段修订", changes=modifications)
    if backfills:
        issue(qc, "WARN", "HISTORICAL_BACKFILL", f"检测到 {len(backfills)} 条历史周回补", amount=str(backfill_amount), products=[record["product"] for record in backfills])
    if additions:
        issue(qc, "INFO", "WEEKLY_DELTA", f"本次新增 {len(additions)} 条 / {added_amount} 亿，其中当周 {len(weekly_additions)} 条 / {weekly_amount} 亿")

    previous_amount = sum((record["amount"] for record in previous), ZERO)
    current_amount = sum((record["amount"] for record in current), ZERO)
    difference = current_amount - previous_amount
    reconciliation_level = "OK" if difference == added_amount else "FAIL"
    issue(qc, reconciliation_level, "DELTA_RECONCILIATION", f"累计规模变动 {difference} 亿，新增规模 {added_amount} 亿", difference=str(difference), additions=str(added_amount))

    summary = {
        "latest_date": latest_date,
        "current_week": current_week,
        "previous_count": len(previous),
        "current_count": len(current),
        "additions": additions,
        "deletions": deletions,
        "modifications": modifications,
        "weekly_additions": weekly_additions,
        "backfills": backfills,
        "amounts": {"previous": previous_amount, "current": current_amount, "difference": difference, "additions": added_amount, "weekly": weekly_amount, "backfills": backfill_amount},
    }
    return summary, qc


# ---------------------------------------------------------------------------
# 聚合
# ---------------------------------------------------------------------------

def is_jd(record: dict[str, Any]) -> bool:
    """京东系 = 原始权益人或基础资产含“京东”（覆盖走信托通道发行的京东资产）。"""
    return "京东" in record["originator"] or "京东" in record["base_asset"]


def asset_family(base_asset: str) -> str:
    """按基础资产名称归并展示所用的资产集团，未命中归入未知资产。"""
    normalized = str(base_asset or "").strip()
    for family in ASSET_FAMILY_ORDER:
        if any(keyword in normalized for keyword in ASSET_FAMILY_KEYWORDS[family]):
            return family
    return UNKNOWN_ASSET_FAMILY


def cluster_by_family(selected: list[tuple[str, Decimal]]) -> list[str]:
    """同集团基础资产相邻展示的稳定聚类排序。

    入参为按 2026 累计规模降序选出的 Top N [(资产名, 规模), ...]。
    规则：集团间按其内部最高规模降序（未知资产集团沉底），集团内按规模降序；
    返回仅含资产名的展示顺序，入围集合与入参完全一致。
    """
    groups: dict[str, list[tuple[str, Decimal]]] = {}
    for name, amount in selected:
        groups.setdefault(asset_family(name), []).append((name, amount))
    for family in groups:
        groups[family].sort(key=lambda item: -item[1])

    def group_rank(family: str):
        head = groups[family][0][1]
        return (1 if family == UNKNOWN_ASSET_FAMILY else 0, -head, family)

    families = sorted(groups.keys(), key=group_rank)
    return [name for family in families for name, _ in groups[family]]


def is_trust_channel(record: dict[str, Any]) -> bool:
    """信托渠道分布仅纳入名称含“信托”的原始权益人。"""
    return "信托" in record["originator"]


def yoy_cutoff(latest: date) -> date:
    """同比窗口截止日 = 2026 最新簿记日前一年。"""
    try:
        return latest.replace(year=latest.year - 1)
    except ValueError:  # 2 月 29 日
        return latest.replace(year=latest.year - 1, day=28)


def weighted_avg(pairs: list[tuple[Decimal, Decimal]]) -> Decimal | None:
    """按规模加权均值；pairs = [(rate, amount), ...]。"""
    valid = [(rate, amount) for rate, amount in pairs if rate is not None and amount > ZERO]
    if not valid:
        return None
    total = sum((amount for _, amount in valid), ZERO)
    return sum((rate * amount for rate, amount in valid), ZERO) / total


def month_key(record: dict[str, Any]) -> int | None:
    return record["date"].month if record["date"] else None


def build_dashboard(records_2026: list[dict[str, Any]], records_2025: list[dict[str, Any]]) -> tuple[dict[str, Any], list[dict]]:
    qc: list[dict] = []
    dated_2026 = [r for r in records_2026 if r["date"]]
    if not dated_2026:
        raise PanelInputError("2026 数据缺少簿记时间，无法确定最新簿记日")
    latest_date = max(r["date"] for r in dated_2026)
    cutoff = yoy_cutoff(latest_date)

    market_2026 = [r for r in records_2026 if not is_jd(r)]
    jd_2026 = [r for r in records_2026 if is_jd(r)]
    market_2025 = [r for r in records_2025 if not is_jd(r) and r["date"] and r["date"] <= cutoff]
    jd_2025 = [r for r in records_2025 if is_jd(r)]

    issue(qc, "OK", "PARSE_2026", f"2026 源解析 {len(records_2026)} 期", rows=len(records_2026))
    issue(qc, "OK", "PARSE_2025", f"2025 源解析 {len(records_2025)} 期", rows=len(records_2025))
    issue(qc, "OK", "JD_EXCLUSION",
          f"剔除京东系 2026 年 {len(jd_2026)} 期 / {sum((r['amount'] for r in jd_2026), ZERO):.0f} 亿，"
          f"2025 年 {len(jd_2025)} 期",
          count_2026=len(jd_2026), count_2025=len(jd_2025))
    if not market_2025:
        issue(qc, "WARN", "YOY_WINDOW_EMPTY", f"2025 同期窗口（≤ {cutoff.isoformat()}）无可用数据")
    else:
        issue(qc, "OK", "YOY_WINDOW", f"2025 同期窗口 {len(market_2025)} 期", cutoff=cutoff.isoformat())

    # 最新簿记周 = 最新簿记日所在周标签
    latest_week = next(r["week"] for r in dated_2026 if r["date"] == latest_date)
    week_records = [r for r in market_2026 if r["week"] == latest_week]
    rate_pairs = [(r["aaa"], r["amount"]) for r in week_records]
    term_pairs = [(r["term"], r["amount"]) for r in week_records]
    aaa_wavg = weighted_avg(rate_pairs)
    term_wavg = weighted_avg(term_pairs)
    aaa_values = sorted(r["aaa"] for r in week_records if r["aaa"] is not None)
    term_values = sorted(r["term"] for r in week_records if r["term"] is not None)
    if aaa_wavg is None:
        issue(qc, "WARN", "LATEST_WEEK_RATE_MISSING", f"最新簿记周（{latest_week}）无有效 AAA 利率")
    if term_wavg is None:
        issue(qc, "WARN", "LATEST_WEEK_TERM_MISSING", f"最新簿记周（{latest_week}）无有效 AAA 期限")

    total_amount = sum((r["amount"] for r in market_2026), ZERO)
    prev_amount = sum((r["amount"] for r in market_2025), ZERO)
    yoy_pct = None if prev_amount == ZERO else (total_amount - prev_amount) / prev_amount * HUNDRED

    family_names = (*ASSET_FAMILY_ORDER, UNKNOWN_ASSET_FAMILY)
    by_family = []
    for name in family_names:
        amount_2026 = sum((r["amount"] for r in market_2026 if asset_family(r["base_asset"]) == name), ZERO)
        amount_2025 = sum((r["amount"] for r in market_2025 if asset_family(r["base_asset"]) == name), ZERO)
        pct = None if amount_2025 == ZERO else (amount_2026 - amount_2025) / amount_2025 * HUNDRED
        by_family.append({"name": name, "amount_2026": amount_2026, "amount_2025": amount_2025, "pct": pct})
    by_family.sort(key=lambda item: (item["name"] == UNKNOWN_ASSET_FAMILY, -item["amount_2026"], item["name"]))
    family_difference = sum((item["amount_2026"] for item in by_family), ZERO) - total_amount
    issue(qc, "OK" if family_difference == ZERO else "WARN", "FAMILY_OVERVIEW_RECONCILIATION",
          f"基础资产集团概览加总与市场总额差额 {family_difference} 亿元")

    trust_totals: dict[str, dict[str, Decimal]] = {}
    unmapped_trust_assets: dict[str, Decimal] = {}
    for record in market_2026:
        if not is_trust_channel(record):
            continue
        family = asset_family(record["base_asset"])
        if family == UNKNOWN_ASSET_FAMILY:
            unmapped_trust_assets[record["base_asset"]] = (
                unmapped_trust_assets.get(record["base_asset"], ZERO) + record["amount"]
            )
        channel = trust_totals.setdefault(record["originator"], {name: ZERO for name in family_names})
        channel[family] += record["amount"]

    trust_ranked = []
    for name, segments in trust_totals.items():
        total = sum(segments.values(), ZERO)
        if total > ZERO:
            trust_ranked.append({"name": name, "amount": total, "segments": segments})
    trust_ranked.sort(key=lambda item: (-item["amount"], item["name"]))
    trust_total = sum((item["amount"] for item in trust_ranked), ZERO)
    trust_display = trust_ranked[:TRUST_TOP_N]
    trust_remaining = trust_ranked[TRUST_TOP_N:]
    if trust_remaining:
        other_segments = {name: ZERO for name in family_names}
        for item in trust_remaining:
            for family, amount in item["segments"].items():
                other_segments[family] += amount
        trust_display.append({
            "name": "其他",
            "amount": sum(other_segments.values(), ZERO),
            "segments": other_segments,
            "merged_channels": len(trust_remaining),
        })
    displayed_trust_total = sum((item["amount"] for item in trust_display), ZERO)
    difference = displayed_trust_total - trust_total
    issue(qc, "OK" if difference == ZERO else "WARN", "TRUST_CHANNEL_RECONCILIATION",
          f"信托渠道展示加总与全量信托渠道总额差额 {difference} 亿元",
          displayed=str(displayed_trust_total), total=str(trust_total))
    for item in trust_display:
        row_difference = sum(item["segments"].values(), ZERO) - item["amount"]
        issue(qc, "OK" if row_difference == ZERO else "WARN", "TRUST_CHANNEL_ROW_RECONCILIATION",
              f"{item['name']} 分段加总与渠道总额差额 {row_difference} 亿元")
    unknown_segment_total = sum((item["segments"].get(UNKNOWN_ASSET_FAMILY, ZERO) for item in trust_display), ZERO)
    unmapped_total = sum(unmapped_trust_assets.values(), ZERO)
    issue(qc, "OK" if unknown_segment_total == unmapped_total else "WARN", "TRUST_UNKNOWN_ASSET_RECONCILIATION",
          f"信托渠道未知资产分段与基础资产明细差额 {unknown_segment_total - unmapped_total} 亿元")
    if unmapped_trust_assets:
        issue(qc, "OK", "TRUST_UNKNOWN_ASSETS",
              f"信托渠道存在 {len(unmapped_trust_assets)} 类未知资产，已归入未知资产分段",
              assets={name: str(amount) for name, amount in sorted(unmapped_trust_assets.items())})
    if not trust_ranked:
        issue(qc, "WARN", "TRUST_CHANNEL_EMPTY", "无可展示的信托渠道发行记录")

    # 周序（按各周内最早簿记日排序，供卡片价格走势使用）
    week_min_date: dict[str, date] = {}
    for r in market_2026:
        if r["date"] and (r["week"] not in week_min_date or r["date"] < week_min_date[r["week"]]):
            week_min_date[r["week"]] = r["date"]
    weeks_ordered = [w for w, _ in sorted(week_min_date.items(), key=lambda item: item[1])]

    base_totals: dict[str, Decimal] = {}
    for r in market_2026:
        base_totals[r["base_asset"]] = base_totals.get(r["base_asset"], ZERO) + r["amount"]
    # 入围口径：严格按 2026 累计规模降序取 Top N；展示顺序：同集团相邻聚类
    selected = sorted(base_totals.items(), key=lambda item: -item[1])[:TOP_N]
    top_bases = cluster_by_family(selected)

    latest_month = latest_date.month
    previous_month = latest_month - 1

    cards = []
    for base in top_bases:
        card_records = [r for r in market_2026 if r["base_asset"] == base]
        monthly = {month: ZERO for month in range(1, 13)}
        for r in card_records:
            month = month_key(r)
            if month:
                monthly[month] += r["amount"]
        weekly_rates = []
        for week in weeks_ordered:
            avg = weighted_avg([(r["aaa"], r["amount"]) for r in card_records if r["week"] == week])
            if avg is not None:
                weekly_rates.append({"week": week, "date": week_min_date.get(week), "avg": avg})
        originator_totals: dict[str, Decimal] = {}
        for r in card_records:
            originator_totals[r["originator"]] = originator_totals.get(r["originator"], ZERO) + r["amount"]
        originators = [{"name": name, "amount": amount}
                       for name, amount in sorted(originator_totals.items(), key=lambda item: -item[1])]
        mom_rows = []
        for entry in originators:
            name = entry["name"]
            cur = sum((r["amount"] for r in card_records
                       if r["originator"] == name and month_key(r) == latest_month), ZERO)
            prev = sum((r["amount"] for r in card_records
                        if r["originator"] == name and month_key(r) == previous_month), ZERO)
            delta = cur - prev
            status = ("新增" if prev == ZERO and cur > ZERO else
                      "增长" if delta > ZERO else "下降" if delta < ZERO else "持平")
            mom_rows.append({"name": name, "prev": prev, "cur": cur, "delta": delta, "status": status})
        mom_rows.sort(key=lambda row: -(row["cur"] + row["prev"]))
        if not originators or not weekly_rates:
            issue(qc, "WARN", "CARD_DATA_INCOMPLETE", f"{base} 卡片数据不完整：权益人 {len(originators)} 条 / 周度利率 {len(weekly_rates)} 点")
        cards.append({
            "name": base,
            "asset_type": card_records[0]["asset_type"] if card_records else "",
            "count": len(card_records),
            "amount": sum((r["amount"] for r in card_records), ZERO),
            "monthly": monthly,
            "weekly_rates": weekly_rates,
            "originators": originators,
            "mom": {"current_month": latest_month, "previous_month": previous_month, "rows": mom_rows},
        })

    # 卡片月度加总与卡片总量一致性（构造性校验）
    for card in cards:
        difference = sum(card["monthly"].values(), ZERO) - card["amount"]
        issue(qc, "OK" if difference == ZERO else "WARN", "MONTHLY_RECONCILIATION",
              f"{card['name']} 月度加总与卡片总量差额 {difference} 亿元")

    data = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "meta": {
            "latest_date": latest_date,
            "latest_week": latest_week,
            "yoy_cutoff": cutoff,
            "excluded_jd": {
                "y2026": {"count": len(jd_2026), "amount": sum((r["amount"] for r in jd_2026), ZERO)},
                "y2025": {"count": len(jd_2025)},
            },
        },
        "kpis": {
            "total_2026": {"count": len(market_2026), "amount": total_amount},
            "yoy": {"count": len(market_2025), "amount": prev_amount, "pct": yoy_pct},
            "by_family": by_family,
            "trust_channels": {
                "total": trust_total,
                "total_channels": len(trust_ranked),
                "top_n": TRUST_TOP_N,
                "families": list(family_names),
                "rows": trust_display,
                "unmapped_assets": unmapped_trust_assets,
            },
            "latest_week": {
                "label": latest_week,
                "count": len(week_records),
                "aaa": None if aaa_wavg is None else {
                    "wavg": aaa_wavg, "min": aaa_values[0], "max": aaa_values[-1], "n": len(aaa_values)},
                "term": None if term_wavg is None else {
                    "wavg": term_wavg, "min": term_values[0], "max": term_values[-1], "n": len(term_values)},
            },
        },
        "cards": cards,
    }
    return data, qc


# ---------------------------------------------------------------------------
# 格式化
# ---------------------------------------------------------------------------

def yi(value: Decimal | None) -> str:
    return "—" if value is None else f"{value:,.0f}"


def pct2(value: Decimal | None, signed: bool = True) -> str:
    if value is None:
        return "—"
    return f"{value:+.2f}%" if signed else f"{value:.2f}%"


def rate_pct(value: Decimal | None) -> str:
    """小数形式利率（0.0175）→ 百分比展示（1.75%）。"""
    return "—" if value is None else pct2(value * HUNDRED, signed=False)


def json_default(value: Any):
    if isinstance(value, (Decimal, date, datetime)):
        return str(value)
    raise TypeError(type(value).__name__)


def delta_text(delta: Decimal | None, pct: Decimal | None, status: str, unit: str = "亿元") -> str:
    if delta is None:
        return "无可比上期"
    if status == "新增":
        return f"新增 {yi(delta)} {unit}"
    return f"{status} {yi(delta)} {unit}（{pct2(pct)}）"


# ---------------------------------------------------------------------------
# 渲染
# ---------------------------------------------------------------------------

PEER_ISSUANCE_CSS = """
.peer-issuance-root{--ink:#0d1b2e;--paper:#FDFBF7;--muted:#666;font-family:-apple-system,"PingFang SC","Microsoft YaHei",sans-serif;color:#0b0b0b;background:#f4f5f7;min-height:100vh;padding:24px;box-sizing:border-box}.peer-issuance-shell{max-width:1420px;margin:auto}.peer-issuance-hero{background:linear-gradient(135deg,#1a3a5c,#0d1b2e);color:#fff;padding:24px 28px;border-radius:10px 10px 0 0}.peer-issuance-hero h1{margin:0;font-size:25px}.peer-issuance-hero p{margin:7px 0 0;color:#dce7f2;font-size:13px}.peer-issuance-date-grid{display:flex;gap:8px;flex-wrap:wrap;margin-top:16px}.peer-issuance-date{background:rgba(255,255,255,.13);padding:7px 10px;border-radius:5px;font-size:12px}.peer-issuance-section,.peer-issuance-card{background:var(--paper);border:1px solid rgba(11,11,11,.1);border-radius:8px;box-shadow:0 1px 3px rgba(0,0,0,.05)}.peer-issuance-section h2{font-size:16px;color:var(--ink);margin:0 0 4px}.peer-issuance-section>p{margin:0 0 15px;font-size:12px;color:var(--muted)}.peer-issuance-bar-row{display:grid;grid-template-columns:110px 1fr 236px;gap:10px;align-items:center;margin:11px 0}.peer-issuance-label{font-size:13px;font-weight:600;color:#52514e}.peer-issuance-track{height:11px;border-radius:7px;background:#edf0f2;overflow:hidden}.peer-issuance-track span{height:100%;display:block;border-radius:7px;min-width:2px}.peer-issuance-value{display:flex;flex-direction:column;align-items:flex-end;font-size:14px;line-height:1.12;text-align:right;font-variant-numeric:tabular-nums}.peer-issuance-value small{font-size:10px;line-height:1.15;margin-top:1px}.peer-issuance-cards{display:grid;grid-template-columns:minmax(0,1fr) minmax(0,1fr);gap:16px;margin-bottom:0}.peer-issuance-card{padding:18px;min-width:0}.peer-issuance-card-head{display:flex;align-items:baseline;gap:8px;flex-wrap:wrap;border-bottom:1px solid rgba(11,11,11,.08);padding-bottom:10px;margin-bottom:6px}.peer-issuance-card-head h3{font-size:16px;color:var(--ink);margin:0}.peer-issuance-tag{background:#edf0f2;color:#52514e;font-size:11px;padding:2px 8px;border-radius:10px}.peer-issuance-card-amount{margin-left:auto;font-size:14px;font-weight:600;font-variant-numeric:tabular-nums}.peer-issuance-card h4{font-size:15px;color:#52514e;margin:16px 0 8px}.peer-issuance-months{display:flex;gap:6px;align-items:flex-end}.peer-issuance-month{flex:1;display:flex;flex-direction:column;align-items:center;gap:3px;min-width:0}.peer-issuance-month b{font-size:10px;font-weight:600;color:#52514e;font-variant-numeric:tabular-nums;white-space:nowrap}.peer-issuance-mbar{width:100%;height:56px;display:flex;align-items:flex-end}.peer-issuance-mbar span{display:block;width:100%;background:#2a78d6;border-radius:3px 3px 0 0}.peer-issuance-month i{font-style:normal;font-size:10px;color:#999}.peer-issuance-month.off b,.peer-issuance-month.off i{visibility:hidden}.peer-issuance-month.off .peer-issuance-mbar span{display:none}.peer-issuance-month.off .peer-issuance-mbar{background:repeating-linear-gradient(45deg,#f2f3f5,#f2f3f5 4px,#fafafa 4px,#fafafa 8px);border-radius:3px}.peer-issuance-chart{width:100%;height:auto;display:block}.peer-issuance-spark-empty{font-size:12px;color:#999;padding:20px 0;text-align:center}.peer-issuance-delta{display:block;font-size:12px;font-weight:600}.peer-issuance-delta.up{color:#006300}.peer-issuance-delta.down{color:#d03b3b}.peer-issuance-delta.flat{color:#52514e}.peer-issuance-foot{font-size:12px;line-height:1.4;color:#52514e}.peer-issuance-foot p{margin:2px 0}.peer-issuance-foot h2{margin-bottom:5px}@media(max-width:900px){.peer-issuance-cards{grid-template-columns:1fr 1fr}}@media(max-width:620px){.peer-issuance-root{padding:12px}.peer-issuance-cards{grid-template-columns:1fr}.peer-issuance-bar-row{grid-template-columns:90px 1fr 150px}.peer-issuance-hero{padding:20px}}
"""


def safe_text(value: Any) -> str:
    return html.escape(str(value), quote=True)


PEER_ISSUANCE_EXTRA_CSS = """
.peer-issuance-section{margin:18px 0 0;padding:18px 20px}.peer-issuance-section>p{margin-bottom:16px}.peer-issuance-cards{gap:22px;margin-top:18px}.peer-issuance-card{padding:20px}.peer-issuance-trust-row{display:grid;grid-template-columns:128px minmax(0,1fr) 152px;gap:12px;align-items:start;margin:14px 0}.peer-issuance-trust-name{font-size:13px;font-weight:600;color:#52514e;padding-top:2px}.peer-issuance-trust-detail{min-width:0}.peer-issuance-stack-track{height:16px;background:#edf0f2;border-radius:8px;overflow:hidden;display:flex}.peer-issuance-stack-segment{height:100%;min-width:0;display:block}.peer-issuance-trust-value{font-size:13px;text-align:right;font-variant-numeric:tabular-nums;color:#353535;padding-top:1px}.peer-issuance-trust-segment-labels{display:flex;flex-wrap:wrap;gap:5px 11px;margin-top:7px}.peer-issuance-trust-segment-label{display:inline-flex;align-items:center;gap:4px;font-size:11px;color:#5d5d5d;line-height:1.25;white-space:nowrap}.peer-issuance-trust-segment-label i{width:7px;height:7px;border-radius:2px;display:inline-block;flex:0 0 auto}.peer-issuance-legend{display:flex;flex-wrap:wrap;gap:8px 15px;margin:0 0 14px;font-size:12px;color:#52514e}.peer-issuance-legend-item{display:inline-flex;align-items:center;gap:5px}.peer-issuance-legend-item i{width:9px;height:9px;border-radius:2px;display:inline-block}.peer-issuance-empty{font-size:13px;color:#999;margin:8px 0}.peer-issuance-foot{margin-top:22px}.peer-issuance-foot p{margin:3px 0}@media(max-width:620px){.peer-issuance-section{padding:16px}.peer-issuance-trust-row{grid-template-columns:86px minmax(0,1fr);gap:8px}.peer-issuance-trust-value{grid-column:2;font-size:12px;text-align:left;padding-top:0}.peer-issuance-legend{gap:6px 10px}.peer-issuance-trust-segment-labels{gap:5px 8px}}
"""


PEER_ISSUANCE_COMPONENT_CSS = PEER_ISSUANCE_CSS + PEER_ISSUANCE_EXTRA_CSS


def type_rows(by_type: list[dict], total: Decimal) -> str:
    rows = []
    for item in by_type:
        share = ZERO if total == ZERO else item["amount_2026"] / total * HUNDRED
        delta = item["amount_2026"] - item["amount_2025"]
        status = ("新增" if item["amount_2025"] == ZERO and item["amount_2026"] > ZERO else
                  "增长" if delta > ZERO else "下降" if delta < ZERO else "持平")
        state = "up" if status in {"增长", "新增"} else "down" if status == "下降" else "flat"
        rows.append(
            f'<div class="peer-issuance-bar-row"><div class="peer-issuance-label">{safe_text(item["name"])}</div>'
            f'<div class="peer-issuance-track"><span style="width:{float(share):.2f}%;background:{ASSET_FAMILY_COLORS.get(item["name"], BAR_BLUE)}"></span></div>'
            f'<div class="peer-issuance-value"><span>{yi(item["amount_2026"])} 亿 · {float(share):.0f}%</span>'
            f'<small class="peer-issuance-delta {state}">去年同期 {yi(item["amount_2025"])} 亿，{delta_text(delta, item["pct"], status)}</small></div></div>')
    return "".join(rows)


def trust_channel_rows(trust_channels: dict[str, Any]) -> str:
    """渲染信托渠道 Top N + 其他的绝对规模堆叠横条。"""
    rows = trust_channels["rows"]
    if not rows:
        return '<p class="peer-issuance-empty">暂无可展示的信托渠道数据</p>'
    max_total = max(row["amount"] for row in rows)
    total = trust_channels["total"]
    rendered = []
    for row in rows:
        channel_share = ZERO if total == ZERO else row["amount"] / total * HUNDRED
        segments = []
        segments = []
        labels = []
        for family in trust_channels["families"]:
            amount = row["segments"].get(family, ZERO)
            if amount <= ZERO:
                continue
            width = amount / max_total * HUNDRED
            label = f"{family} {yi(amount)} 亿（{float(amount / row['amount'] * HUNDRED):.0f}%）"
            color = ASSET_FAMILY_COLORS[family]
            segments.append(
                f'<span class="peer-issuance-stack-segment" style="width:{float(width):.3f}%;background:{color}" title="{safe_text(label)}"></span>'
            )
            labels.append(
                f'<span class="peer-issuance-trust-segment-label"><i style="background:{color}"></i>{safe_text(label)}</span>'
            )
        merged_note = f' · 合并 {row["merged_channels"]} 家' if row.get("merged_channels") else ""
        rendered.append(
            f'<div class="peer-issuance-trust-row"><div class="peer-issuance-trust-name">{safe_text(row["name"])}</div>'
            f'<div class="peer-issuance-trust-detail"><div class="peer-issuance-stack-track" aria-label="{safe_text(row["name"])}发行规模">{"".join(segments)}</div>'
            f'<div class="peer-issuance-trust-segment-labels">{"".join(labels)}</div></div>'
            f'<div class="peer-issuance-trust-value">{yi(row["amount"])} 亿 · {float(channel_share):.0f}%{merged_note}</div></div>'
        )
    legend = "".join(
        f'<span class="peer-issuance-legend-item"><i style="background:{ASSET_FAMILY_COLORS[family]}"></i>{family}</span>'
        for family in trust_channels["families"]
    )
    return f'<div class="peer-issuance-legend">{legend}</div>{"".join(rendered)}'


def month_cells(monthly: dict[int, Decimal]) -> str:
    max_value = max(monthly.values()) or Decimal("1")
    cells = []
    for month in range(1, 13):
        value = monthly.get(month, ZERO)
        if value > ZERO:
            height = max(float(value / max_value * 56), 3)
            cells.append(
                f'<div class="peer-issuance-month"><b>{yi(value)}</b>'
                f'<div class="peer-issuance-mbar"><span style="height:{height:.0f}px"></span></div><i>{month}月</i></div>')
        else:
            cells.append(
                f'<div class="peer-issuance-month off"><b>0</b>'
                f'<div class="peer-issuance-mbar"></div><i>{month}月</i></div>')
    return "".join(cells)


def rate_sparkline(points: list[dict], width: int = 560, height: int = 150) -> str:
    """带坐标轴与网格线的周度利率折线图（内联 SVG，全数据点标记）。"""
    if not points:
        return '<div class="peer-issuance-spark-empty">暂无利率数据</div>'
    values = [point["avg"] for point in points]
    count = len(values)
    vmin, vmax = min(values), max(values)
    # 纵轴留 15% 余量，网格 4 条水平线
    pad = (vmax - vmin) * Decimal("0.15") if vmax > vmin else Decimal("0.001")
    lo, hi = vmin - pad, vmax + pad
    left, right = Decimal(46), Decimal(width - 10)
    top, bottom = Decimal(10), Decimal(height - 26)

    def x_at(index: int) -> Decimal:
        if count == 1:
            return (left + right) / 2
        return left + (right - left) * Decimal(index) / Decimal(count - 1)

    def y_at(value: Decimal) -> Decimal:
        return bottom - (value - lo) / (hi - lo) * (bottom - top)

    parts = [f'<svg class="peer-issuance-chart" viewBox="0 0 {width} {height}" role="img" aria-label="AAA利率走势">']
    # 网格线 + 纵轴刻度（4 档）
    for i in range(5):
        value = lo + (hi - lo) * Decimal(i) / Decimal(4)
        y = y_at(value)
        parts.append(f'<line x1="{float(left):.0f}" y1="{float(y):.1f}" x2="{float(right):.0f}" y2="{float(y):.1f}" stroke="#e8eaee" stroke-width="1"/>')
        parts.append(f'<text x="{float(left) - 6:.0f}" y="{float(y) + 3:.1f}" font-size="9" fill="#999" text-anchor="end">{rate_pct(value)}</text>')
    # 坐标轴
    parts.append(f'<line x1="{float(left):.0f}" y1="{float(top):.0f}" x2="{float(left):.0f}" y2="{float(bottom):.0f}" stroke="#c9cdd4" stroke-width="1"/>')
    parts.append(f'<line x1="{float(left):.0f}" y1="{float(bottom):.0f}" x2="{float(right):.0f}" y2="{float(bottom):.0f}" stroke="#c9cdd4" stroke-width="1"/>')
    # 折线 + 全数据点
    poly = " ".join(f"{float(x_at(i)):.1f},{float(y_at(v)):.1f}" for i, v in enumerate(values))
    if count > 1:
        parts.append(f'<polyline points="{poly}" fill="none" stroke="#2a78d6" stroke-width="1.6"/>')
    for i, v in enumerate(values):
        parts.append(f'<circle cx="{float(x_at(i)):.1f}" cy="{float(y_at(v)):.1f}" r="2.2" fill="#2a78d6"/>')
    # 横轴日期标签:每隔一个数据点标注一个(索引 0,2,4,...;末点强制包含避免尾部留白)
    label_indexes = set(range(0, count, 2))
    label_indexes.add(count - 1)
    for i in sorted(label_indexes):
        label = points[i].get("date")
        text = label.strftime("%m/%d") if hasattr(label, "strftime") else points[i]["week"][:10]
        anchor = "start" if i == 0 else "end" if i == count - 1 else "middle"
        parts.append(f'<text x="{float(x_at(i)):.1f}" y="{float(bottom) + 14:.0f}" font-size="9" fill="#999" text-anchor="{anchor}">{text}</text>')
    # 数值标注:同样每隔一个数据点标注一个,与横轴标签索引对齐;位于点上方
    for i in sorted(label_indexes):
        v = values[i]
        x = float(x_at(i))
        anchor = "start" if i == 0 else "end" if i == count - 1 else "middle"
        parts.append(f'<text x="{x:.1f}" y="{max(float(y_at(v)) - 7, 9):.0f}" font-size="9" fill="#52514e" text-anchor="{anchor}">{rate_pct(v)}</text>')
    parts.append('</svg>')
    return "".join(parts)


def originator_rows(originators: list[dict], card_total: Decimal, mom: dict, latest_date: date) -> str:
    """权益人全年分布 + 当月环比 合并区块：横条为全年规模，右侧为环比变化。"""
    cur_m, prev_m = mom["current_month"], mom["previous_month"]
    mom_by_name = {row["name"]: row for row in mom["rows"]}
    rows = []
    for entry in originators:
        share = ZERO if card_total == ZERO else entry["amount"] / card_total * HUNDRED
        row = mom_by_name.get(entry["name"], {"prev": ZERO, "cur": ZERO, "delta": ZERO, "status": "持平"})
        if row["status"] == "新增":
            delta_html = f'新增 {yi(row["cur"])} 亿'
        else:
            sign = "+" if row["delta"] > ZERO else ""
            delta_html = f"{sign}{yi(row['delta'])} 亿"
        state = "up" if row["status"] in {"增长", "新增"} else "down" if row["status"] == "下降" else "flat"
        rows.append(
            f'<div class="peer-issuance-bar-row"><div class="peer-issuance-label">{entry["name"]}</div>'
            f'<div class="peer-issuance-track"><span style="width:{float(share):.2f}%;background:{BAR_BLUE}"></span></div>'
            f'<div class="peer-issuance-value"><span>{yi(entry["amount"])} 亿 · {float(share):.0f}%</span>'
            f'<small class="peer-issuance-delta {state}">{prev_m}月 {yi(row["prev"])} → {cur_m}月 {yi(row["cur"])}（{delta_html}）</small></div></div>')
    note = (f'<p style="font-size:10px;color:#999;margin:2px 0 0">横条为 2026 年累计发行规模；'
            f'{cur_m}月为进行中月份，环比统计至 {latest_date.isoformat()}</p>')
    return "".join(rows) + note


def card_html(card: dict, latest_date: date) -> str:
    mom = card["mom"]
    return f'''<section class="peer-issuance-card"><div class="peer-issuance-card-head"><h3>{card['name']}</h3><span class="peer-issuance-tag">{card['asset_type']}</span><span class="peer-issuance-card-amount">{yi(card['amount'])} 亿 · {card['count']} 期</span></div><h4>月度发行规模（亿元）</h4><div class="peer-issuance-months">{month_cells(card['monthly'])}</div><h4>AAA 发行利率走势（周度加权）</h4>{rate_sparkline(card['weekly_rates'])}<h4>原始权益人发行规模与月度变化</h4>{originator_rows(card['originators'], card['amount'], mom, latest_date)}</section>'''


def render_body(data: dict[str, Any]) -> str:
    kpis, meta = data["kpis"], data["meta"]
    total = kpis["total_2026"]
    cards = "".join(card_html(card, meta["latest_date"]) for card in data["cards"])
    trust_channels = kpis["trust_channels"]
    trust_rows = trust_channel_rows(trust_channels)
    return f'''<main class="peer-issuance-root"><div class="peer-issuance-shell"><header class="peer-issuance-hero"><h1>同业发行面板</h1><p>互联网金融 ABS/ABN 同业发行动态 · 剔除京东系资产</p><div class="peer-issuance-date-grid"><span class="peer-issuance-date">2026 数据更新至 {meta['latest_date'].isoformat()}</span><span class="peer-issuance-date">同比窗口 2025-01-01 ~ {meta['yoy_cutoff'].isoformat()}</span><span class="peer-issuance-date">剔除京东系 {meta['excluded_jd']['y2026']['count']} 期 / {yi(meta['excluded_jd']['y2026']['amount'])} 亿</span></div></header><section class="peer-issuance-section"><h2>同业发行概览（2026 累计 {yi(total["amount"])} 亿 / {total["count"]} 期）</h2><p>按具体基础资产归并至资产集团，展示累计发行规模、去年同期规模与变化；未命中归并规则的资产列为“未知资产”。同比基期为 2025 年簿记时间 ≤ {meta['yoy_cutoff'].isoformat()} 的同期窗口</p>{type_rows(kpis["by_family"], total["amount"])}</section><section class="peer-issuance-section"><h2>信托渠道分布</h2><p>仅统计原始权益人名称含“信托”的渠道；按 2026 年累计发行规模取前 {trust_channels["top_n"]} 家，其余合并为“其他”。横条长度表示渠道规模；条下明确展示各资产集团的规模及其在该渠道内占比。</p>{trust_rows}</section><section class="peer-issuance-section"><h2>基础资产发行 Top {TOP_N}</h2><p>按 2026 年累计发行规模排序；每张卡片含月度规模、AAA 利率周度走势、原始权益人分布及月度变化</p></section><div class="peer-issuance-cards">{cards}</div><section class="peer-issuance-section peer-issuance-foot"><h2>数据口径</h2><p>剔除京东系 = 原始权益人或基础资产含“京东”（并集口径，覆盖走信托通道发行的京东资产）；统计范围为全市场互联网金融 ABS/ABN 发行。</p><p>同比 = 2025 年簿记时间 ≤ 2026 最新簿记日前一年的同期窗口；信托渠道仅统计名称含“信托”的原始权益人，按基础资产归并至六类资产集团，未命中规则的资产归入未知资产。</p><p>权益人月度变化中当月为进行中月份，统计至最新簿记日；周度利率走势仅含有有效 AAA 利率的周。</p></section></div></main>'''


def render_html(data: dict[str, Any]) -> str:
    return '<!doctype html><html lang="zh-CN"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>同业发行面板</title><style>' + PEER_ISSUANCE_COMPONENT_CSS + '</style></head><body>' + render_body(data) + '</body></html>'


# ---------------------------------------------------------------------------
# 对外接口与 CLI
# ---------------------------------------------------------------------------

def compute_data(xlsx_2026: str | Path, xlsx_2025: str | Path | None = None) -> tuple[dict[str, Any], list[dict]]:
    current_path = Path(xlsx_2026)
    baseline_path = Path(xlsx_2025) if xlsx_2025 else DEFAULT_BASELINE_PATH
    current_records = parse_dynamics(current_path)
    baseline_records = parse_dynamics(baseline_path)
    data, qc = build_dashboard(current_records, baseline_records)
    data["meta"]["inputs"] = {
        "current": input_fingerprint(current_path, len(current_records)),
        "baseline": input_fingerprint(baseline_path, len(baseline_records)),
    }
    return data, qc


def render_peer_issuance_panel(xlsx_2026: str | Path, xlsx_2025: str | Path | None = None) -> str:
    data, qc = compute_data(xlsx_2026, xlsx_2025)
    warnings = [item for item in qc if item["level"] == "WARN"]
    if warnings:
        raise PanelInputError(f"同业发行 QC 存在告警：{warnings}")
    return render_body(data)


def main() -> None:
    parser = argparse.ArgumentParser(description="生成同业发行面板独立预览")
    parser.add_argument("--xlsx-2026", required=True, help="当期同业发行动态 Excel")
    parser.add_argument("--xlsx-2025", default=None, help=f"同比基准 Excel（默认：{DEFAULT_BASELINE_PATH}）")
    parser.add_argument("--output-html")
    parser.add_argument("--output-json")
    parser.add_argument("--qc-json")
    parser.add_argument("--compare-previous", default=None, help="上周同业发行快照 Excel；提供时生成增量与漂移校验")
    parser.add_argument("--drift-json", default=None, help="增量与漂移校验 JSON 输出路径")
    parser.add_argument("--snapshot-json", default=None, help="规范化事实快照与输入 manifest 输出路径")
    parser.add_argument("--strict-qc", action="store_true")
    args = parser.parse_args()
    data, qc = compute_data(Path(args.xlsx_2026), Path(args.xlsx_2025) if args.xlsx_2025 else None)
    tag = data["meta"]["latest_date"].strftime("%Y%m%d")
    DEFAULT_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    html_path = Path(args.output_html) if args.output_html else DEFAULT_OUTPUT_DIR / f"peer_issuance_panel_{tag}.html"
    json_path = Path(args.output_json) if args.output_json else DEFAULT_OUTPUT_DIR / f"peer_issuance_panel_{tag}.json"
    qc_path = Path(args.qc_json) if args.qc_json else DEFAULT_OUTPUT_DIR / f"peer_issuance_panel_{tag}_qc.json"
    for output in (html_path, json_path, qc_path):
        output.parent.mkdir(parents=True, exist_ok=True)
    html_path.write_text(render_html(data), encoding="utf-8")
    json_path.write_text(json.dumps(data, ensure_ascii=False, indent=2, default=json_default), encoding="utf-8")
    qc_path.write_text(json.dumps(qc, ensure_ascii=False, indent=2, default=json_default), encoding="utf-8")
    snapshot_path = Path(args.snapshot_json) if args.snapshot_json else DEFAULT_OUTPUT_DIR / f"peer_issuance_snapshot_{tag}.json"
    snapshot_path.parent.mkdir(parents=True, exist_ok=True)
    current_records = parse_dynamics(Path(args.xlsx_2026))
    snapshot_path.write_text(json.dumps({"input": input_fingerprint(Path(args.xlsx_2026), len(current_records)), "records": current_records}, ensure_ascii=False, indent=2, default=json_default), encoding="utf-8")
    drift_qc: list[dict] = []
    drift_path = None
    if args.compare_previous:
        previous_path = Path(args.compare_previous)
        summary, drift_qc = compare_weekly_snapshot(parse_dynamics(previous_path), parse_dynamics(Path(args.xlsx_2026)))
        drift_path = Path(args.drift_json) if args.drift_json else DEFAULT_OUTPUT_DIR / f"peer_issuance_drift_{tag}.json"
        drift_path.parent.mkdir(parents=True, exist_ok=True)
        drift_path.write_text(json.dumps({"previous_input": input_fingerprint(previous_path, summary["previous_count"]), "current_input": input_fingerprint(Path(args.xlsx_2026), summary["current_count"]), "summary": summary, "qc": drift_qc}, ensure_ascii=False, indent=2, default=json_default), encoding="utf-8")
    warnings = sum(item["level"] == "WARN" for item in qc + drift_qc)
    failures = sum(item["level"] == "FAIL" for item in qc + drift_qc)
    drift_line = f"\n[完成] Drift: {drift_path}" if drift_path else ""
    print(f"[完成] HTML: {html_path}\n[完成] 数据: {json_path}\n[完成] QC: {qc_path}\n[完成] Snapshot: {snapshot_path}{drift_line}\n[QC] WARN={warnings} FAIL={failures}")
    if failures or (args.strict_qc and warnings):
        raise SystemExit(2)


if __name__ == "__main__":
    main()
