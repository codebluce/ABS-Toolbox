"""ABS发行定价公共模块：合并单元格预处理、列名映射、QC框架

供三个工具脚本共享：
  gen_abs_cost_report.py（工具一：成本分布）
  gen_compare_tool.py    （工具二：机构比对）
  gen_spread_report.py   （工具三：利差分析）
"""

import os
import tempfile
from copy import copy

import pandas as pd

# ── 成本区间配置（工具一使用）──────────────────────────────────────
# 整体 bin 下限 1.30%（保理类资产专用，保理成本低是正常现象）
# 非保理类资产的成本下限由 COST_BIN_LOWER_NON_PREF 控制（1.50%），QC 时单独校验
COST_BIN_LOWER = 1.30              # bin 起始下限
COST_BIN_LOWER_NON_PREF = 1.50     # 非保理类资产的成本下限（QC 校验用）
COST_BIN_UPPER = 3.00
COST_BIN_STEP = 0.05
COST_BINS = [round(COST_BIN_LOWER + i * COST_BIN_STEP, 4)
             for i in range(int(round((COST_BIN_UPPER - COST_BIN_LOWER) / COST_BIN_STEP)) + 1)] + [3.05]
COST_LABELS = [f"{COST_BINS[i]:.2f}%~{COST_BINS[i+1]:.2f}%"
               for i in range(len(COST_BINS) - 1)]

# ── 利差区间配置（工具三使用）──────────────────────────────────────
# 下限 0.00%：覆盖接近零利差的非保理类资产（如金采 0.02%）；负利差仍落入 NaN
SPREAD_BINS   = [round(0.00 + i * 0.05, 4) for i in range(33)] + [1.65]
SPREAD_LABELS = [f"+{SPREAD_BINS[i]:.2f}%~+{SPREAD_BINS[i+1]:.2f}%"
                 for i in range(len(SPREAD_BINS) - 1)]

# ── 台账列名（25列，A~Y）──────────────────────────────────────
COLNAMES = [
    '序号', '月份', '类别', '资产类型', '项目名称', '发行场所',
    '计划管理人', '联席承销商', '托管行', '规模', '期限',
    '簿记时间', '国股CD', 'ALL-IN成本', '认购倍数',
    '分层情况', '分层占比', '对应金额（亿）', '评级',
    '成本', '认购机构', '认购份额',
    '申购利率', '穿透机构', '申购规模',
]

# ── WXY/TUV 列名常量 ──────────────────────────────────────────
RESOLVED_COST_COL = '实际成本'
RESOLVED_INST_COL = '实际机构'
RESOLVED_SHARE_COL = '实际份额'
NON_NUMERIC_COST_VALS = {'成本', '自持', '不出表', '-'}

# ── 期限分类 ───────────────────────────────────────────────────
OVER_1Y_ASSETS = {'信托企业主贷', '小贷企业主贷', '京企贷'}

TENOR_COL_CANDIDATES = ['期限（月）', '期限']

PRODUCT_ORDER_TEMPLATE = [
    '1年期赊销白条', '超1年期赊销白条',
    '1年期信托金条', '超1年期信托金条',
    '1年期白条取现', '超1年期白条取现',
    '1年期信托企业主贷', '超1年期信托企业主贷',
    '1年期小贷企业主贷', '超1年期小贷企业主贷',
]

# ── 优先分层过滤条件 ──────────────────────────────────────────
PRIORITY_LAYERS = ['优先A', '优先级', '优先A1', '优先A2']

# ── 机构黑名单（生成看板时全角色剔除）─────────────────────────
# 子串包含匹配：机构名任一字段包含关键词即剔除整行
EXCLUDED_INSTITUTIONS = ['盛际', '邦汇', '厦门汇正']

# 黑名单作用的所有角色字段（任一命中即剔除整行）
EXCLUDED_ROLE_COLS = ['计划管理人', '联席承销商', '托管行', '认购机构', '穿透机构']

# ── 保理类资产豁免（仅豁免利差 bin QC）──────────────────────────
# 资产类型字段包含关键词即认定为保理类
PREFERENTIAL_ASSET_KEYWORDS = ['保理']

# ── 配色 ───────────────────────────────────────────────────────
ASSET_COLORS = {
    '赊销白条':              {'header': '#1a3a5c', 'tag': '#2563a8'},
    '信托金条':              {'header': '#1c3a2a', 'tag': '#2d7a4f'},
    '信托企业主贷':           {'header': '#3a2010', 'tag': '#c05a20'},
    '小贷企业主贷':           {'header': '#2a1040', 'tag': '#6b3daa'},
    '白条取现':              {'header': '#1a3048', 'tag': '#3a6090'},
    '动产质押':              {'header': '#2a2a10', 'tag': '#8a8a30'},
    '保理':                 {'header': '#0d3a3a', 'tag': '#1a7a7a'},
    '京企贷':               {'header': '#3a1010', 'tag': '#a03030'},
    '金条&白取&信托白条混发':  {'header': '#2a2a3a', 'tag': '#5a5a7a'},
    '信托白条':              {'header': '#3a3010', 'tag': '#8a7020'},
    '金条&白取混发':          {'header': '#2a1020', 'tag': '#6a2a6a'},
    '信托白取':              {'header': '#103a2a', 'tag': '#207a5a'},
    '金采':                 {'header': '#1a1a1a', 'tag': '#4a4a4a'},
}
DEFAULT_COLOR = {'header': '#2a2a2a', 'tag': '#555555'}

BASE_COLORS = ASSET_COLORS  # 利差报告用BASE_COLORS作为别名

MONTH_COLORS = [
    '#0d2a4a', '#1a3a6c', '#2a4a8c', '#3a5a9c',
    '#4a6aac', '#5a7abc', '#6a8acc', '#7a9adc',
]
MONTH_ACCENT = '#e8f0fc'


# ── 合并单元格预处理（pandas只读场景）──────────────────────────
def preprocess_xlsx_for_pandas(xlsx_path):
    """预处理合并单元格Excel，返回临时文件路径供pandas读取。

    步骤：
    1. 保存P/U/V/W/X/Y列原始值（pandas只读，无需保存格式）
    2. 取消所有合并单元格
    3. A~T列向下填充（跳过P列），继承格式
    4. 恢复P/U/V/W/X/Y列原始值
    5. 保存为临时文件，返回路径

    调用方负责在pandas读取后删除临时文件（或由系统自动清理）。
    """
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path)
    ws = wb.active

    # 1. 保存P(16)/U(21)/V(22)列合并单元格原始值
    protected_data = {}
    for mrange in list(ws.merged_cells.ranges):
        for row in range(mrange.min_row, mrange.max_row + 1):
            for col in range(mrange.min_col, mrange.max_col + 1):
                if col in (16, 21, 22):
                    val = ws.cell(row=mrange.min_row, column=col).value
                    if val is not None:
                        protected_data[(row, col)] = val

    # 保存W(23)/X(24)/Y(25)列全部数据（只保存值，pandas不需要格式）
    wxy_values = {}
    for r in range(1, ws.max_row + 1):
        for c in (23, 24, 25):
            cell = ws.cell(row=r, column=c)
            if cell.value is not None:
                wxy_values[(r, c)] = cell.value

    # 2. 取消所有合并单元格
    for mrange in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(mrange))

    # 3. A~T列向下填充（跳过P列=16），继承格式
    last_values = {}
    last_cells = {}
    for r in range(2, ws.max_row + 1):
        for c in range(1, 21):
            if c == 16:
                continue
            cell = ws.cell(row=r, column=c)
            if cell.value is not None:
                last_values[c] = cell.value
                last_cells[c] = cell
            elif c in last_values and last_values[c] is not None:
                cell.value = last_values[c]
                src = last_cells[c]
                if src:
                    cell.font = copy(src.font)
                    cell.border = copy(src.border)
                    cell.alignment = copy(src.alignment)
                    cell.number_format = src.number_format

    # 4. 恢复P/U/V列原始值
    for (row, col), val in protected_data.items():
        ws.cell(row=row, column=col).value = val

    # 恢复W/X/Y列原始值（填充范围1-20不含23-25，不会覆盖WXY，此处仅恢复原始值）
    for (r, c), val in wxy_values.items():
        ws.cell(row=r, column=c).value = val

    # 5. 保存为临时文件
    tmp_dir = tempfile.gettempdir()
    tmp_name = f'abs_preprocessed_{os.getpid()}_{id(wb)}.xlsx'
    tmp_path = os.path.join(tmp_dir, tmp_name)
    wb.save(tmp_path)
    wb.close()
    return tmp_path


# ── 公共函数 ───────────────────────────────────────────────────
def classify_tenor(tenor_str, asset_type):
    """将期限字符串解析为 1年期 / 超1年期"""
    import re
    nums = re.findall(r'\d+', str(tenor_str))
    if not nums:
        return '未知'
    x = int(nums[0])
    if x == 12 and asset_type in OVER_1Y_ASSETS:
        return '超1年期'
    return '1年期' if x <= 12 else '超1年期'


def resolve_tenor_col(df):
    """返回实际存在的期限列名，找不到则报错"""
    for c in TENOR_COL_CANDIDATES:
        if c in df.columns:
            return c
    raise ValueError(
        f"[INPUT ERROR] 缺少期限列，期望列名之一：{TENOR_COL_CANDIDATES}\n"
        f"当前列：{list(df.columns)}"
    )


def clean_colnames(df_raw):
    """检查列数并返回25列标准列名"""
    if df_raw.shape[1] < 25:
        raise ValueError(
            f"[INPUT ERROR] 输入文件列数不足（{df_raw.shape[1]}列），"
            "请使用含25列的总表（含WXY补充列）而非子表明细"
        )
    return COLNAMES


def resolve_columns(df):
    """WXY列三元组优先，TUV列整行回退；忽略非数值脏数据

    W/X/Y 代表同一条穿透申购记录，必须作为三元组整体使用。
    旧版逐列 fillna 会在 WXY 部分缺失时制造"W 来自穿透、X/Y 来自表面"的拼接记录。
    """
    w = pd.to_numeric(df['申购利率'], errors='coerce')
    x_raw = df['穿透机构']
    x_str = x_raw.astype(str).str.strip()
    x_present = x_raw.notna() & (x_str != '') & (x_str.str.lower() != 'nan')
    y = pd.to_numeric(df['申购规模'], errors='coerce')

    cost_raw = df['成本'].astype(str).str.strip()
    cost_clean = cost_raw.where(~cost_raw.isin(NON_NUMERIC_COST_VALS), pd.NA)
    t = pd.to_numeric(cost_clean, errors='coerce')

    u = df['认购机构'].where(
        ~df['认购机构'].astype(str).str.strip().isin(NON_NUMERIC_COST_VALS), pd.NA
    )
    v = pd.to_numeric(df['认购份额'], errors='coerce')

    wxy_count = w.notna().astype(int) + x_present.astype(int) + y.notna().astype(int)
    wxy_complete = wxy_count == 3
    wxy_partial = (wxy_count > 0) & (wxy_count < 3)

    if wxy_partial.any():
        sample_cols = [c for c in ['项目名称', '分层', '认购机构', '申购利率', '穿透机构', '申购规模'] if c in df.columns]
        sample = df.loc[wxy_partial, sample_cols].head(5).to_dict('records')
        print(f"[WARN] WXY三元组部分缺失 {int(wxy_partial.sum())} 行，已整行回退 TUV，避免拼接记录。样例: {sample}")

    df[RESOLVED_COST_COL] = t.where(~wxy_complete, w)
    df[RESOLVED_INST_COL] = u.where(~wxy_complete, x_raw)
    df[RESOLVED_SHARE_COL] = v.where(~wxy_complete, y)

    # 机构名自动归并：去公司后缀 + 投行上报后缀
    before_count = df[RESOLVED_INST_COL].nunique()
    df[RESOLVED_INST_COL] = df[RESOLVED_INST_COL].apply(normalize_investor_name)
    after_count = df[RESOLVED_INST_COL].nunique()
    if before_count != after_count:
        print(f'[MERGE] 机构名归并：{before_count} → {after_count}（减少 {before_count - after_count} 个变体）')

    self_hold_mask = df['申购利率'].astype(str).str.strip().str.match(r'^自持$', na=False)
    return df[~self_hold_mask]


# ── 机构名自动归并 ─────────────────────────────────────────────
# 规则：
#   1. 去末尾公司后缀：有限责任公司/股份有限公司/有限公司/股份公司
#   2. 去末尾投行上报后缀：（投行上报）/(投行上报)（全角/半角括号混用）
#   3. 简称归并（显式映射表）：建投→中信建投、中金→中金公司 等
#   4. 括号统一：无括号部门后缀→有括号（XX证券投行→XX证券（投行））
#   5. 不动其他括号后缀：（资管）/（自营）/（机构部）/（衍生品）/（投顾）等保留

import re as _re

_COMPANY_SUFFIXES = [
    '有限责任公司', '股份有限公司', '有限公司', '股份公司',
]
_INVEST_BANK_SUFFIX_PATTERN = _re.compile(
    r'[（(]\s*投行上报\s*[)）]\s*$'
)

# 简称→全称 显式映射表（精确可控）
_INVESTOR_ALIAS_MAP = {
    '建投': '中信建投',
    '建投机构部': '中信建投（机构部）',
    '建投自营': '中信建投（自营）',
    '建投投行': '中信建投（投行）',
    '中金': '中金公司',
    '中金自营': '中金公司（自营）',
    '中金资管': '中金公司（资管）',
    '中金投顾': '中金公司（投顾）',
    '华宝资管': '华宝证券资管',
    '华宝证券（资管）': '华宝证券资管',
    '申万宏源证券资管': '申万宏源资管',
    '申万宏源证券（投顾）': '申万宏源（投顾）',
    '申万宏源证券（自营）': '申万宏源（自营）',
    '申万宏源证券投顾': '申万宏源（投顾）',
    '申万宏源证券自营': '申万宏源（自营）',
}

# 无括号部门后缀 → 有括号（统一写法）
# 匹配 "XX证券投行" / "XX海通投行" / "XX建投投行" 等
# 不匹配 "XX资管"（资管是独立机构名，不是部门）
# 不匹配已含括号的（如 "XX证券（投行）"）
_DEPT_SUFFIX_PATTERN = _re.compile(
    r'^(.+?)(投行|自营|投顾)$'
)


def normalize_investor_name(name):
    """归并机构名变体

    规则顺序：
      1. 去投行上报后缀
      2. 去公司后缀
      3. 显式映射表（简称→全称）
      4. 括号统一（XX证券投行→XX证券（投行））
    """
    if name is None or (isinstance(name, float) and pd.isna(name)):
        return name
    s = str(name).strip()
    if not s:
        return name

    # 规则 2：去投行上报后缀（先去，因为可能在公司后缀之前）
    s = _INVEST_BANK_SUFFIX_PATTERN.sub('', s).strip()

    # 规则 1：去公司后缀
    for suffix in _COMPANY_SUFFIXES:
        if s.endswith(suffix):
            s = s[:-len(suffix)].strip()
            break

    # 规则 3：显式映射表
    if s in _INVESTOR_ALIAS_MAP:
        s = _INVESTOR_ALIAS_MAP[s]

    # 规则 4：括号统一（XX证券投行→XX证券（投行），不处理资管）
    m = _DEPT_SUFFIX_PATTERN.match(s)
    if m:
        s = f'{m.group(1)}（{m.group(2)}）'

    return s


def validate_input(df, extra_required=None):
    """Fail-fast 入口校验：缺失列 + 关键列空值比例"""
    required = ['簿记时间', '资产类型', '分层情况',
                RESOLVED_SHARE_COL, RESOLVED_INST_COL, RESOLVED_COST_COL]
    if extra_required:
        required.extend(extra_required)

    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(
            f"[INPUT ERROR] 缺少必需列：{missing}\n"
            f"当前列：{list(df.columns)}\n"
            f"请检查Excel表头是否与示例一致"
        )

    # 关键列空值比例检查
    null_check_cols = [c for c in required if c in df.columns]
    null_ratio = df[null_check_cols].isna().mean()
    high_null = null_ratio[null_ratio > 0.3].index.tolist()
    if high_null:
        from warnings import warn
        warn(f"[INPUT WARN] 以下关键列空值比例>30%：{high_null}，数据可能不完整")


def is_excluded_institution(name):
    """机构名是否命中黑名单关键词（子串包含匹配）"""
    if name is None or pd.isna(name):
        return False
    s = str(name)
    return any(kw in s for kw in EXCLUDED_INSTITUTIONS)


def is_preferential_asset(asset_type):
    """资产类型是否为保理类（用于利差 bin QC 豁免）"""
    if asset_type is None or pd.isna(asset_type):
        return False
    s = str(asset_type)
    return any(kw in s for kw in PREFERENTIAL_ASSET_KEYWORDS)


# ── 簿记日期纠错（2025→2026）──────────────────────────────────
# 业务规则：所有存续项目和新增项目一定发生在 2026 年，不可能是 2025 年。
# 原始台账人工录入时偶尔会把 2026 误写为 2025，本函数自动纠正。
# 纠错策略：年份 2025 → 2026（同月同日），但仅当该项目名下没有 2026 年记录时改。
# 若同一项目名下既有 2025 又有 2026 记录（可能跨年发新），不自动改，warn 提示人工确认。

WRONG_YEAR = 2025
CORRECT_YEAR = 2026


def fix_bookkeeping_year(df, project_col='项目名称', date_col='簿记时间', label='日期纠错'):
    """纠正簿记日期里 2025 年的错误年份（改为 2026 年同月同日）

    Args:
        df: DataFrame，含 project_col 和 date_col 两列
        project_col: 项目名称列名
        date_col: 簿记时间列名（datetime 类型）
        label: 日志标签

    Returns:
        纠错后的 DataFrame（不修改原 df）

    行为：
        - 找出 date_col 年份=2025 的行
        - 对每行：检查该项目名下是否还有 2026 年的记录
          - 没有 2026 年记录（即整个项目都录错了年份）：自动改为 2026 年同月同日
          - 有 2026 年记录（可能跨年发新）：不改，warn 提示人工确认
        - 其他年份不动
    """
    if date_col not in df.columns or project_col not in df.columns:
        return df

    # 确保 date_col 是 datetime
    if not pd.api.types.is_datetime64_any_dtype(df[date_col]):
        return df

    wrong_mask = df[date_col].dt.year == WRONG_YEAR
    if not wrong_mask.any():
        return df

    df_fixed = df.copy()
    n_fixed = 0
    # 每个项目名下的所有年份集合
    proj_years = df.groupby(project_col)[date_col].apply(
        lambda s: set(s.dropna().dt.year.unique())
    )

    fixed_projects = set()
    for idx in df[wrong_mask].index:
        proj = df.at[idx, project_col]
        if pd.isna(proj):
            continue
        years_of_proj = proj_years.get(proj, set())
        # 若该项目只有 2025 年记录、没有 2026 年记录 → 自动纠错
        if CORRECT_YEAR not in years_of_proj and WRONG_YEAR in years_of_proj:
            old_dt = df.at[idx, date_col]
            try:
                new_dt = old_dt.replace(year=CORRECT_YEAR)
            except ValueError:
                # 处理 2/29 这种闰年日期
                new_dt = old_dt.replace(year=CORRECT_YEAR, day=28)
            df_fixed.at[idx, date_col] = new_dt
            n_fixed += 1
            if proj not in fixed_projects:
                print(f'[FIX] {label}: 项目"{proj}" 簿记日期 {old_dt.strftime("%Y-%m-%d")} → {new_dt.strftime("%Y-%m-%d")}（年份 2025→2026，项目仅有 2025 年记录）')
                fixed_projects.add(proj)
        else:
            # 该项目既有 2025 又有 2026，可能是跨年发新，不改
            print(f'[WARN] {label}: 项目"{proj}" 簿记日期 {df.at[idx, date_col].strftime("%Y-%m-%d")} 落在 2025 年，但该项目同时有 2026 年记录（可能跨年发新），未自动纠错，请人工确认')

    if n_fixed > 0:
        print(f'[FIX] {label}: 共纠正 {n_fixed} 条 2025 年簿记日期 → 2026 年（涉及 {len(fixed_projects)} 个项目）')
    return df_fixed


def filter_excluded_institutions(df, role_cols=None, label='数据加载'):
    """剔除黑名单机构相关记录

    Args:
        df: 原始 DataFrame
        role_cols: 参与匹配的角色字段列表（默认 EXCLUDED_ROLE_COLS）
                   只匹配 df 中实际存在的列
        label: 日志标签（用于打印剔除摘要）

    Returns:
        过滤后的 DataFrame（剔除任一角色字段命中黑名单的整行）

    行为说明：
        - 子串包含匹配：'盛际' / '邦汇'
        - 任一角色字段命中即剔除整行（最严格口径）
        - 不修改原 df
    """
    if role_cols is None:
        role_cols = EXCLUDED_ROLE_COLS
    cols = [c for c in role_cols if c in df.columns]
    if not cols:
        return df

    mask = pd.Series(False, index=df.index)
    for c in cols:
        mask = mask | df[c].apply(is_excluded_institution)

    n_excluded = int(mask.sum())
    if n_excluded > 0:
        # 摘要：每个角色字段命中数
        details = []
        for c in cols:
            n = int(df[c].apply(is_excluded_institution).sum())
            if n > 0:
                details.append(f'{c}={n}')
        print(f'[FILTER] {label}: 剔除黑名单机构记录 {n_excluded} 行 ({", ".join(details)})')
    return df[~mask].copy()


def load_and_filter(xlsx_path, need_tenor=False, extra_required=None, fix_year=True,
                    preprocessed_path=None):
    """统一的数据读取与预处理入口

    Args:
        xlsx_path: 台账Excel路径
        need_tenor: 是否需要期限分类（工具二/三需要）
        extra_required: 额外必需列
        fix_year: 是否执行"2025→2026簿记日期纠错"（默认True，兼容现有调用）。
                  该纠错逻辑假设台账业务年份为2026，若传入的是2025年历史台账本身
                  （而非当前年台账），必须传 False，否则2025年的真实日期会被误改成2026年。
        preprocessed_path: 已预处理的共享临时文件路径（可选）。传入则跳过自身 preprocess，
                  直接读该文件。综合看板 main 单次预处理后注入各子模块，消除 ~7 次重复预处理。
                  传入时本函数不负责删除 tmp（由创建方统一删）；为 None 时走原逻辑（自己
                  preprocess，调用方负责删除返回的 tmp_path）。

    Returns:
        df: 全量数据DataFrame（过滤自持后）
        dfa: 仅优先A/优先级子集
        tmp_path: 预处理临时文件路径（调用方应负责删除）
        tenor_col: 期限列名（need_tenor=True时）或None
    """
    # 合并单元格预处理
    if preprocessed_path is not None:
        tmp_path = preprocessed_path          # 共享 tmp:跳过 preprocess,由创建方统一删
    else:
        tmp_path = preprocess_xlsx_for_pandas(xlsx_path)
    try:
        df_raw = pd.read_excel(tmp_path, engine='openpyxl', header=None)
    except Exception as e:
        # 预处理已成功生成临时文件，读取失败应报错而非静默回退
        # （回退到未预处理文件会导致合并单元格数据失真）
        # 共享 tmp(preprocessed_path 传入)不删,由创建方统一管理
        if preprocessed_path is None:
            try:
                os.remove(tmp_path)
            except OSError:
                pass
        raise ValueError(
            f"[INPUT ERROR] 预处理后的临时文件读取失败：{e}\n"
            f"临时文件路径：{tmp_path}\n"
            f"请检查原始Excel文件是否损坏"
        )

    colnames = clean_colnames(df_raw)
    df_raw.columns = colnames
    df = df_raw.iloc[2:].copy()
    title_mask = df['成本'].astype(str).str.match(r'^成本$', na=False)
    df = df[~title_mask]
    df.reset_index(drop=True, inplace=True)

    for c in ['规模', '国股CD']:
        df[c] = pd.to_numeric(
            df[c].astype(str).str.strip()
            .replace({v: pd.NA for v in NON_NUMERIC_COST_VALS}),
            errors='coerce'
        )
    df['簿记时间'] = pd.to_datetime(df['簿记时间'], errors='coerce')

    # 簿记日期纠错：2025 年 → 2026 年（仅项目名唯一时自动改；历史年份台账需关闭）
    if fix_year:
        df = fix_bookkeeping_year(df, label='load_and_filter')

    df = resolve_columns(df)
    dfa = df[df['分层情况'].isin(PRIORITY_LAYERS)].copy()

    if need_tenor:
        tenor_col = resolve_tenor_col(dfa)
    else:
        tenor_col = None

    extra = list(extra_required or [])
    if need_tenor:
        extra.append('国股CD')
    validate_input(dfa, extra_required=extra if extra else None)

    # 黑名单剔除：全角色匹配（管理人/联席承销商/托管行/认购机构/穿透机构）
    # 在 validate_input 之后执行，避免影响空值比例检查的诊断价值
    df = filter_excluded_institutions(df, label='load_and_filter(全量)')
    dfa = filter_excluded_institutions(dfa, label='load_and_filter(优先A)')

    return df, dfa, tmp_path, tenor_col


# ── QC 框架基类 ───────────────────────────────────────────────
class QCRunner:
    """QC检查框架：ok/warn/fail + 汇总 + 返回值"""

    def __init__(self, title):
        self.title = title
        self.issues = []
        self.ok_count = 0

    def ok(self, msg):
        self.ok_count += 1
        print(f'  [OK]   {msg}')

    def warn(self, msg):
        self.issues.append(('WARN', msg))
        print(f'  [WARN] {msg}')

    def fail(self, msg):
        self.issues.append(('FAIL', msg))
        print(f'  [FAIL] {msg}')

    def header(self):
        print('\n' + '=' * 52)
        print(f'  QC Report — {self.title}')
        print('=' * 52)

    def summary(self):
        print('\n' + '=' * 52)
        fail_items = [x for x in self.issues if x[0] == 'FAIL']
        warn_items = [x for x in self.issues if x[0] == 'WARN']

        if not fail_items and not warn_items:
            print(f'  QC PASSED — {self.ok_count}项全部通过')
        elif not fail_items:
            print(f'  QC PASSED WITH WARNINGS — {self.ok_count}项通过，{len(warn_items)}项需关注：')
            for _, msg in warn_items:
                print(f'    ! {msg}')
        else:
            print(f'  QC FAILED — {len(fail_items)}项严重错误，{len(warn_items)}项警告：')
            for level, msg in self.issues:
                mark = 'X' if level == 'FAIL' else '!'
                print(f'    {mark} {msg}')
        print('=' * 52 + '\n')
        return not bool(fail_items)


def product_color(asset_type, tenor_label):
    """1年期用基色；超1年期用浅30%亮度的同色相"""
    base = BASE_COLORS.get(asset_type, DEFAULT_COLOR)
    h = base['header']
    t = base['tag']

    def shift_hex(hex_color, amount=50):
        r = min(255, int(hex_color[1:3], 16) + amount)
        g = min(255, int(hex_color[3:5], 16) + amount)
        b = min(255, int(hex_color[5:7], 16) + amount)
        return f'#{r:02x}{g:02x}{b:02x}'

    if '超1年' in tenor_label:
        return {'header': shift_hex(h, 45), 'tag': shift_hex(t, 25)}
    return {'header': h, 'tag': t}
