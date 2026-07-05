#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
增量台账合并 v2.1
功能：
  - 增量合并：将本周新原始台账中的增量项目追加到上周已加工台账，并录入簿记明细
  - 补充簿记（--supplement）：对已有台账中的指定项目补充录入簿记明细，不做增量合并
核心流程（增量合并）：
1. 预处理两个台账（已加工台账保护WXY列）
2. 项目差分识别增量
3. 从新台账提取增量项目A~V行
4. 追加到已加工台账末尾
5. 对目标项目执行簿记录入
6. 全表格式优化 + QC（含存量WXY保护验证）
核心流程（补充簿记）：
1. 预处理已加工台账（保护WXY列）
5. 对有明细的项目执行簿记录入
6. 全表格式优化 + QC（含存量WXY保护验证）
"""
import openpyxl, sys, os, argparse
from collections import OrderedDict, defaultdict
from copy import copy

# ============================================================
# 复用 insert_bookkeeping.py 的核心函数
# ============================================================

def normalize(name):
    if not name: return ""
    n = str(name).strip().replace("（", "(").replace("）", ")").replace(" ", "").replace("\u3000", "")
    for suffix in ["(投行上报)", "(固定收益部)", "(投行)", "(资管)", "自营", "经营", "产品", "部门"]:
        n = n.replace(suffix, "")
    return n


def unmerge_and_fill_raw(ws):
    """预处理新原始台账：取消合并+向下填充，保护P/U/V"""
    protected_data = {}
    for mrange in list(ws.merged_cells.ranges):
        for row in range(mrange.min_row, mrange.max_row + 1):
            for col in range(mrange.min_col, mrange.max_col + 1):
                if col in (16, 21, 22):
                    val = ws.cell(row=mrange.min_row, column=col).value
                    if val is not None:
                        protected_data[(row, col)] = val
    for mrange in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(mrange))
    last_values = {}
    last_cells = {}
    for r in range(2, ws.max_row + 1):
        for c in range(1, 21):
            if c == 16: continue
            cell = ws.cell(row=r, column=c)
            if cell.value is not None:
                last_values[c] = cell.value
                last_cells[c] = cell
            elif c in last_values and last_values[c] is not None:
                cell.value = last_values[c]
                src = last_cells[c]
                if src:
                    cell.font = copy(src.font)
                    cell.border = copy(src.border)
                    cell.alignment = copy(src.alignment)
                    cell.number_format = src.number_format
    for (row, col), val in protected_data.items():
        ws.cell(row=row, column=col).value = val
    return ws


def unmerge_and_fill_processed(ws):
    """预处理已加工台账：取消合并+向下填充，保护P/U/V/W/X/Y"""
    # 1. Save P/U/V/W/X/Y from merged cells AND non-merged WXY data
    protected_data = {}
    for mrange in list(ws.merged_cells.ranges):
        for row in range(mrange.min_row, mrange.max_row + 1):
            for col in range(mrange.min_col, mrange.max_col + 1):
                if col in (16, 21, 22):
                    val = ws.cell(row=mrange.min_row, column=col).value
                    if val is not None:
                        protected_data[(row, col)] = val

    # Save ALL WXY data (not just merged cells - WXY may not be merged)
    wxy_data = {}  # (row, col) -> (value, font, border, alignment, number_format)
    for r in range(1, ws.max_row + 1):
        for c in (23, 24, 25):
            cell = ws.cell(row=r, column=c)
            if cell.value is not None:
                wxy_data[(r, c)] = (
                    cell.value,
                    copy(cell.font), copy(cell.border),
                    copy(cell.alignment), cell.number_format
                )

    # 2. Unmerge all
    for mrange in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(mrange))

    # 3. Fill A~T (skip P), inherit format
    last_values = {}
    last_cells = {}
    for r in range(2, ws.max_row + 1):
        for c in range(1, 21):
            if c == 16: continue
            cell = ws.cell(row=r, column=c)
            if cell.value is not None:
                last_values[c] = cell.value
                last_cells[c] = cell
            elif c in last_values and last_values[c] is not None:
                cell.value = last_values[c]
                src = last_cells[c]
                if src:
                    cell.font = copy(src.font)
                    cell.border = copy(src.border)
                    cell.alignment = copy(src.alignment)
                    cell.number_format = src.number_format

    # 4. Restore P/U/V
    for (row, col), val in protected_data.items():
        ws.cell(row=row, column=col).value = val

    # 5. Restore WXY data (critical: fill may have overwritten WXY in empty rows)
    # First clear any WXY that was set by fill (shouldn't happen since we only fill A~T, but safe)
    for r in range(1, ws.max_row + 1):
        for c in (23, 24, 25):
            if (r, c) not in wxy_data:
                ws.cell(row=r, column=c).value = None
    # Then restore original WXY
    for (r, c), (val, font, border, alignment, nf) in wxy_data.items():
        cell = ws.cell(row=r, column=c)
        cell.value = val
        cell.font = font
        cell.border = border
        cell.alignment = alignment
        cell.number_format = nf

    return ws


def get_all_projects(ws):
    """Extract all project names with their row ranges from E column."""
    projects = {}
    current_name = None
    for r in range(3, ws.max_row + 1):
        e = ws.cell(row=r, column=5).value
        u = ws.cell(row=r, column=21).value
        p = ws.cell(row=r, column=16).value
        if e and str(e).strip():
            name = str(e).strip()
            if name != current_name:
                current_name = name
                if name not in projects:
                    projects[name] = {'start': r, 'end': r}
            projects[name]['end'] = r
        elif current_name and (u or p):
            projects[current_name]['end'] = r
    return projects


def get_project_range(ws, target):
    """Find the row range of a project (by E column).
    Uses exact match first, then contains match as fallback.
    """
    t = normalize(target)

    # Pass 1: Exact match (normalized equality)
    for r in range(2, ws.max_row + 1):
        e = ws.cell(row=r, column=5).value
        if e and normalize(str(e)) == t:
            start = r
            end = start
            for rr in range(start + 1, ws.max_row + 1):
                e2 = ws.cell(row=rr, column=5).value
                p2 = ws.cell(row=rr, column=16).value
                u2 = ws.cell(row=rr, column=21).value
                if e2 and normalize(str(e2)) != t:
                    break
                if p2 is None and u2 is None and e2 is None:
                    break
                end = rr
            return (start, end)

    # Pass 2: Contains match (fallback for partial names)
    for r in range(2, ws.max_row + 1):
        e = ws.cell(row=r, column=5).value
        if e and t in normalize(str(e)):
            start = r
            end = start
            for rr in range(start + 1, ws.max_row + 1):
                e2 = ws.cell(row=rr, column=5).value
                p2 = ws.cell(row=rr, column=16).value
                u2 = ws.cell(row=rr, column=21).value
                if e2 and t not in normalize(str(e2)):
                    break
                if p2 is None and u2 is None and e2 is None:
                    break
                end = rr
            return (start, end)
    return None


def get_layers_in_project(ws, start, end):
    row_layers = []
    for r in range(start, end + 1):
        p = ws.cell(row=r, column=16).value
        row_layers.append((r, str(p).strip() if p else ""))
    layers = []
    if not row_layers: return layers
    current_name = row_layers[0][1]
    layer_start = row_layers[0][0]
    for r, name in row_layers[1:]:
        if name != current_name:
            layers.append((current_name or "优先级", layer_start, r - 1))
            current_name = name
            layer_start = r
    layers.append((current_name or "优先级", layer_start, end))
    return layers


def normalize_rate(rate_val):
    """Normalize subscription rate to decimal form (e.g. 0.018 for 1.80%).
    Booking details may store rates as:
      - Decimal (0.018) -- already correct
      - Percentage integer (1.80, 2.00) -- needs /100
    Rule: if numeric value >= 1.0, treat as percentage representation and divide by 100.
    ABS subscription rates are always in [0.5%, 5.0%] range (0.005-0.05 in decimal).
    String values like 'self-held' are returned as-is.
    """
    if rate_val is None:
        return rate_val
    if isinstance(rate_val, str):
        s = rate_val.strip().rstrip('%')
        try:
            v = float(s)
        except (ValueError, TypeError):
            return rate_val  # non-numeric string (e.g. "self-held")
        if v >= 1.0:
            return v / 100.0
        return v
    if isinstance(rate_val, (int, float)):
        v = float(rate_val)
        if v >= 1.0:
            return v / 100.0
        return v
    return rate_val


def read_detail(path, expected_layer="auto"):
    """Read booking detail from an Excel file.
    Supports single-section and multi-section formats where different layers
    (e.g. priority A1, priority A2) are separated by header rows containing
    keywords like '申购利率' in column 1.
    """
    import re
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    items = []
    current_layer = None
    for r in range(1, ws.max_row + 1):
        c1 = ws.cell(row=r, column=1).value
        c2 = ws.cell(row=r, column=2).value
        c3 = ws.cell(row=r, column=3).value
        # Detect section header: col1 contains rate keyword, col2/3 are header labels
        c1_str = str(c1).strip() if c1 else ""
        if c1_str and ('申购利率' in c1_str or '投标利率' in c1_str):
            # Parse layer name from header, e.g. "优先A1级申购利率（%）" -> "优先A1"
            layer = c1_str.split('申购利率')[0].split('投标利率')[0].strip()
            # Remove suffixes: 级, 级(%), （%） etc
            layer = re.sub(r'级.*$', '', layer).strip()
            if layer:
                current_layer = layer
            else:
                current_layer = "优先A"
            continue
        # Skip non-data rows
        if not c2 or str(c2).strip() in ['投资者', '投资者名称', '机构', '穿透机构', '']:
            continue
        # Skip project info rows (簿记日期, 管理人, 项目名称)
        if c1_str in ['项目名称', '簿记日期', '管理人', '']:
            if c1_str == '' and c2 is None:
                continue
            if c1_str in ['项目名称', '簿记日期', '管理人']:
                continue
        try:
            sf = float(c3) if c3 else 0
        except (ValueError, TypeError):
            continue
        if sf > 0 and c2:
            rate = normalize_rate(c1)
            layer = current_layer if current_layer else (expected_layer if expected_layer != "auto" else "优先A")
            items.append({'rate': rate, 'name': str(c2).strip(), 'size': sf, 'layer': layer})
    # Merge: same institution + same rate + same layer
    merged = OrderedDict()
    for it in items:
        key = (normalize(it['name']), it['rate'], it['layer'])
        if key in merged:
            merged[key]['size'] += it['size']
        else:
            merged[key] = dict(it)
    return list(merged.values())


def map_detail_to_project(filename, project_names=None):
    """Map detail filename to ledger project name.
    Strategy: try exact match first, then transformed variants.
    Returns the best-matching project name, or the transformed name as fallback.
    """
    import re
    b = os.path.basename(filename)
    # Strip common suffixes
    name = b.replace(".xlsx", "").replace(".xls", "")
    name = re.sub(r'[_-]?簿记.*$', '', name)
    name = re.sub(r'[_-]?明细.*$', '', name)
    # Strip trailing "项目" suffix (e.g. "京诚14-9项目簿记明细" -> "京诚14-9")
    name = re.sub(r'项目$', '', name)
    name = name.strip()

    # If project names are provided, try exact match first
    if project_names:
        # Exact match
        if name in project_names:
            return name
        # Normalize match (case-insensitive, strip whitespace)
        name_lower = normalize(name)
        for pn in project_names:
            if normalize(pn) == name_lower:
                return pn

    # Transform: 15号8期 -> 15-8 (ordinal pattern with period/期 after 号)
    transformed = re.sub(r'号(\d+)[期]', r'-\1', name)
    transformed = re.sub(r'第', '', transformed)
    # If transformation changed the name, check against project names
    if project_names and transformed != name:
        if transformed in project_names:
            return transformed
        t_lower = normalize(transformed)
        for pn in project_names:
            if normalize(pn) == t_lower:
                return pn

    # Return original name (not transformed) as fallback -- preserves "号" in names like "东裕9号叔裕"
    return name


# ============================================================
# Enhanced QC checks (based on pitfall analysis)
# ============================================================

def run_enhanced_qc(ws_out, ws_orig_protected, set_a, supplemented_keys,
                    target_projects, dmap, detail_map_results,
                    detail_layers_available, detail_layers_used):
    """Enhanced QC checks derived from pitfall analysis.
    Checks: #2,#3,#11,#15,#17,#20,#27,#28,#29,#30,#41,#42
    Returns (qc_fails, qc_warns) counts.
    """
    qc_fails = 0
    qc_warns = 0

    # Compute last_data_row for output
    last_data_row = 2
    for r in range(2, ws_out.max_row + 1):
        if ws_out.cell(row=r, column=5).value or ws_out.cell(row=r, column=21).value:
            last_data_row = r

    # Get all projects in output
    out_projects = get_all_projects(ws_out)

    # --- QC 7.4: W column rate range (#41) ---
    # Priority layers: strict [0.005, 0.05] (0.5%~5.0%)
    # Other layers (中间级/次级/受益权): relaxed upper bound 0.15 (15%)
    # Any value >= 1.0 is always a FAIL (normalize_rate failure)
    print("\n=== QC 7.4: W column rate range ===")
    bad_rates = []
    warn_rates = []
    for r in range(3, last_data_row + 1):
        w = ws_out.cell(row=r, column=23).value
        if w is not None and isinstance(w, (int, float)):
            v = float(w)
            p = ws_out.cell(row=r, column=16).value
            p_str = str(p).strip() if p else ''
            is_priority = '优先' in p_str
            x = ws_out.cell(row=r, column=24).value or ''
            if v >= 1.0:
                bad_rates.append((r, v, '>=1.0 (normalize_rate failed?)', x))
            elif is_priority and v > 0.05:
                bad_rates.append((r, v, f'>5% in priority layer ({p_str})', x))
            elif is_priority and v < 0.005 and v > 0:
                warn_rates.append((r, v, f'<0.5% in priority layer ({p_str})', x))
            elif not is_priority and v > 0.15:
                bad_rates.append((r, v, f'>15% in non-priority layer ({p_str})', x))
    if bad_rates:
        qc_fails += 1
        for r, v, reason, x in bad_rates:
            print(f"  FAIL: Row{r} W={v:.6f} ({reason}) X={x}")
    else:
        print("  PASS: All W column rates in valid range")
    if warn_rates:
        qc_warns += 1
        for r, v, reason, x in warn_rates:
            print(f"  WARN: Row{r} W={v:.6f} ({reason}) X={x}")

    # --- QC 7.5: Cross-layer merge check (#27) ---
    # Pitfall #27: same institution + same rate merged across layers into ONE row
    # Legitimate case: same institution same rate in different layers on SEPARATE rows (allowed)
    # Bug: two layer entries merged into a single row (WXY has combined size from both layers)
    print("\n=== QC 7.5: Cross-layer merge check ===")
    cross_layer_issues = []
    for proj_name, info in out_projects.items():
        s, e = info['start'], info['end']
        layers = get_layers_in_project(ws_out, s, e)
        for layer_name, lstart, lend in layers:
            # Within each layer, check for duplicate (normalize(X), W) on different rows
            # with DIFFERENT U values (indicating they were merged from separate U rows)
            xw_groups = defaultdict(list)
            for r in range(lstart, lend + 1):
                w = ws_out.cell(row=r, column=23).value
                x = ws_out.cell(row=r, column=24).value
                u = ws_out.cell(row=r, column=21).value
                if w is not None and x is not None:
                    key = (normalize(str(x)), w)
                    xw_groups[key].append((r, u))
            # Same (X,W) appearing more than once in the SAME layer with NO U value
            # means they were appended as unmatched items - OK
            # But same (X,W) with same U value = duplicate match = bug
            for key, rows in xw_groups.items():
                if len(rows) > 1:
                    u_vals = [u for _, u in rows if u]
                    if len(u_vals) > 1:
                        # Multiple U values with same (X,W) in same layer - likely OK (different rows)
                        pass
                    # Check if multiple rows in same layer have same (X,W) AND same U
                    u_set = set(str(u).strip() if u else '' for _, u in rows)
                    if len(u_set) < len(rows):
                        cross_layer_issues.append((proj_name, layer_name, key, rows))
    if cross_layer_issues:
        qc_fails += 1
        for proj, layer, key, rows in cross_layer_issues:
            print(f"  FAIL: {proj}/{layer} - duplicate (X,W) match: {rows}")
    else:
        print("  PASS: No cross-layer merge detected")

    # --- QC 7.6: U row duplicate match check (#20) ---
    print("\n=== QC 7.6: U row duplicate match check ===")
    dup_matches = []
    for proj_name, info in out_projects.items():
        s, e = info['start'], info['end']
        layers = get_layers_in_project(ws_out, s, e)
        for layer_name, lstart, lend in layers:
            u_wxy = defaultdict(list)
            for r in range(lstart, lend + 1):
                u = ws_out.cell(row=r, column=21).value
                w = ws_out.cell(row=r, column=23).value
                if u and w is not None:
                    u_wxy[normalize(str(u))].append((r, str(u).strip(), w))
            for un, entries in u_wxy.items():
                if len(entries) > 1:
                    rates = set(w for _, _, w in entries)
                    if len(rates) == 1:
                        dup_matches.append((proj_name, layer_name, un, entries))
    if dup_matches:
        qc_warns += 1
        for proj, layer, un, entries in dup_matches:
            print(f"  WARN: {proj}/{layer} - '{un}' matched {len(entries)} times with same rate: {[(r,w) for r,_,w in entries]}")
    else:
        print("  PASS: No duplicate U row matches with same rate")

    # --- QC 7.7: Layer order check (#3) ---
    print("\n=== QC 7.7: Layer order check ===")
    LAYER_ORDER = {'优先A': 0, '优先B': 1, '中间级': 2, '次级': 3, '受益权': 4, '优先级': 0}
    order_issues = []
    for proj_name, info in out_projects.items():
        s, e = info['start'], info['end']
        layers = get_layers_in_project(ws_out, s, e)
        prev_order = -1
        for layer_name, lstart, lend in layers:
            order = LAYER_ORDER.get(layer_name, -1)
            if order >= 0 and order < prev_order:
                order_issues.append((proj_name, layer_name, order, prev_order))
            if order >= 0:
                prev_order = order
    if order_issues:
        qc_warns += 1
        for proj, layer, order, prev in order_issues:
            print(f"  WARN: {proj} - '{layer}' (order={order}) after layer (order={prev})")
    else:
        print("  PASS: All layer orders correct")

    # --- QC 7.8: W column single value check (#2) ---
    print("\n=== QC 7.8: W column single value check ===")
    multi_values = []
    for r in range(3, last_data_row + 1):
        w = ws_out.cell(row=r, column=23).value
        if w is not None and isinstance(w, str):
            if any(sep in w for sep in [',', '\uff0c', '\n', ';', '\uff1b', '/']):
                multi_values.append((r, w[:50]))
    if multi_values:
        qc_fails += 1
        for r, w in multi_values:
            print(f"  FAIL: Row{r} W contains multiple values: '{w}'")
    else:
        print("  PASS: All W column values are single-valued")

    # --- QC 7.9: Detail mapping report (#30) ---
    print("\n=== QC 7.9: Detail mapping report ===")
    unmapped = [dp for dp, result in detail_map_results.items() if result[0] == 'unmatched']
    if unmapped:
        qc_fails += 1
        for dp in unmapped:
            print(f"  FAIL: Detail '{os.path.basename(dp)}' not mapped to any project")
    else:
        print(f"  PASS: All {len(detail_map_results)} detail files mapped successfully")

    # --- QC 7.10: Layer coverage check (#42) ---
    print("\n=== QC 7.10: Layer coverage check ===")
    coverage_issues = []
    for proj_name in supplemented_keys:
        available = detail_layers_available.get(proj_name, set())
        used = detail_layers_used.get(proj_name, set())
        unused = available - used
        if unused:
            coverage_issues.append((proj_name, unused, available, used))
    if coverage_issues:
        qc_warns += 1
        for proj, unused, available, used in coverage_issues:
            print(f"  WARN: {proj} - unused detail layers: {unused} (available: {available}, used: {used})")
    else:
        print("  PASS: All detail layers matched")

    # --- QC 7.11: X column consistency check (#28) ---
    print("\n=== QC 7.11: X column consistency check ===")
    REDUNDANT_SUFFIXES = ['(投行上报)', '(固定收益部)', '(投行)', '(资管)',
                          '\uff08投行上报\uff09', '\uff08固定收益部\uff09', '\uff08投行\uff09', '\uff08资管\uff09']
    x_issues = []
    for r in range(3, last_data_row + 1):
        u = ws_out.cell(row=r, column=21).value
        x = ws_out.cell(row=r, column=24).value
        if u and x:
            has_suffix = any(s in str(x) for s in REDUNDANT_SUFFIXES)
            if has_suffix:
                x_issues.append((r, str(u)[:30], str(x)[:30]))
    if x_issues:
        qc_warns += 1
        for r, u, x in x_issues[:10]:
            print(f"  WARN: Row{r} U='{u}' vs X='{x}' (X has redundant suffix)")
        if len(x_issues) > 10:
            print(f"  ... and {len(x_issues) - 10} more")
    else:
        print("  PASS: X column names consistent with U column")

    # --- QC 7.12: Header completeness check (#11) ---
    print("\n=== QC 7.12: Header completeness check ===")
    header_issues = []
    for c, title in [(23, '申购利率'), (24, '穿透机构'), (25, '申购规模')]:
        for hr in [1, 2]:
            v = ws_out.cell(row=hr, column=c).value
            if v is None or str(v).strip() == '':
                header_issues.append((hr, c, title))
    if header_issues:
        qc_warns += 1
        for hr, c, title in header_issues:
            print(f"  WARN: Row{hr} Col{c} ({title}) is empty")
    else:
        print("  PASS: All WXY headers present")

    # --- QC 7.13: N column number format check (#15) ---
    print("\n=== QC 7.13: N column number format check ===")
    n_format_issues = []
    for r in range(3, last_data_row + 1):
        n = ws_out.cell(row=r, column=14)
        if n.value is not None and isinstance(n.value, (int, float)):
            if n.number_format != '0.000%':
                n_format_issues.append((r, n.number_format))
    if n_format_issues:
        count = len(n_format_issues)
        qc_warns += 1
        print(f"  WARN: {count} N column cells have wrong format (expected 0.000%)")
        for r, fmt in n_format_issues[:5]:
            print(f"    Row{r}: '{fmt}'")
        if count > 5:
            print(f"    ... and {count - 5} more")
    else:
        print("  PASS: All N column numeric cells have 0.000% format")

    # --- QC 7.14: V column format protection check (#17) ---
    print("\n=== QC 7.14: V column format protection check ===")
    v_format_issues = []
    if ws_orig_protected:
        sample_rows = list(range(3, min(ws_out.max_row + 1, last_data_row + 1), 50))
        for r in sample_rows:
            v_orig = ws_orig_protected.cell(row=r, column=22)
            v_out = ws_out.cell(row=r, column=22)
            if v_orig.value is not None and v_out.value is not None:
                if v_orig.number_format != v_out.number_format:
                    v_format_issues.append((r, v_orig.number_format, v_out.number_format))
    if v_format_issues:
        qc_warns += 1
        for r, orig_fmt, out_fmt in v_format_issues[:5]:
            print(f"  WARN: Row{r} V column format changed: '{orig_fmt}' -> '{out_fmt}'")
    else:
        print("  PASS: V column formats preserved (sample check)")

    # --- QC 7.15: P column integrity check (#29) ---
    print("\n=== QC 7.15: P column integrity check ===")
    p_issues = []
    if ws_orig_protected:
        for proj_name in set_a:
            orig_range = get_project_range(ws_orig_protected, proj_name)
            out_range = get_project_range(ws_out, proj_name)
            if not orig_range or not out_range:
                continue
            orig_p = set()
            for r in range(orig_range[0], orig_range[1] + 1):
                p = ws_orig_protected.cell(row=r, column=16).value
                if p and str(p).strip():
                    orig_p.add(str(p).strip())
            out_p = set()
            for r in range(out_range[0], out_range[1] + 1):
                p = ws_out.cell(row=r, column=16).value
                if p and str(p).strip():
                    out_p.add(str(p).strip())
            if orig_p and out_p != orig_p:
                p_issues.append((proj_name, orig_p, out_p))
    if p_issues:
        qc_fails += 1
        for proj, orig, out in p_issues[:10]:
            print(f"  FAIL: {proj} P layers changed: {orig} -> {out}")
        if len(p_issues) > 10:
            print(f"  ... and {len(p_issues) - 10} more")
    else:
        print("  PASS: All P column layer values preserved")

    # --- QC 7.16: Error value scan (#N/A, #VALUE!, #REF!, etc.) ---
    print("\n=== QC 7.16: Error value scan ===")
    ERROR_VALUES = {'#N/A', '#VALUE!', '#REF!', '#DIV/0!', '#NAME?', '#NULL!', '#NUM!'}
    error_cells = []
    for r in range(3, last_data_row + 1):
        for c in range(1, 26):
            v = ws_out.cell(row=r, column=c).value
            if v is not None and isinstance(v, str) and v.strip() in ERROR_VALUES:
                error_cells.append((r, c, v.strip()))
    if error_cells:
        qc_fails += 1
        for r, c, v in error_cells[:20]:
            print(f"  FAIL: Row{r} Col{c} contains error value: {v}")
        if len(error_cells) > 20:
            print(f"  ... and {len(error_cells) - 20} more")
    else:
        print("  PASS: No error values found")

    # --- QC 7.17: Formula check in WXY columns ---
    print("\n=== QC 7.17: Formula check in WXY columns ===")
    formula_cells = []
    for r in range(3, last_data_row + 1):
        for c in (23, 24, 25):
            cell = ws_out.cell(row=r, column=c)
            if cell.value is not None and isinstance(cell.value, str) and str(cell.value).startswith('='):
                formula_cells.append((r, c, str(cell.value)[:50]))
    if formula_cells:
        qc_fails += 1
        for r, c, v in formula_cells:
            print(f"  FAIL: Row{r} Col{c} contains formula: {v}")
    else:
        print("  PASS: No formulas in WXY columns")

    # --- QC 7.18: Same-institution WXY continuity (#45 Plan C) ---
    # Same institution's multi-bid WXY rows must be row-adjacent
    print("\n=== QC 7.18: Same-institution WXY continuity ===")
    continuity_issues = []
    for proj_name, info in out_projects.items():
        s, e = info['start'], info['end']
        layers = get_layers_in_project(ws_out, s, e)
        for layer_name, lstart, lend in layers:
            # Collect rows with X value in this layer
            x_rows = defaultdict(list)  # normalize(X) -> list of row numbers
            for r in range(lstart, lend + 1):
                x = ws_out.cell(row=r, column=24).value
                w = ws_out.cell(row=r, column=23).value
                if x is not None and w is not None:
                    x_rows[normalize(str(x))].append(r)
            # Check each institution: rows must be consecutive
            for xn, rows in x_rows.items():
                if len(rows) > 1:
                    for i in range(1, len(rows)):
                        if rows[i] != rows[i-1] + 1:
                            continuity_issues.append((proj_name, layer_name, xn, rows))
                            break
    if continuity_issues:
        qc_warns += 1
        for proj, layer, xn, rows in continuity_issues:
            print(f"  WARN: {proj}/{layer} - '{xn}' WXY rows not consecutive: {rows}")
    else:
        print("  PASS: All same-institution WXY rows are consecutive")

    # --- QC 7.19: V column (ledger) vs Y column (detail) independence (#45) ---
    # V column = ledger winning size (partial bid may win)
    # Y column = detail bid size (full bid amount)
    # V != Y sum is expected, just report mismatches as INFO (not FAIL)
    print("\n=== QC 7.19: V vs Y column independence (INFO) ===")
    mismatch_count = 0
    for proj_name, info in out_projects.items():
        s, e = info['start'], info['end']
        layers = get_layers_in_project(ws_out, s, e)
        for layer_name, lstart, lend in layers:
            # Group by institution: V from U row, Y sum from all WXY rows of same institution
            inst_data = defaultdict(lambda: {'v': None, 'y_sum': 0.0, 'u_row': None})
            for r in range(lstart, lend + 1):
                u = ws_out.cell(row=r, column=21).value
                v = ws_out.cell(row=r, column=22).value
                x = ws_out.cell(row=r, column=24).value
                y = ws_out.cell(row=r, column=25).value
                if u and v is not None:
                    # V may be '-' or other non-numeric placeholder
                    try:
                        v_num = float(v) if v else 0.0
                    except (ValueError, TypeError):
                        v_num = None
                    if v_num is not None:
                        inst_data[normalize(str(u))]['v'] = v_num
                        inst_data[normalize(str(u))]['u_row'] = r
                if x and y is not None:
                    xn = normalize(str(x))
                    if inst_data[xn]['v'] is None:
                        # No U row for this institution, skip
                        pass
                    else:
                        try:
                            inst_data[xn]['y_sum'] += float(y) if y else 0.0
                        except (ValueError, TypeError):
                            pass
            for xn, d in inst_data.items():
                if d['v'] is not None and d['y_sum'] > 0:
                    if abs(d['v'] - d['y_sum']) > 0.001:
                        mismatch_count += 1
                        if mismatch_count <= 10:
                            print(f"  INFO: {proj_name}/{layer_name} Row{d['u_row']} '{xn}' V={d['v']:.3f} vs Y_sum={d['y_sum']:.3f} (diff={d['v']-d['y_sum']:.3f})")
    if mismatch_count > 0:
        print(f"  INFO: {mismatch_count} V/Y mismatches found (expected: V=winning size, Y=bid size, partial-bid scenarios cause diff)")
    else:
        print("  PASS: All V values match Y sums")

    # --- Summary ---
    print("\n=== Enhanced QC Summary ===")
    print(f"  Fails: {qc_fails}, Warns: {qc_warns}")
    if qc_fails > 0:
        print("  RESULT: FAIL - blocking issues found")
    elif qc_warns > 0:
        print("  RESULT: WARN - review warnings before using output")
    else:
        print("  RESULT: PASS - all enhanced checks passed")

    return qc_fails, qc_warns


# ============================================================
# 增量合并核心逻辑
# ============================================================

def run_increment_merge(processed_path, new_raw_path, detail_paths, output_path, supplement=False, rebook=False):
    mode_label = "Rebook WXY" if rebook else ("Supplement WXY entry" if supplement else "Increment merge")
    print("=" * 60)
    print(f"v2.1 {mode_label}")
    print("=" * 60)

    # === Step 1: Preprocess ledger(s) ===
    print("\n=== Step 1: Preprocessing ===")

    print("  Loading processed ledger (with WXY)...")
    wb_a = openpyxl.load_workbook(processed_path, data_only=False)
    ws_a = wb_a.active
    ws_a = unmerge_and_fill_processed(ws_a)
    projects_a = get_all_projects(ws_a)
    set_a = set(projects_a.keys())
    print(f"  Processed ledger: {len(set_a)} projects")

    # --- Supplement/Rebook mode: skip Steps 2-4 ---
    if supplement or rebook:
        increment = set()  # no new projects
        print(f"  [{mode_label}] Skipping new raw ledger load and diff")
    else:
        print("  Loading new raw ledger...")
        wb_b = openpyxl.load_workbook(new_raw_path, data_only=False)
        ws_b = wb_b.active
        ws_b = unmerge_and_fill_raw(ws_b)
        projects_b = get_all_projects(ws_b)
        set_b = set(projects_b.keys())
        print(f"  New raw ledger: {len(set_b)} projects")

        # Pre-check: new_raw should not have WXY data (#36)
        wxy_in_raw = 0
        for r in range(3, ws_b.max_row + 1):
            for c in (23, 24, 25):
                if ws_b.cell(row=r, column=c).value is not None:
                    wxy_in_raw += 1
                    break
            if wxy_in_raw >= 3:
                break
        if wxy_in_raw > 0:
            print(f"  !! WARNING: New raw ledger contains WXY data in {wxy_in_raw}+ rows!")
            print(f"  !! This may be an already-processed ledger, not a raw one. (#36)")

        # === Step 2: Diff projects ===
        print("\n=== Step 2: Project diff ===")
        increment = set_b - set_a
        missing = set_a - set_b  # Old projects not in new ledger

        # Preserve source file row order (avoid sorted() which breaks A-column seq)
        increment_ordered = [p for p in projects_b.keys() if p in increment]
        missing_ordered = [p for p in projects_a.keys() if p in missing]

        print(f"  Increment projects (new): {len(increment)}")
        for p in increment_ordered:
            print(f"    + {p}")
        if missing:
            print(f"  !   Missing projects (in old but not in new): {len(missing)}")
            for p in missing_ordered:
                print(f"    - {p}")

        if not increment:
            print("  No increment projects found. Nothing to merge.")
            wb_a.save(output_path)
            print(f"Saved (unchanged): {output_path}")
            return

        # === Step 3: Extract increment project rows from B ===
        print("\n=== Step 3: Extract increment rows from new ledger ===")
        increment_rows = {}  # project_name -> list of row data dicts
        for proj_name in increment_ordered:
            info = projects_b[proj_name]
            s, e = info['start'], info['end']
            rows = []
            for r in range(s, e + 1):
                row_data = {}
                for c in range(1, 26):
                    cell = ws_b.cell(row=r, column=c)
                    row_data[c] = {
                        'value': cell.value,
                        'font': copy(cell.font),
                        'border': copy(cell.border),
                        'alignment': copy(cell.alignment),
                        'number_format': cell.number_format
                    }
                rows.append(row_data)
            increment_rows[proj_name] = rows
            print(f"  {proj_name}: {len(rows)} rows (B rows {s}-{e})")

        # === Step 4: Append increment rows to A ===
        print("\n=== Step 4: Append increment rows to processed ledger ===")
        # Find last data row in A
        last_row_a = 2
        for r in range(2, ws_a.max_row + 1):
            if ws_a.cell(row=r, column=5).value or ws_a.cell(row=r, column=21).value:
                last_row_a = r

        append_start = last_row_a + 1
        current_row = append_start

        for proj_name in increment_ordered:
            rows = increment_rows[proj_name]
            for row_data in rows:
                for c in range(1, 26):
                    cell = ws_a.cell(row=current_row, column=c)
                    cell.value = row_data[c]['value']
                    cell.font = row_data[c]['font']
                    cell.border = row_data[c]['border']
                    cell.alignment = row_data[c]['alignment']
                    cell.number_format = row_data[c]['number_format']
                current_row += 1
            print(f"  Appended {proj_name}: {len(rows)} rows starting at row {append_start}")
            append_start = current_row

    # === Step 5: Bookkeeping ===
    step5_label = "rebook projects" if rebook else ("supplement projects" if supplement else "increment projects")
    print(f"\n=== Step 5: Bookkeeping for {step5_label} ===")

    # Build set of all project names for matching
    all_projects_current = get_all_projects(ws_a)
    project_name_set = set(all_projects_current.keys())

    # Map detail files to project names
    dmap = {}
    detail_map_results = {}  # dp -> ('matched', proj_name) or ('unmatched', name)
    detail_layers_available = {}  # proj_name -> set of layer names from detail
    for dp in detail_paths:
        k = map_detail_to_project(dp, project_name_set)
        # In increment mode: prefer matching against increment project names
        if not supplement and not rebook:
            matched = False
            for inc_name in increment:
                if normalize(k) in normalize(inc_name) or normalize(inc_name) in normalize(k):
                    dmap[inc_name] = read_detail(dp)
                    detail_map_results[dp] = ('matched', inc_name)
                    # Track available layers from detail
                    detail_layers_available[inc_name] = set(it.get('layer', '优先A') for it in dmap[inc_name])
                    matched = True
                    print(f"  Mapped detail '{os.path.basename(dp)}' -> '{inc_name}'")
                    break
            if not matched:
                dmap[k] = read_detail(dp)
                detail_map_results[dp] = ('unmatched', k)
                detail_layers_available[k] = set(it.get('layer', '优先A') for it in dmap[k])
                print(f"  !   Detail '{os.path.basename(dp)}' mapped to '{k}' (not in increment list, may not match)")
        else:
            # In supplement mode: map directly, but verify project exists
            dmap[k] = read_detail(dp)
            if k in project_name_set:
                detail_map_results[dp] = ('matched', k)
            else:
                detail_map_results[dp] = ('unmatched', k)
            detail_layers_available[k] = set(it.get('layer', '优先A') for it in dmap[k])
            print(f"  Mapped detail '{os.path.basename(dp)}' -> '{k}'" +
                  (" (NOT in ledger!)" if k not in project_name_set else ""))

    # Process target projects (from back to front to avoid row shift)
    target_projects = []
    for k in dmap:
        r = get_project_range(ws_a, k)
        if r:
            target_projects.append((k, r[0], r[1]))
        else:
            print(f"  !   Project '{k}' not found in ledger, skipping bookkeeping")
    target_projects.sort(key=lambda x: x[1], reverse=True)

    # === Rebook mode: clear target projects' WXY before re-entry ===
    # Step 5.0: For each target project, delete "unmatched new rows" (U=None but WXY non-empty)
    #           at layer ends, then clear remaining WXY cells
    if rebook:
        print("\n  [Rebook mode] Clearing target projects' WXY before re-entry...")
        # Process from back to front to avoid row shift during deletion
        for key, pstart, pend in target_projects:
            print(f"    Clearing {key} rows {pstart}-{pend}")
            # Find layers and identify rows to delete
            layers = get_layers_in_project(ws_a, pstart, pend)
            rows_to_delete = []  # list of row numbers (will delete from back to front)
            rows_to_clear = []   # list of row numbers (WXY cells to clear, keep row)

            for layer_name, lstart, lend in layers:
                # Within each layer, find rows where U is None but W or X or Y is non-empty
                # These are "unmatched new rows" appended by previous runs (Plan C前 bug产物)
                for r in range(lstart, lend + 1):
                    u = ws_a.cell(row=r, column=21).value
                    w = ws_a.cell(row=r, column=23).value
                    x = ws_a.cell(row=r, column=24).value
                    y = ws_a.cell(row=r, column=25).value
                    wxy_has_data = (w is not None or x is not None or y is not None)
                    if u is None and wxy_has_data:
                        rows_to_delete.append(r)
                    elif u is not None and wxy_has_data:
                        # U row with WXY (matched row from previous run) - clear WXY, keep U/V
                        rows_to_clear.append(r)

            # Delete rows from back to front
            for r in sorted(rows_to_delete, reverse=True):
                ws_a.delete_rows(r, 1)
                print(f"      Deleted Row{r} (unmatched new row)")
            # Clear WXY cells on remaining U rows
            for r in rows_to_clear:
                # Note: row numbers may have shifted after deletion, need to recompute
                # But since we only delete rows BELOW the U row (U rows are at top of each layer),
                # the U row numbers themselves are stable. Actually no - if a U row's layer
                # had unmatched rows above it deleted... wait, we delete from back to front
                # and U rows are at top of layer, so U rows shift only if a previous layer's
                # unmatched rows were deleted. Let's just clear by scanning again after deletion.
                pass  # Handle in re-scan below

            # Re-scan and clear WXY on remaining U rows (row numbers may have shifted)
            # Re-find project range since rows were deleted
            r_new = get_project_range(ws_a, key)
            if r_new:
                pstart_new, pend_new = r_new
                layers_new = get_layers_in_project(ws_a, pstart_new, pend_new)
                cleared_count = 0
                for layer_name, lstart, lend in layers_new:
                    for r in range(lstart, lend + 1):
                        u = ws_a.cell(row=r, column=21).value
                        if u is not None:
                            # Clear WXY (columns 23, 24, 25)
                            ws_a.cell(row=r, column=23).value = None
                            ws_a.cell(row=r, column=24).value = None
                            ws_a.cell(row=r, column=25).value = None
                            cleared_count += 1
                print(f"      Cleared WXY on {cleared_count} U rows")

        # Re-fetch target_projects with updated row ranges after deletion
        target_projects = []
        for k in dmap:
            r = get_project_range(ws_a, k)
            if r:
                target_projects.append((k, r[0], r[1]))
        target_projects.sort(key=lambda x: x[1], reverse=True)
        print(f"  [Rebook mode] Clearing complete. Re-entry with Plan C logic...\n")

    detail_layers_used = {}  # proj_name -> set of layer names that were matched
    for key, pstart, pend in target_projects:
        print(f"\n  Processing {key} rows {pstart}-{pend}")
        layers = get_layers_in_project(ws_a, pstart, pend)

        # Group detail items by their layer
        items = dmap[key]
        detail_by_layer = OrderedDict()
        for it in items:
            dl = it.get('layer', '优先A')
            if dl not in detail_by_layer:
                detail_by_layer[dl] = []
            detail_by_layer[dl].append(it)

        # Process each layer that has matching details
        # Collect all new_items with their target insert position
        all_inserts = []  # list of (insert_row, layer_lend, items)

        for layer_name, lstart, lend in layers:
            # Find matching detail layer
            detail_items = None
            # Exact match first (e.g. "优先A1" == "优先A1")
            if layer_name in detail_by_layer:
                detail_items = detail_by_layer.pop(layer_name)
            # Fuzzy: detail "优先A" matches ledger "优先A1" if no specific match
            elif '优先A' in detail_by_layer and '优先' in layer_name:
                # Only use "优先A" fallback if no more specific match exists
                has_specific = any(layer_name in k or k in layer_name for k in detail_by_layer if k != '优先A')
                if not has_specific:
                    detail_items = detail_by_layer.pop('优先A')
            # Contains match
            else:
                for dk in list(detail_by_layer.keys()):
                    if dk in layer_name or layer_name in dk:
                        detail_items = detail_by_layer.pop(dk)
                        break

            if detail_items is None:
                continue

            # Track which layers were actually used
            if key not in detail_layers_used:
                detail_layers_used[key] = set()
            detail_layers_used[key].add(layer_name)

            print(f"    Target layer: {layer_name} rows {lstart}-{lend}")

            urows = []
            for r in range(lstart, lend + 1):
                u = ws_a.cell(row=r, column=21).value
                v = ws_a.cell(row=r, column=22).value
                if u:
                    urows.append([r, str(u).strip(), float(v) if v else 0.0, False])

            print(f"      U rows: {len(urows)}, Details: {len(detail_items)}")

            header = {}
            for c in range(1, 21):
                header[c] = ws_a.cell(row=lstart, column=c).value

            # === Plan C: same-institution multi-bid continuity ===
            # Group details by institution (preserve detail order)
            # First bid of each institution matches the U row (writes WXY into U row)
            # Subsequent bids of same institution insert right below the U row
            # Institutions not in ledger append to layer end
            detail_by_inst = OrderedDict()
            for it in detail_items:
                key = normalize(it['name'])
                if key not in detail_by_inst:
                    detail_by_inst[key] = []
                detail_by_inst[key].append(it)

            # Build U-row index: normalize(name) -> list of urow indices
            inst_to_urows = defaultdict(list)
            for idx, ur in enumerate(urows):
                inst_to_urows[normalize(ur[1])].append(idx)

            new_items = []  # list of (detail_item, anchor_row)
            for inst_key, items in detail_by_inst.items():
                matched_ur_idx = None
                # Find first unmatched U row for this institution
                for ur_idx in inst_to_urows.get(inst_key, []):
                    if not urows[ur_idx][3]:
                        matched_ur_idx = ur_idx
                        break
                # Fuzzy match fallback (handles name variants)
                if matched_ur_idx is None:
                    for ur_idx, ur in enumerate(urows):
                        if ur[3]:
                            continue
                        un = normalize(ur[1])
                        if (len(inst_key) >= 4 and len(un) >= 4 and
                            (inst_key in un or un in inst_key)):
                            matched_ur_idx = ur_idx
                            break

                if matched_ur_idx is not None:
                    ur = urows[matched_ur_idx]
                    # First bid writes WXY into U row
                    first = items[0]
                    ws_a.cell(row=ur[0], column=23).value = first['rate']
                    ws_a.cell(row=ur[0], column=24).value = first['name']
                    ws_a.cell(row=ur[0], column=25).value = first['size']
                    ur[3] = True
                    print(f"      MATCH: {first['name']} -> Row{ur[0]} (rate={first['rate']}, size={first['size']})")
                    # Subsequent bids anchor below this U row
                    for extra in items[1:]:
                        new_items.append((extra, ur[0]))
                        print(f"      EXTRA: {extra['name']} rate={extra['rate']} size={extra['size']} (anchor Row{ur[0]})")
                else:
                    # Institution not in ledger, append to layer end (anchor=lend, insert_at=lend+1)
                    for it in items:
                        new_items.append((it, lend))

            if new_items:
                all_inserts.append((new_items, header, layer_name))

        # Insert new rows grouped by anchor row, from back to front
        # Same anchor_row = same insert position; items keep detail order
        inserts_by_anchor = defaultdict(list)  # anchor_row -> list of (item, header, layer_name)
        for new_items, header, layer_name in all_inserts:
            for it, anchor_row in new_items:
                inserts_by_anchor[anchor_row].append((it, header, layer_name))

        for anchor_row in sorted(inserts_by_anchor.keys(), reverse=True):
            items_with_header = inserts_by_anchor[anchor_row]
            insert_at = anchor_row + 1
            ws_a.insert_rows(insert_at, len(items_with_header))
            for idx, (it, header, layer_name) in enumerate(items_with_header):
                nr = insert_at + idx
                for c in range(1, 21):
                    ws_a.cell(row=nr, column=c).value = header.get(c)
                ws_a.cell(row=nr, column=23).value = it['rate']
                ws_a.cell(row=nr, column=24).value = it['name']
                ws_a.cell(row=nr, column=25).value = it['size']
                print(f"    NEW: Row{nr} {it['name']} rate={it['rate']} size={it['size']} (anchor Row{anchor_row}, layer={layer_name})")

    # === Step 6: Format optimization + QC ===
    print("\n=== Step 6: Format optimization ===")
    from openpyxl.styles import Font, Alignment, Border, Side

    thin_side = Side(style='thin', color='000000')
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    font_header = Font(name='华文楷体', size=11, bold=True)
    font_data = Font(name='华文楷体', size=11, bold=False)
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    last_data_row = 2
    for r in range(2, ws_a.max_row + 1):
        if ws_a.cell(row=r, column=5).value or ws_a.cell(row=r, column=21).value:
            last_data_row = r

    fmt_count = 0
    for r in range(3, last_data_row + 1):
        has_data = any(ws_a.cell(row=r, column=c).value for c in range(1, 26))
        if not has_data: continue
        fmt_count += 1
        for c in range(1, 26):
            cell = ws_a.cell(row=r, column=c)
            cell.border = thin_border
            if cell.font.name != 'KaiTi':
                cell.font = font_data
            cell.alignment = align_center
        # Column-specific formats
        for c, fmt in [(12, 'yyyy-mm-dd'), (13, '0.00%'), (14, '0.000%'),
                       (17, '0.00%'), (20, '0.00%')]:
            cc = ws_a.cell(row=r, column=c)
            if cc.value is not None and isinstance(cc.value, (int, float)):
                cc.number_format = fmt
        rc = ws_a.cell(row=r, column=18)
        if rc.value is not None and isinstance(rc.value, (int, float)):
            rc.number_format = '0.00_);[Red]\\(0.00\\)'
        wc = ws_a.cell(row=r, column=23)
        if wc.value is not None and isinstance(wc.value, (int, float)):
            wc.number_format = '0.00%'
        yc = ws_a.cell(row=r, column=25)
        if yc.value is not None and isinstance(yc.value, (int, float)):
            yc.number_format = '0.00'
    print(f"  Formatted {fmt_count} data rows")

    # Header format
    for hr in range(1, 3):
        for c in range(1, 26):
            cell = ws_a.cell(row=hr, column=c)
            cell.font = font_header
            cell.alignment = align_center
            cell.border = thin_border
    # WXY headers
    for c, title in [(23, '申购利率'), (24, '穿透机构'), (25, '申购规模')]:
        for hr in [1, 2]:
            cell = ws_a.cell(row=hr, column=c)
            if cell.value is None:
                cell.value = title

    wb_a.save(output_path)
    print(f"\nSaved: {output_path}")

    # === QC ===
    wb_out = openpyxl.load_workbook(output_path, data_only=False)
    ws_out = wb_out.active

    # QC 7.1: Target project bookkeeping completeness
    # Only count rows where WXY was written (not original U/V rows without WXY)
    print("\n=== QC 7.1: Target project amount ===")
    # Track which projects were supplemented (had WXY added)
    supplemented_keys = set(k for k, _, _ in target_projects)
    for key, _, _ in target_projects:
        r = get_project_range(ws_out, key)
        if not r: continue
        s, e = r
        ty = 0.0
        for rr in range(s, e + 1):
            w = ws_out.cell(row=rr, column=23).value
            x = ws_out.cell(row=rr, column=24).value
            if w is not None or x is not None:
                ty += float(ws_out.cell(row=rr, column=25).value or 0)
        td = sum(it['size'] for it in dmap.get(key, []))
        diff = abs(ty - td)
        status = 'PASS' if diff < 0.01 else 'FAIL'
        print(f"  {key}: WXY_Y={ty:.3f} detail={td:.3f} diff={diff:.3f} {status}")


    # QC 7.2: WXY preservation for non-target projects
    print("\n=== QC 7.2: Existing WXY preservation ===")
    # CRITICAL: Must preprocess the original to expand merged cells for WXY comparison
    ws_orig_protected = unmerge_and_fill_processed(openpyxl.load_workbook(processed_path, data_only=False).active)
    wxy_mismatches = 0
    checked_count = 0
    for proj_name in set_a:
        # Skip projects that were targets of bookkeeping (their WXY is expected to change)
        if proj_name in supplemented_keys:
            continue
        orig_range = get_project_range(ws_orig_protected, proj_name)
        if not orig_range:
            continue
        orig_wxy = []
        for r in range(orig_range[0], orig_range[1] + 1):
            w = ws_orig_protected.cell(row=r, column=23).value
            x = ws_orig_protected.cell(row=r, column=24).value
            y = ws_orig_protected.cell(row=r, column=25).value
            if w is not None or x is not None or y is not None:
                orig_wxy.append((w, x, y))

        out_range = get_project_range(ws_out, proj_name)
        if not out_range:
            continue
        out_wxy = []
        for r in range(out_range[0], out_range[1] + 1):
            w = ws_out.cell(row=r, column=23).value
            x = ws_out.cell(row=r, column=24).value
            y = ws_out.cell(row=r, column=25).value
            if w is not None or x is not None or y is not None:
                out_wxy.append((w, x, y))

        checked_count += 1
        if orig_wxy != out_wxy:
            wxy_mismatches += 1
            print(f"  !  {proj_name}: mismatch orig={len(orig_wxy)} out={len(out_wxy)}")
            for i, o in enumerate(orig_wxy[:3]):
                if i < len(out_wxy) and o != out_wxy[i]:
                    print(f"    r{i}: {o} -> {out_wxy[i]}")

    if wxy_mismatches == 0:
        print(f"  PASS: All {checked_count} non-target projects WXY preserved")
    else:
        print(f"  FAIL: {wxy_mismatches} projects WXY changed")

    # QC 7.3: Target projects present in output
    print("\n=== QC 7.3: Target projects present ===")
    check_set = increment if not supplement else supplemented_keys
    for proj_name in sorted(check_set):
        r = get_project_range(ws_out, proj_name)
        if r:
            print(f"  {proj_name}: rows {r[0]}-{r[1]} [OK]")
        else:
            print(f"  {proj_name}: NOT FOUND [X]")

    # QC 7.4-7.15: Enhanced QC checks
    run_enhanced_qc(ws_out, ws_orig_protected, set_a, supplemented_keys,
                    target_projects, dmap, detail_map_results,
                    detail_layers_available, detail_layers_used)

    print("\n" + "=" * 60)
    print(f"{mode_label} complete.")
    print("=" * 60)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="增量台账合并 v2.1")
    parser.add_argument("--processed", required=True, help="已加工台账（含WXY）")
    parser.add_argument("--new-raw", required=False, help="本周新原始台账（无WXY），增量合并模式必填")
    parser.add_argument("--details", nargs="*", default=[], help="簿记明细文件")
    parser.add_argument("--output", required=True, help="输出文件路径")
    parser.add_argument("--supplement", action="store_true", help="补充簿记模式：仅对有明细的项目录入WXY，不做增量合并")
    parser.add_argument("--rebook", action="store_true", help="重录模式：清空目标项目WXY+删除分层末尾未匹配新行，然后重新录入（用于修复方案C前的旧WXY数据）")
    args = parser.parse_args()

    if not args.supplement and not args.rebook and not args.new_raw:
        parser.error("--new-raw is required when not using --supplement or --rebook mode")
    if (args.supplement or args.rebook) and args.new_raw:
        parser.error("--new-raw is not allowed in --supplement or --rebook mode")
    if args.supplement and args.rebook:
        parser.error("--supplement and --rebook are mutually exclusive")

    run_increment_merge(args.processed, args.new_raw, args.details, args.output,
                        supplement=args.supplement, rebook=args.rebook)
