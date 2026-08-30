"""白条/金条原始 Excel 的消金资产实验面板生成器。"""

from __future__ import annotations

import argparse
import json
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from consumer_asset_mapping import (
    CATEGORY_COLORS,
    CASH_FUNDING_RULE,
    CASH_LOAN_ASSETS,
    CASH_LOAN_FUNDING_ORDER,
    CASH_LOAN_RULE,
    CONSUMER_LOAN_ASSETS,
    CONSUMER_LOAN_FUNDING_ORDER,
    CONSUMER_LOAN_RULE,
    GOLD_BALANCE_SHEET,
    GOLD_FUNDING_ORDER,
    PANEL_SUMMARY_RULE,
    TOTAL_ASSET_RULE,
    WHITE_FUNDING_ORDER,
    WHITE_FUNDING_SHEET,
    WHITE_SECTION_LABELS,
    WHITE_SECTIONS,
)

ROOT = Path(__file__).resolve().parents[2]
DEFAULT_OUTPUT_DIR = ROOT / "deliverables" / "dashboards" / "02_history" / "lab_viz"
ZERO = Decimal("0")


class PanelInputError(ValueError):
    """无法可靠生成面板时抛出。"""


def as_decimal(value: Any) -> Decimal:
    if value is None or str(value).strip() in {"", "-", "—"}:
        return ZERO
    try:
        return Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, AttributeError) as exc:
        raise PanelInputError(f"非法金额：{value!r}") from exc


def as_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                pass
    return None


def issue(items: list[dict], level: str, code: str, message: str, **detail: Any) -> None:
    items.append({"level": level, "code": code, "message": message, **detail})


def open_sheet(path: Path, sheet_name: str):
    if not path.exists():
        raise PanelInputError(f"输入文件不存在：{path}")
    workbook = load_workbook(path, data_only=True, read_only=False)
    if sheet_name not in workbook.sheetnames:
        raise PanelInputError(f"{path.name} 缺少必需工作表：{sheet_name}")
    return workbook[sheet_name]


def parse_white(path: Path, qc: list[dict]) -> tuple[dict[date, dict[str, dict[str, Decimal]]], dict[date, dict[str, Decimal]]]:
    ws = open_sheet(path, WHITE_FUNDING_SHEET)
    # 白条 Sheet 为“日期纵向、资金类型横向”的分组表；日期位于 B 列。
    snapshots: dict[date, dict[str, dict[str, Decimal]]] = {}
    reported_totals: dict[date, dict[str, Decimal]] = {}
    current_section = None
    for row in range(4, ws.max_row + 1):
        section = ws.cell(row, 1).value or current_section
        if section in WHITE_SECTIONS:
            current_section = section
        else:
            continue
        row_date = as_date(ws.cell(row, 2).value)
        # 当前输入的 Sheet 是按行放日期、列放资金类型；这里按 row 解析。
        stat_date = row_date
        if not stat_date:
            continue
        if stat_date not in snapshots:
            snapshots[stat_date] = {item: {} for item in WHITE_SECTIONS}
            reported_totals[stat_date] = {}
        for col, funding in enumerate(WHITE_FUNDING_ORDER, start=3):
            raw_value = ws.cell(row, col).value
            # 分分卡、白取的赊销列为空，不将“空”伪造成展示用的零余额类型。
            if raw_value is not None and str(raw_value).strip() not in {"", "-", "—"}:
                snapshots[stat_date][section][funding] = as_decimal(raw_value)
        reported_total = ws.cell(row, 8).value
        if reported_total is not None and str(reported_total).strip() not in {"", "-", "—"}:
            reported_totals[stat_date][section] = as_decimal(reported_total)
    if not snapshots:
        raise PanelInputError(f"{WHITE_FUNDING_SHEET} 未读取到有效余额记录")
    return snapshots, reported_totals


def parse_gold_balance(path: Path) -> tuple[dict[date, dict[str, Decimal]], dict[date, Decimal]]:
    ws = open_sheet(path, GOLD_BALANCE_SHEET)
    labels = {str(ws.cell(2, col).value).strip(): col for col in range(2, ws.max_column + 1)}
    missing = set(GOLD_FUNDING_ORDER) - set(labels)
    if missing:
        raise PanelInputError(f"{GOLD_BALANCE_SHEET} 缺少资金类型列：{sorted(missing)}")
    total_col = labels.get("汇总")
    records, reported_totals = {}, {}
    for row in range(4, ws.max_row + 1):
        stat_date = as_date(ws.cell(row, 1).value)
        if stat_date:
            records[stat_date] = {label: as_decimal(ws.cell(row, labels[label]).value) for label in GOLD_FUNDING_ORDER}
            if total_col:
                reported_totals[stat_date] = as_decimal(ws.cell(row, total_col).value)
    if not records:
        raise PanelInputError(f"{GOLD_BALANCE_SHEET} 未读取到有效余额记录")
    return records, reported_totals


def choose_dates(records: dict[date, Any], domain: str, qc: list[dict]) -> tuple[date, date]:
    available = sorted(records)
    if not available:
        raise PanelInputError(f"{domain}没有可用统计日")
    current = available[-1]
    previous = date.fromordinal(current.toordinal() - 5)
    if previous not in records:
        raise PanelInputError(
            f"{domain}缺少五天前可比数据：最新统计日 {current.isoformat()}，"
            f"要求 {previous.isoformat()}"
        )
    return current, previous


def metric(name: str, current: Decimal, previous: Decimal | None, source: str = "") -> dict[str, Any]:
    delta = None if previous is None else current - previous
    if previous is None:
        rate, status = None, "无可比上期"
    elif previous == ZERO and current != ZERO:
        rate, status = None, "新增"
    elif previous == ZERO:
        rate, status = ZERO, "持平"
    else:
        rate = delta / previous * Decimal("100")
        status = "增长" if delta > ZERO else "下降" if delta < ZERO else "持平"
    return {"name": name, "amount_yuan": current, "previous_yuan": previous, "change_yuan": delta,
            "change_pct": rate, "status": status, "source": source}


def sum_items(values: dict[str, Decimal], labels: tuple[str, ...]) -> Decimal:
    return sum((values.get(label, ZERO) for label in labels), ZERO)


def check_total(qc: list[dict], name: str, displayed: Decimal, reported: Decimal | None, locator: str) -> None:
    if reported is None:
        return
    difference = displayed - reported
    level = "OK" if abs(difference) <= Decimal("2") else "WARN"
    issue(qc, level, "TOTAL_RECONCILIATION", f"{name}明细与来源汇总差额 {difference} 元",
          displayed_yuan=str(displayed), reported_yuan=str(reported), difference_yuan=str(difference), source_locator=locator)


def build_dashboard(baitiao_path: Path, jintiao_path: Path) -> tuple[dict[str, Any], list[dict]]:
    qc: list[dict] = []
    white, white_reported_totals = parse_white(baitiao_path, qc)
    gold, gold_reported_totals = parse_gold_balance(jintiao_path)
    white_current, white_previous = choose_dates(white, "白条", qc)
    gold_current, gold_previous = choose_dates(gold, "金条", qc)

    def white_section(section: str) -> list[dict]:
        current = white[white_current].get(section, {})
        previous = white[white_previous].get(section, {})
        if not current:
            issue(qc, "WARN", "WHITE_SECTION_UNAVAILABLE", f"白条来源未提供{WHITE_SECTION_LABELS[section]}明细")
        return [metric(label, current.get(label, ZERO), previous.get(label, ZERO),
                       f"{WHITE_FUNDING_SHEET}/{section}/{label}")
                for label in WHITE_FUNDING_ORDER if label in current or label in previous]

    def total_item(name: str, items: list[dict], source: str = "") -> dict:
        return metric(name, sum((item["amount_yuan"] for item in items), ZERO),
                      sum((item["previous_yuan"] or ZERO for item in items), ZERO), source)

    def sum_metric(name: str, items: list[dict], source: str = "") -> dict:
        return metric(name, sum((item["amount_yuan"] for item in items), ZERO),
                      sum((item["previous_yuan"] or ZERO for item in items), ZERO), source)

    white_consume = white_section("白条消费")
    fenfenka = white_section("分分卡")
    cashout = white_section("取现")
    white_consume_total = total_item("白条消费", white_consume)
    fenfenka_total = total_item("分分卡", fenfenka)
    cashout_total = total_item("白取", cashout)
    check_total(qc, "白条消费", white_consume_total["amount_yuan"], white_reported_totals.get(white_current, {}).get("白条消费"),
                f"{WHITE_FUNDING_SHEET}/白条消费/汇总")
    check_total(qc, "分分卡", fenfenka_total["amount_yuan"], white_reported_totals.get(white_current, {}).get("分分卡"),
                f"{WHITE_FUNDING_SHEET}/分分卡/汇总")
    check_total(qc, "白取", cashout_total["amount_yuan"], white_reported_totals.get(white_current, {}).get("取现"),
                f"{WHITE_FUNDING_SHEET}/取现/汇总")

    gold_metrics = [metric(label, gold[gold_current][label], gold[gold_previous][label],
                           f"{GOLD_BALANCE_SHEET}/{label}") for label in GOLD_FUNDING_ORDER]
    gold_total = total_item("金条", gold_metrics, GOLD_BALANCE_SHEET)
    check_total(qc, "金条结构", gold_total["amount_yuan"], gold_reported_totals.get(gold_current), GOLD_BALANCE_SHEET)

    consumer_assets = [white_consume_total, fenfenka_total]
    consumer_funding = []
    for label in CONSUMER_LOAN_FUNDING_ORDER:
        source_labels = {"助贷100%", "助贷联合贷"} if label == "助贷合计" else {label}
        parts = [item for item in white_consume + fenfenka if item["name"] in source_labels]
        consumer_funding.append(sum_metric(label, parts, "白条消费 + 分分卡"))
    consumer_total = total_item("消费贷", consumer_assets, CONSUMER_LOAN_RULE)

    cash_assets = [gold_total, cashout_total]
    cash_funding = [
        sum_metric("信托", [item for item in gold_metrics + cashout if item["name"] == "信托"], "金条信托 + 白取信托"),
        sum_metric("小贷", [item for item in gold_metrics + cashout if item["name"] == "小贷"], "金条小贷 + 白取小贷"),
        sum_metric("助贷合计", [item for item in gold_metrics + cashout if item["name"] in {"助贷", "助贷100%", "助贷联合贷"}], CASH_FUNDING_RULE),
    ]
    cash_total = total_item("现金贷", cash_assets, CASH_LOAN_RULE)
    total_assets = [consumer_total, cash_total]
    consumer_finance_total = total_item("消金资产合计", total_assets, TOTAL_ASSET_RULE)

    for category, assets, funding, total in (("消费贷", consumer_assets, consumer_funding, consumer_total),
                                             ("现金贷", cash_assets, cash_funding, cash_total)):
        check_total(qc, f"{category}资产类型", sum((item["amount_yuan"] for item in assets), ZERO), total["amount_yuan"], f"{category}/资产类型")
        check_total(qc, f"{category}资金类型", sum((item["amount_yuan"] for item in funding), ZERO), total["amount_yuan"], f"{category}/资金类型")

    data = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "sources": {"baitiao": baitiao_path.name, "jintiao": jintiao_path.name},
        "dates": {"baitiao": {"current": white_current, "previous": white_previous}, "jintiao": {"current": gold_current, "previous": gold_previous}},
        "consumer_loan": {"assets": consumer_assets, "funding": consumer_funding, "total": consumer_total},
        "cash_loan": {"assets": cash_assets, "funding": cash_funding, "total": cash_total},
        "consumer_finance_total": {"assets": total_assets, "total": consumer_finance_total},
    }
    return data, qc


def billion(value: Decimal | None) -> str:
    return "—" if value is None else f"{value / Decimal('100000000'):,.0f}"


def percent(value: Decimal | None) -> str:
    return "—" if value is None else f"{value:+.2f}%"


def json_default(value: Any):
    if isinstance(value, (Decimal, date, datetime)):
        return str(value)
    raise TypeError(type(value).__name__)


def change_text(item: dict[str, Any]) -> str:
    if item["change_yuan"] is None:
        return "无可比上期"
    if item["status"] == "新增":
        return f"新增 {billion(item['change_yuan'])} 亿元"
    return f"{item['status']} {billion(item['change_yuan'])} 亿元（{percent(item['change_pct'])}）"


def bar_rows(items: list[dict[str, Any]], total: Decimal, asset_structure: bool = False) -> str:
    rows = []
    for item in sorted(items, key=lambda value: value["amount_yuan"], reverse=True):
        share = ZERO if total == ZERO else item["amount_yuan"] / total * Decimal("100")
        color = CATEGORY_COLORS.get(item["name"], "#2a78d6")
        state = "up" if item["status"] in {"增长", "新增"} else "down" if item["status"] == "下降" else "flat"
        rows.append(f'''<div class="consumer-asset-bar-row"><div class="consumer-asset-label">{item['name']}</div>
<div class="consumer-asset-track"><span style="width:{float(share):.2f}%;background:{color}"></span></div>
<div class="consumer-asset-value"><span>{billion(item['amount_yuan'])} 亿 · {share:.0f}%</span><small class="consumer-asset-delta {state}">{change_text(item)}</small></div></div>''')
    return "".join(rows)


def metric_card(item: dict[str, Any], note: str = "") -> str:
    state = "up" if item["status"] in {"增长", "新增"} else "down" if item["status"] == "下降" else "flat"
    return f'''<article class="consumer-asset-kpi"><p>{item['name']}</p><strong>{billion(item['amount_yuan'])}<em>亿</em></strong>
<span class="consumer-asset-delta {state}">{change_text(item)}</span><small>{note}</small></article>'''


CONSUMER_ASSET_CSS = """
.consumer-asset-root{--ink:#0d1b2e;--paper:#FDFBF7;--muted:#666;font-family:-apple-system,"PingFang SC","Microsoft YaHei",sans-serif;color:#0b0b0b;background:#f4f5f7;min-height:100vh;padding:24px;box-sizing:border-box}.consumer-asset-shell{max-width:1420px;margin:auto}.consumer-asset-hero{background:linear-gradient(135deg,#1a3a5c,#0d1b2e);color:#fff;padding:24px 28px;border-radius:10px 10px 0 0}.consumer-asset-hero h1{margin:0;font-size:25px}.consumer-asset-hero p{margin:7px 0 0;color:#dce7f2;font-size:13px}.consumer-asset-date-grid{display:flex;gap:8px;flex-wrap:wrap;margin-top:16px}.consumer-asset-date{background:rgba(255,255,255,.13);padding:7px 10px;border-radius:5px;font-size:12px}.consumer-asset-kpis{display:grid;grid-template-columns:repeat(3,1fr);gap:12px;margin:16px 0}.consumer-asset-kpi,.consumer-asset-section{background:var(--paper);border:1px solid rgba(11,11,11,.1);border-radius:8px;box-shadow:0 1px 3px rgba(0,0,0,.05)}.consumer-asset-kpi{padding:17px}.consumer-asset-kpi p{color:#52514e;font-size:13px;margin:0 0 8px}.consumer-asset-kpi strong{font-size:27px;line-height:1;font-variant-numeric:tabular-nums}.consumer-asset-kpi em{font-style:normal;font-size:13px;margin-left:4px;font-weight:500}.consumer-asset-kpi small{display:block;color:#777;margin-top:8px;font-size:11px}.consumer-asset-delta{display:block;font-size:12px;margin-top:9px;font-weight:600}.consumer-asset-delta.up{color:#006300}.consumer-asset-delta.down{color:#d03b3b}.consumer-asset-delta.flat{color:#52514e}.consumer-asset-grid{display:grid;grid-template-columns:minmax(0,1fr) minmax(0,1fr);gap:16px}.consumer-asset-section{padding:18px;margin-bottom:16px;min-width:0}.consumer-asset-section h2{font-size:16px;color:var(--ink);margin:0 0 4px}.consumer-asset-section h3{font-size:15px;color:#52514e;margin:16px 0 7px}.consumer-asset-section h3 + .consumer-asset-bar-row{margin-top:7px}.consumer-asset-bar-row + h3{margin-top:32px}.consumer-asset-section>p{margin:0 0 15px;font-size:12px;color:var(--muted)}.consumer-asset-bar-row{display:grid;grid-template-columns:92px 1fr 182px;gap:10px;align-items:center;margin:11px 0}.consumer-asset-label{font-size:13px;font-weight:600;color:#52514e}.consumer-asset-track{height:11px;border-radius:7px;background:#edf0f2;overflow:hidden}.consumer-asset-track span{height:100%;display:block;border-radius:7px;min-width:2px}.consumer-asset-value{display:flex;flex-direction:column;align-items:flex-end;font-size:14px;line-height:1.12;text-align:right;font-variant-numeric:tabular-nums}.consumer-asset-value small{font-size:10px;line-height:1.15;margin-top:1px}.consumer-asset-foot{font-size:12px;line-height:1.4;color:#52514e}.consumer-asset-foot p{margin:2px 0}.consumer-asset-foot h2{margin-bottom:5px}@media(max-width:900px){.consumer-asset-kpis,.consumer-asset-grid{grid-template-columns:1fr 1fr}}@media(max-width:620px){.consumer-asset-root{padding:12px}.consumer-asset-kpis,.consumer-asset-grid{grid-template-columns:1fr}.consumer-asset-bar-row{grid-template-columns:76px 1fr 125px}.consumer-asset-hero{padding:20px}}
"""


def render_body(data: dict[str, Any]) -> str:
    white_dates, gold_dates = data["dates"]["baitiao"], data["dates"]["jintiao"]
    date_caption = lambda values: "统计日 {} · 对比 {}".format(values["current"].isoformat(), values["previous"].isoformat())
    return """<main class="consumer-asset-root"><div class="consumer-asset-shell"><header class="consumer-asset-hero"><h1>消金资产面板</h1><p>按消费贷与现金贷分类展示资产类型、资金类型结构及五天变化</p><div class="consumer-asset-date-grid"><span class="consumer-asset-date">白条消费 / 分分卡 / 白取：{white_date}</span><span class="consumer-asset-date">金条：{gold_date}</span></div></header><section class="consumer-asset-kpis">{consumer_kpi}{cash_kpi}{total_kpi}</section><div class="consumer-asset-grid"><section class="consumer-asset-section"><h2>消费贷资产及资金结构</h2><p>消费贷 = 白条消费 + 分分卡；均按当前规模降序</p><h3>资产类型结构</h3>{consumer_assets}<h3>资金类型结构</h3>{consumer_funding}</section><section class="consumer-asset-section"><h2>现金贷资产及资金结构</h2><p>现金贷 = 金条 + 白取；均按当前规模降序</p><h3>资产类型结构</h3>{cash_assets}<h3>资金类型结构</h3>{cash_funding}</section></div><section class="consumer-asset-section consumer-asset-foot"><h2>数据口径</h2><p>{consumer_rule} {cash_rule} {total_rule}</p><p>{funding_rule}</p><p>白条侧、金条侧均严格对比各自最新统计日前第 5 天；现金贷及消金资产合计为按来源最新可用统计日汇总的混合统计日口径。</p></section></div></main>""".format(
        white_date=date_caption(white_dates), gold_date=date_caption(gold_dates),
        consumer_kpi=metric_card(data["consumer_loan"]["total"], "白条侧统计日"),
        cash_kpi=metric_card(data["cash_loan"]["total"], "混合统计日：金条 + 白取"),
        total_kpi=metric_card(data["consumer_finance_total"]["total"], "混合统计日：消费贷 + 现金贷"),
        consumer_assets=bar_rows(data["consumer_loan"]["assets"], data["consumer_loan"]["total"]["amount_yuan"]),
        consumer_funding=bar_rows(data["consumer_loan"]["funding"], data["consumer_loan"]["total"]["amount_yuan"]),
        cash_assets=bar_rows(data["cash_loan"]["assets"], data["cash_loan"]["total"]["amount_yuan"]),
        cash_funding=bar_rows(data["cash_loan"]["funding"], data["cash_loan"]["total"]["amount_yuan"]),
        consumer_rule=CONSUMER_LOAN_RULE, cash_rule=CASH_LOAN_RULE, total_rule=TOTAL_ASSET_RULE, funding_rule=CASH_FUNDING_RULE)


def render_html(data: dict[str, Any]) -> str:
    return '<!doctype html><html lang="zh-CN"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>消金资产面板</title><style>' + CONSUMER_ASSET_CSS + '</style></head><body>' + render_body(data) + '</body></html>'


def compute_data(baitiao_path: str | Path, jintiao_path: str | Path) -> tuple[dict[str, Any], list[dict]]:
    return build_dashboard(Path(baitiao_path), Path(jintiao_path))


def render_consumer_asset_panel(baitiao_path: str | Path, jintiao_path: str | Path) -> str:
    data, qc = compute_data(baitiao_path, jintiao_path)
    warnings = [item for item in qc if item["level"] == "WARN"]
    if warnings:
        raise PanelInputError("消金资产 QC 存在告警：{}".format(warnings))
    return render_body(data)


def main() -> None:
    parser = argparse.ArgumentParser(description="生成消金资产面板独立预览")
    parser.add_argument("--baitiao-xlsx", required=True)
    parser.add_argument("--jintiao-xlsx", required=True)
    parser.add_argument("--output-html")
    parser.add_argument("--output-json")
    parser.add_argument("--qc-json")
    parser.add_argument("--strict-qc", action="store_true")
    args = parser.parse_args()
    data, qc = build_dashboard(Path(args.baitiao_xlsx), Path(args.jintiao_xlsx))
    tag = max(data["dates"]["baitiao"]["current"], data["dates"]["jintiao"]["current"]).strftime("%Y%m%d")
    DEFAULT_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    html_path = Path(args.output_html) if args.output_html else DEFAULT_OUTPUT_DIR / f"consumer_asset_panel_{tag}.html"
    json_path = Path(args.output_json) if args.output_json else DEFAULT_OUTPUT_DIR / f"consumer_asset_panel_{tag}.json"
    qc_path = Path(args.qc_json) if args.qc_json else DEFAULT_OUTPUT_DIR / f"consumer_asset_panel_{tag}_qc.json"
    for output in (html_path, json_path, qc_path): output.parent.mkdir(parents=True, exist_ok=True)
    html_path.write_text(render_html(data), encoding="utf-8")
    json_path.write_text(json.dumps(data, ensure_ascii=False, indent=2, default=json_default), encoding="utf-8")
    qc_path.write_text(json.dumps(qc, ensure_ascii=False, indent=2, default=json_default), encoding="utf-8")
    warnings = sum(item["level"] == "WARN" for item in qc)
    print(f"[完成] HTML: {html_path}\n[完成] 数据: {json_path}\n[完成] QC: {qc_path}\n[QC] WARN={warnings}")
    if args.strict_qc and warnings:
        raise SystemExit(2)


if __name__ == "__main__":
    main()
