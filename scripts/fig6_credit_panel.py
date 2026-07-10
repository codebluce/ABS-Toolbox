"""fig6_investor_analysis.py — 投资人分析看板(lab实验)

读取 20260630额度盘点.xlsx:
- 非标授信 sheet: A基石投资人 / B非标剩余额度 / C非标授信总额 / D-I为7-12月摊还(待填)
- 非标历史发行 sheet: U认购机构 / V认购份额 / X循环期结束日

算法:
  对非标授信每个机构,从非标历史发行 sheet 找 U列近似匹配 + X列在2026年N月的行,
  V列(认购份额)合计 = 该机构N月摊还到期额度

输出: 投资人分析看板 HTML(风格与综合看板一致)
"""
import os, sys, re, datetime
_SCRIPTS_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, _SCRIPTS_DIR)
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
import html
import pandas as pd
from abs_common import preprocess_xlsx_for_pandas

CREDIT_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                           '..', 'deliverables', 'dashboards', '04_reference',
                           '20260630额度盘点.xlsx')
LEDGER_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                           '..', 'deliverables', 'ledger', '03_final',
                           '2026年ABS发行台账-0703-定稿.xlsx')
OUTPUT = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                      '..', 'deliverables', 'dashboards', '02_history',
                      'lab_viz', 'fig6_investor_analysis.html')


def normalize(name):
    """机构名归一化:去空格/括号统一/去常见后缀"""
    if not name:
        return ''
    n = str(name).strip()
    n = n.replace('（', '(').replace('）', ')')
    n = n.replace(' ', '').replace('　', '')
    # 去常见后缀
    for suffix in ['有限责任公司', '股份有限公司', '股份有限公司', '理财有限责任公司',
                   '理财', '资管', '资产管理']:
        if n.endswith(suffix) and len(n) > len(suffix) + 2:
            n = n[:-len(suffix)]
    return n


def approximate_match(inst_a, inst_b):
    """近似匹配:归一化后相等,或一方包含另一方(len>=3)"""
    na = normalize(inst_a)
    nb = normalize(inst_b)
    if not na or not nb:
        return False
    if na == nb:
        return True
    if len(na) >= 3 and len(nb) >= 3:
        if na in nb or nb in na:
            return True
    return False


def compute_maturity_amounts(ledger_path=None):
    """计算每个非标授信机构 7-12月摊还到期额度 + 实时更新非标剩余额度

    ledger_path: 发行台账路径(综合看板传入),None 时用默认 0703 定稿
    """
    if ledger_path is None:
        ledger_path = LEDGER_FILE
    wb = openpyxl.load_workbook(CREDIT_FILE, read_only=True, data_only=True)

    # 1. 读非标授信 sheet(A-C列)
    ws_credit = wb['非标授信']
    institutions = []
    for r in range(2, ws_credit.max_row + 1):
        name = ws_credit.cell(row=r, column=1).value
        remaining_orig = ws_credit.cell(row=r, column=2).value  # B非标剩余额度(表格值)
        total = ws_credit.cell(row=r, column=3).value            # C非标授信总额
        if name:
            institutions.append({
                'name': str(name).strip(),
                'remaining_orig': float(remaining_orig) if isinstance(remaining_orig, (int, float)) else 0.0,
                'remaining': float(remaining_orig) if isinstance(remaining_orig, (int, float)) else 0.0,  # 实时更新后
                'new_subscription': 0.0,  # 新增认购规模(7月起非标/保登)
                'total': float(total) if isinstance(total, (int, float)) else 0.0,
                'maturities': {7: 0.0, 8: 0.0, 9: 0.0, 10: 0.0, 11: 0.0, 12: 0.0},
            })
    print(f'非标授信机构数: {len(institutions)}')

    # 2. 读非标历史发行 sheet,按月聚合每个机构的认购份额(用于7-12月摊还)
    ws_hist = wb['非标历史发行']
    hist_records = []
    for r in range(2, ws_hist.max_row + 1):
        u = ws_hist.cell(row=r, column=21).value  # U认购机构
        v = ws_hist.cell(row=r, column=22).value  # V认购份额
        x = ws_hist.cell(row=r, column=24).value  # X循环期结束日
        if not u or not isinstance(x, datetime.datetime):
            continue
        if x.year != 2026 or x.month < 7 or x.month > 12:
            continue
        try:
            size = float(v) if isinstance(v, (int, float)) else 0.0
        except (ValueError, TypeError):
            size = 0.0
        hist_records.append((str(u).strip(), x.month, size))
    print(f'2026年7-12月历史发行记录数: {len(hist_records)}')

    # 3. 对每个非标授信机构,近似匹配聚合摊还
    for inst in institutions:
        for hist_name, month, size in hist_records:
            if approximate_match(inst['name'], hist_name):
                inst['maturities'][month] += size

    wb.close()

    # 4. 读最新发行台账,计算新增认购规模(F列含非标/保登 + L列>=2026-07-01 + 同机构V列合计)
    print(f'\n读取台账计算新增认购规模: {ledger_path}')
    tmp = preprocess_xlsx_for_pandas(ledger_path)
    raw = pd.read_excel(tmp, header=None)
    headers = raw.iloc[0].tolist()
    df_ledger = raw.iloc[2:].copy()
    df_ledger.columns = [str(h).strip() if h is not None and str(h).strip() else f'col{i}'
                         for i, h in enumerate(headers)]

    # 类型转换
    df_ledger['发行场所'] = df_ledger['发行场所'].astype(str)
    df_ledger['簿记时间'] = pd.to_datetime(df_ledger['簿记时间'], errors='coerce')
    df_ledger['认购份额'] = pd.to_numeric(df_ledger['认购份额'], errors='coerce')

    # 筛选:F列含"非标"或"保登" + L列>=2026-07-01
    mask = (df_ledger['发行场所'].str.contains('非标|保登', na=False)) & \
           (df_ledger['簿记时间'] >= '2026-07-01') & \
           (df_ledger['认购机构'].notna()) & \
           (df_ledger['认购份额'].notna()) & (df_ledger['认购份额'] > 0)
    df_new = df_ledger[mask].copy()
    print(f'7月起非标/保登新增认购记录数: {len(df_new)}')

    new_sub_records = [(str(u).strip(), float(s))
                       for u, s in zip(df_new['认购机构'], df_new['认购份额'])]

    # 5. 对每个非标授信机构,近似匹配聚合新增认购规模
    for inst in institutions:
        for sub_name, size in new_sub_records:
            if approximate_match(inst['name'], sub_name):
                inst['new_subscription'] += size
        # 实时更新:非标剩余额度 = 表格值 - 新增认购规模
        inst['remaining'] = inst['remaining_orig'] - inst['new_subscription']

    # 打印结果
    print()
    print('=== 摊还到期额度 + 新增认购 + 实时剩余额度 ===')
    print(f'{"机构":12s} {"授信总额":>8s} {"表格剩余":>8s} {"新增认购":>8s} {"实时剩余":>8s} {"7月":>6s} {"8月":>6s} {"9月":>6s} {"10月":>6s} {"11月":>6s} {"12月":>6s}')
    for inst in institutions:
        m = inst['maturities']
        print(f'{inst["name"]:12s} {inst["total"]:8.1f} {inst["remaining_orig"]:8.1f} {inst["new_subscription"]:8.1f} {inst["remaining"]:8.1f} {m[7]:6.1f} {m[8]:6.1f} {m[9]:6.1f} {m[10]:6.1f} {m[11]:6.1f} {m[12]:6.1f}')

    return institutions


CREDIT_CSS = """
/* 非标投资人额度监控 panel 样式 */
.credit-dashboard {
    background: #ffffff;
    border-radius: 8px;
    overflow: hidden;
}
.credit-header {
    background: linear-gradient(135deg, #1a3a5c 0%, #0d1b2e 100%);
    color: white;
    padding: 16px 24px;
}
.credit-title {
    font-size: 18px;
    font-weight: 600;
}
.credit-subtitle {
    font-size: 12px;
    opacity: 0.75;
    margin-top: 4px;
}
.credit-panel {
    padding: 20px 24px;
}
.credit-table {
    width: 100%;
    border-collapse: collapse;
    font-size: 13px;
}
.credit-table thead th {
    background: #f8f9fa;
    color: #495057;
    font-weight: 600;
    padding: 10px 6px;
    text-align: center;
    border-bottom: 2px solid #dee2e6;
    white-space: nowrap;
}
.credit-table thead th:first-child { text-align: left; padding-left: 12px; }
.credit-table tbody td {
    padding: 8px 6px;
    text-align: center;
    border-bottom: 1px solid #e9ecef;
    color: #1d1d1f;
}
.credit-table tbody td.inst-name {
    text-align: left;
    padding-left: 12px;
    font-weight: 500;
    color: #0d1b2e;
}
.credit-table tbody td.num {
    font-variant-numeric: tabular-nums;
    font-weight: 600;
    white-space: nowrap;
}
.credit-table tbody td.num.neg {
    color: #c0392b;
    font-weight: 600;
}
.credit-table tbody tr:hover { background: #f8f9fa; }
.credit-table tbody tr.totals-row {
    background: #e8eef5;
    font-weight: 600;
}
.credit-table tbody tr.totals-row td { border-top: 2px solid #0d1b2e; }
.credit-legend {
    margin-top: 14px;
    padding: 10px 14px;
    background: #f8f9fa;
    border-radius: 6px;
    font-size: 12px;
    color: #6c757d;
    line-height: 1.6;
}
.credit-legend strong { color: #495057; }
"""


def render_credit_panel(ledger_path=None):
    """生成非标投资人额度监控 panel HTML body(供综合看板集成)

    返回: HTML 字符串(不含 <html>/<head>/<body>,仅 panel div)
    """
    institutions = compute_maturity_amounts(ledger_path=ledger_path)

    # 计算合计行
    totals = {'remaining': 0.0, 'total': 0.0, 'new_subscription': 0.0,
              'maturities': {7: 0.0, 8: 0.0, 9: 0.0, 10: 0.0, 11: 0.0, 12: 0.0}}
    for inst in institutions:
        totals['remaining'] += inst['remaining']
        totals['total'] += inst['total']
        totals['new_subscription'] += inst['new_subscription']
        for m in range(7, 13):
            totals['maturities'][m] += inst['maturities'][m]

    # 表格行
    rows_html = ''
    for inst in institutions:
        m = inst['maturities']
        def fmt(v):
            return f'{v:.1f}' if v > 0 else '-'
        remaining_class = 'num neg' if inst['remaining'] < 0 else 'num'
        rows_html += f'''
        <tr>
            <td class="inst-name">{html.escape(inst['name'])}</td>
            <td class="num">{inst['total']:.1f}</td>
            <td class="num">{fmt(inst['new_subscription'])}</td>
            <td class="{remaining_class}">{inst['remaining']:.1f}</td>
            <td class="num">{fmt(m[7])}</td>
            <td class="num">{fmt(m[8])}</td>
            <td class="num">{fmt(m[9])}</td>
            <td class="num">{fmt(m[10])}</td>
            <td class="num">{fmt(m[11])}</td>
            <td class="num">{fmt(m[12])}</td>
        </tr>'''

    # 合计行
    tm = totals['maturities']
    totals_html = f'''
        <tr class="totals-row">
            <td class="inst-name">合计</td>
            <td class="num">{totals['total']:.1f}</td>
            <td class="num">{totals['new_subscription']:.1f}</td>
            <td class="num">{totals['remaining']:.1f}</td>
            <td class="num">{tm[7]:.1f}</td>
            <td class="num">{tm[8]:.1f}</td>
            <td class="num">{tm[9]:.1f}</td>
            <td class="num">{tm[10]:.1f}</td>
            <td class="num">{tm[11]:.1f}</td>
            <td class="num">{tm[12]:.1f}</td>
        </tr>'''

    return f'''
<div class="credit-dashboard">
    <div class="credit-header">
        <div class="credit-title">非标投资人额度监控</div>
        <div class="credit-subtitle">截至 2026-06-30 | 非标授信额度与 7-12月摊还到期额度</div>
    </div>
    <div class="credit-panel">
        <table class="credit-table">
            <thead>
                <tr>
                    <th>基石投资人</th>
                    <th>非标授信总额<br/>(亿元)</th>
                    <th>新增认购规模<br/>(亿元)</th>
                    <th>非标剩余额度<br/>(实时,亿元)</th>
                    <th>7月摊还<br/>(亿元)</th>
                    <th>8月摊还<br/>(亿元)</th>
                    <th>9月摊还<br/>(亿元)</th>
                    <th>10月摊还<br/>(亿元)</th>
                    <th>11月摊还<br/>(亿元)</th>
                    <th>12月摊还<br/>(亿元)</th>
                </tr>
            </thead>
            <tbody>
                {rows_html}
                {totals_html}
            </tbody>
        </table>
        <div class="credit-legend">
            <strong>字段说明:</strong><br/>
            • <strong>非标授信总额</strong>:来自 20260630额度盘点.xlsx「非标授信」sheet C 列<br/>
            • <strong>新增认购规模</strong>:最新发行台账中,F列含"非标"或"保登"关键词,且 L列(簿记时间)≥ 2026-07-01 的同机构 V列(认购份额)合计<br/>
            • <strong>非标剩余额度(实时)</strong>= 表格原始剩余值 - 新增认购规模(红色表示已超额度)<br/>
            • <strong>7-12月摊还</strong>:该机构在「非标历史发行」sheet 中,U列同机构(近似匹配)且 X列(循环期结束日)在 2026 年对应月份的认购份额合计<br/>
            • <strong>近似匹配规则</strong>:机构名归一化(去空格/括号统一/去"理财/资管"等后缀)后相等,或一方包含另一方(len≥3)<br/>
            • <strong>"-"</strong>:该机构当月无到期项目或无新增认购
        </div>
    </div>
</div>'''


def main():
    """独立运行模式:生成完整 HTML 文件"""
    print('=== 非标投资人额度监控看板生成 ===')
    body = render_credit_panel()

    os.makedirs(os.path.dirname(OUTPUT), exist_ok=True)
    full_html = f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<title>非标投资人额度监控 - ABS工具箱</title>
<style>
body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', 'PingFang SC', 'Microsoft YaHei', sans-serif; background: #f5f5f7; padding: 24px; }}
{CREDIT_CSS}
</style>
</head>
<body>
{body}
</body>
</html>'''
    with open(OUTPUT, 'w', encoding='utf-8') as f:
        f.write(full_html)
    print(f'\n已保存: {OUTPUT}')


if __name__ == '__main__':
    main()
